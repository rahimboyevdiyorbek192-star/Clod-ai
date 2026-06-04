# tg_scrapers.py
import re
import os
import asyncio
import random
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telethon.tl.functions.users import GetFullUserRequest
from telethon.tl.functions.contacts import ImportContactsRequest, DeleteContactsRequest
from telethon.tl.types import InputPhoneContact
from telethon.tl.functions.channels import GetFullChannelRequest, JoinChannelRequest
from telethon.tl.functions.messages import ImportChatInviteRequest
from telethon.errors import FloodWaitError, ChannelPrivateError
from datetime import datetime
import aiosqlite
import database as db_mod
import music_scanner as music_mod

SCANNER_PAUSED    = False
MONITORING_PAUSED = False
_SCAN_COUNT = 0
_WATCH_ALERTS = None  # asyncio.Queue() — main() da ishga tushiriladi

async def _check_batch_alerts(batch: list):
    """Keshga qo'shilgan xabarlar uchun alertlarni tekshirish (fon taskda)."""
    try:
        for msg_id, source, sender_id, sender_name, sender_un, text, msg_date in batch:
            if not text:
                continue
            hits = await check_message_alerts(msg_id, source, sender_id, sender_name, text, msg_date)
            if hits and _WATCH_ALERTS:
                for hit in hits:
                    await _WATCH_ALERTS.put(('alert', hit))
    except Exception:
        pass

# Adaptiv flood tracker: flood kelsa avtomatik sekinlashadi
_FLOOD_PENALTY = 0.0   # qo'shimcha uyqu (soniyalarda), flood kelsa oshadi

def _record_flood(seconds: float):
    """Flood kelganda penalty oshirish — keyingi so'rovlar sekinlashadi."""
    global _FLOOD_PENALTY
    _FLOOD_PENALTY = min(_FLOOD_PENALTY + seconds * 0.1, 30.0)

def _decay_flood():
    """Har muvaffaqiyatli so'rovda penalty ozayadi."""
    global _FLOOD_PENALTY
    if _FLOOD_PENALTY > 0:
        _FLOOD_PENALTY = max(0.0, _FLOOD_PENALTY - 0.05)

async def _safe_api_call(make_coro):
    """API chaqiruvni flood himoyasi bilan bajaradi.
    make_coro: callable, har safar yangi coroutine qaytaradi (lambda yoki partial).
    Flood wait semaphore TASHQARISIDA uxlaydi — boshqa tasklar bloklanmaydi.
    """
    for attempt in range(3):
        _decay_flood()
        extra = _FLOOD_PENALTY
        if extra > 0:
            await asyncio.sleep(min(extra, 10))
        try:
            return await asyncio.wait_for(make_coro(), timeout=30)
        except asyncio.TimeoutError:
            return None
        except FloodWaitError as e:
            _record_flood(e.seconds)
            log_flood("api_call", e.seconds)
            await asyncio.sleep(min(e.seconds + 2, 60))
            if attempt == 2:
                return None
        except Exception:
            return None
    return None

# ─── AQLLI RESURS MENEJERI ───────────────────────────────────────────
# Bot o'zi qaysi jarayon og'ir ekanini biladi va resurslarni taqsimlaydi
# Kanal musiqa skanerlash tugadimi?
_CHANNEL_MUSIC_DONE = False

# Flood statistikasi
_FLOOD_STATS = {}

def log_flood(func_name, seconds):
    """Flood statistikasini saqlash"""
    if func_name not in _FLOOD_STATS:
        _FLOOD_STATS[func_name] = {'count': 0, 'total_secs': 0, 'max_secs': 0}
    _FLOOD_STATS[func_name]['count'] += 1
    _FLOOD_STATS[func_name]['total_secs'] += seconds
    _FLOOD_STATS[func_name]['max_secs'] = max(_FLOOD_STATS[func_name]['max_secs'], seconds)
    
    # flood_log.txt ga yozish
    from datetime import datetime as _dt
    line = f"{_dt.now().strftime('%H:%M:%S')} | {func_name} | {seconds}s\n"
    try:
        import os as _os
        log_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'flood_log.txt')
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(line)
    except Exception:
        pass
    print(f"[FLOOD] {func_name}: {seconds}s")

_RESOURCE = {
    'heavy_scan':    False,   # Og'ir skanerlash (guruh, xabar, comment)
    'music_paused':  False,   # Musiqa tracker pauza
    'profile_slow':  False,   # Profil tracker sekin rejim
    'current_task':  None,    # Hozir nima ishlayapti
    'task_start':    None,    # Qachon boshlangan
}

def resource_start(task_name: str):
    """Og'ir jarayon boshlananda chaqiriladi."""
    _RESOURCE['heavy_scan']   = True
    _RESOURCE['music_paused'] = True
    _RESOURCE['profile_slow'] = True
    _RESOURCE['current_task'] = task_name
    _RESOURCE['task_start']   = datetime.now()
    print(f"[RESURS] {task_name} boshlandi — fon jarayonlar sekinlashtirildi")

def resource_stop(task_name: str):
    """Og'ir jarayon tugaganda chaqiriladi."""
    _RESOURCE['heavy_scan']   = False
    _RESOURCE['music_paused'] = False
    _RESOURCE['profile_slow'] = False
    _RESOURCE['current_task'] = None
    _RESOURCE['task_start']   = None
    print(f"[RESURS] {task_name} tugadi — fon jarayonlar davom ettirildi")

def resource_status() -> str:
    """Hozirgi resurs holati."""
    if _RESOURCE['heavy_scan']:
        started = _RESOURCE['task_start']
        elapsed = ""
        if started:
            secs = int((datetime.now() - started).total_seconds())
            elapsed = f" ({secs//60} daq {secs%60} sek)"
        return f"🔴 Og'ir jarayon: {_RESOURCE['current_task']}{elapsed}"
    return "🟢 Normal rejim"


# ─────────────────────────────────────────────────────────────────────
# YORDAMCHI FUNKSIYALAR
# ─────────────────────────────────────────────────────────────────────

def extract_exact_birth_date(text):
    if not text:
        return ""
    m = re.search(r'\b(\d{1,2}[.\/\-]\d{1,2}[.\/\-](?:19|20)\d{2})\b', text)
    if m:
        return m.group(1)
    years = re.findall(r'\b(19\d{2}|20[0-2]\d)\b', text)
    return years[0] if years else ""


async def get_user_by_phone(userbot, phone: str):
    """
    Telefon raqam orqali foydalanuvchi ma'lumotlarini oladi.
    FloodWait kelsa — kutadi va qayta urinadi.
    Qaytaradi: (user, full_info) yoki (None, None)
    """
    imported_user_id = None
    try:
        # 1. Kontaktga saqlash — FloodWait bo'lsa kutib qayta urinish
        for attempt in range(3):
            try:
                result = await userbot(ImportContactsRequest([
                    InputPhoneContact(
                        client_id=0,
                        phone=f"+{phone}",
                        first_name="TempContact",
                        last_name=""
                    )
                ]))
                break
            except FloodWaitError as e:
                log_flood("get_user_by_phone", e.seconds)
                await asyncio.sleep(e.seconds + 5)
                if attempt == 2:
                    return None, None
            except Exception as e:
                print(f"get_user_by_phone xatosi ({phone}): {e}")
                return None, None

        if not result.users:
            return None, None

        user = result.users[0]
        imported_user_id = user.id

        # Kontaktlar orasida kichik pauza
        await asyncio.sleep(random.uniform(1.5, 3.0))

        # 2. To'liq profil ma'lumotlarini olish
        try:
            full_info = await userbot(GetFullUserRequest(user.id))
        except FloodWaitError as e:
            log_flood("get_user_by_phone", e.seconds)
            await asyncio.sleep(min(e.seconds + 5, 300))
            try:
                full_info = await userbot(GetFullUserRequest(user.id))
            except Exception:
                full_info = None
        except Exception:
            full_info = None

        return user, full_info

    except Exception as e:
        print(f"get_user_by_phone xatosi ({phone}): {e}")
        return None, None
    finally:
        # 3. Kontaktni o'chirish (har doim)
        if imported_user_id:
            try:
                await userbot(DeleteContactsRequest(id=[imported_user_id]))
                await asyncio.sleep(1)
            except Exception:
                pass



def extract_bio_links(bio_text):
    """Bio dagi barcha Telegram havolalarini topadi."""
    if not bio_text:
        return []
    results = []
    seen = set()
    pattern = r'(?:https?://)?(?:t\.me|telegram\.me)(/[^\s\)\]>\"\']+)'
    for m in re.finditer(pattern, bio_text):
        path = m.group(1)
        if not path or path == '/':
            continue
        full_url = "https://t.me" + path
        if full_url in seen:
            continue
        seen.add(full_url)
        first = path.strip('/').split('/')[0]
        if first.startswith('+') or first in ('joinchat', 'addlist', 'c'):
            results.append(full_url)
        elif re.match(r'^[a-zA-Z0-9_]{3,}$', first):
            results.append("@" + first)
        else:
            results.append(full_url)
    return results


def extract_invite_links(bio_text):
    """
    Bio dan FAQAT maxfiy kanal invite linklarini ajratadi.
    t.me/+XXXX yoki t.me/joinchat/XXXX formatlar.
    """
    if not bio_text:
        return []
    results = []
    pattern = r'(?:https?://)?(?:t\.me|telegram\.me)(/(?:\+|joinchat/)[^\s\)\]>\"\']+)'
    for m in re.finditer(pattern, bio_text):
        path = m.group(1)
        full_url = "https://t.me" + path
        if full_url not in results:
            results.append(full_url)
    return results


async def safe_get_entity(userbot, target):
    """get_entity ni FloodWait bilan xavfsiz chaqirish."""
    from telethon.errors import FloodWaitError
    for attempt in range(3):
        try:
            return await userbot.get_entity(target)
        except FloodWaitError as e:
            log_flood("safe_get_entity", e.seconds)
            # Uzoq flood bo'lsa — o'tkazib yuborish (kutmaslik)
            if e.seconds > 120:
                return None
            await asyncio.sleep(e.seconds + 2)
        except Exception:
            return None
    return None


async def send_join_request(userbot, invite_link):
    """
    t.me/+XXXX yoki t.me/joinchat/XXXX invite link orqali
    kanalga qo'shilish so'rovnomasi yuboradi.
    """
    try:
        if "/+" in invite_link:
            hash_part = invite_link.split("/+")[-1].rstrip("/")
        elif "joinchat/" in invite_link:
            hash_part = invite_link.split("joinchat/")[-1].rstrip("/")
        else:
            return False
        await userbot(ImportChatInviteRequest(hash=hash_part))
        return True
    except Exception as e:
        err = str(e).lower()
        if "already" in err or "request" in err:
            return True  # Avval yuborilgan — normal
        print(f"send_join_request xatosi ({invite_link}): {e}")
        return False


def validate_target(text: str) -> tuple[bool, str]:
    """
    Guruh/kanal linkini tekshiradi.
    Qaytaradi: (is_valid, cleaned_text)
    """
    if not text or len(text) > 512:
        return False, ""
    text = text.strip()
    # Ruxsat etilgan formatlar: @username, https://t.me/..., raqam
    import re
    if re.match(r'^@[a-zA-Z0-9_]{3,}$', text):
        return True, text
    if re.match(r'^https?://t\.me/', text):
        return True, text
    if re.match(r'^-?\d+$', text):
        return True, text
    if re.match(r'^[a-zA-Z0-9_]{3,}$', text):
        return True, text
    return False, text


def validate_keyword(text: str) -> tuple[bool, str]:
    """Kalit so'zni tekshiradi. Max 200 belgi."""
    if not text or len(text.strip()) == 0:
        return False, ""
    text = text.strip()
    if len(text) > 200:
        return False, text
    return True, text


async def resolve_personal_channel(userbot, ch_id):
    """
    Shaxsiy kanal linkini hal qiladi.
    Numeric ID ni resolved_channel_ids jadvaliga saqlaydi (keshlayd).
    Qaytaradi: (link, is_private)
    """
    try:
        ch_entity = await userbot.get_entity(ch_id)

        # Numeric ID ni kesh jadvaliga saqlash — a'zo bo'lmasdan ham olish mumkin
        if hasattr(ch_entity, 'id') and ch_entity.id:
            _eid = str(ch_entity.id).lstrip('-')
            _num_id = f"-100{_eid}" if not str(ch_entity.id).startswith('-100') else str(ch_entity.id)
            _link_key = str(ch_id)
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            try:
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                    await _db.execute(
                        "INSERT OR REPLACE INTO resolved_channel_ids "
                        "(channel_link, numeric_id, resolved_at) VALUES (?, ?, ?)",
                        (_link_key, _num_id, now_str)
                    )
                    await _db.commit()
            except Exception:
                pass

        ch_uname  = getattr(ch_entity, 'username', None)
        if ch_uname:
            return f"https://t.me/{ch_uname}", False
        try:
            full_ch = await userbot(GetFullChannelRequest(ch_entity))
            inv = getattr(full_ch.full_chat, 'exported_invite', None)
            if inv and getattr(inv, 'link', None):
                return inv.link, False
        except Exception:
            pass
        ch_title = getattr(ch_entity, 'title', '')
        return ch_title or str(ch_id), False
    except ChannelPrivateError:
        return f"🔒 Maxfiy (ID:{ch_id})", True
    except Exception:
        return "", False


def apply_excel_styles(ws, total_rows):
    """Excel faylni chiroyli formatlaydi."""
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin")
    )
    even_fill = PatternFill("solid", fgColor="DEEAF1")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    # Sarlavha formatlash
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # Ma'lumot qatorlari — faqat 1000 qatorgacha per-cell styling
    style_limit = min(total_rows + 1, 1001)
    for row_num in range(2, style_limit + 1):
        idx      = row_num - 1
        row_fill = even_fill if idx % 2 == 0 else odd_fill
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = left_align if col_num >= 2 else center_align

    # Ustun kengliklari
    for col in ws.columns:
        max_w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22


def apply_excel_styles_light(ws):
    """Katta fayllarda faqat sarlavha + ustun kengligi + freeze — per-cell styling yo'q."""
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin")
    )
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    # Ustun kengliklari (faqat birinchi 3 qator asosida tez hisob)
    for col in ws.columns:
        sample_vals = [str(c.value or '') for c in list(col)[:50]]
        max_w = max((len(v) for v in sample_vals), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 55)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────
# GURUH SKANERI — bot.py dagi ub_members_process asosida
# ─────────────────────────────────────────────────────────────────────

async def deep_scan_group(userbot, target_group, output_path, status_msg,
                           resume_offset=0, resume_count=0, scan_id=None,
                           pre_entity=None):
    """
    Guruh a'zolarini skanerlab Excel ga yozadi.
    Bio dagi t.me/+XXXX linklar — alohida "Maxfiy Kanal" ustuniga.
    Maxfiy kanallarga so'rovnoma yuboriladi va DB ga saqlanadi.
    """
    global MONITORING_PAUSED, _SCAN_COUNT, _FLOOD_PENALTY
    _SCAN_COUNT += 1
    MONITORING_PAUSED = True
    _FLOOD_PENALTY = 0.0  # yangi skan — eski flood penaltyni nolga tushirish
    resource_start("Ochiq Guruh Skanerlash")

    if scan_id is None:
        sender_id = getattr(status_msg, 'chat_id', 0)
        scan_id = await db_mod.create_scan_session(str(target_group), output_path, sender_id)

    if resume_offset > 0 and os.path.exists(output_path):
        try:
            _loop = asyncio.get_event_loop()
            wb = await asyncio.wait_for(
                _loop.run_in_executor(None, openpyxl.load_workbook, output_path),
                timeout=30
            )
            ws = wb.active
        except Exception:
            resume_offset = 0
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Guruh Skaner"
            ws.append([
                "№", "Ism", "Familiya", "Username", "Telegram ID",
                "Telefon", "Bot?", "Premium?", "Bio",
                "Shaxsiy Kanal", "Maxfiy Kanal", "Ochiq Kanal", "Profil havolasi"
            ])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guruh Skaner"
        ws.append([
            "№", "Ism", "Familiya", "Username", "Telegram ID",
            "Telefon", "Bot?", "Premium?", "Bio",
            "Shaxsiy Kanal",    # personal_channel_id dan olingan
            "Maxfiy Kanal",     # Bio dagi t.me/+XXXX linklar
            "Ochiq Kanal",      # Bio dagi @username linklar
            "Profil havolasi"
        ])

    count   = resume_count
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    SAVE_EVERY = 50

    try:
        # pre_entity berilgan bo'lsa — qayta resolve qilinmaydi (ResolveUsernameRequest tejaladi)
        entity = pre_entity
        if entity is None:
            entity = await safe_get_entity(userbot, target_group)
        if entity is None:
            raise Exception("Guruh topilmadi yoki Telegram cheklovi. Biroz kutib qayta urining.")

        # --- 1-bosqich: a'zolar ro'yxati (aggressive=True) ---
        try:
            await status_msg.edit("📋 **1-bosqich:** A'zolar ro'yxati yuklanmoqda...")
        except Exception:
            pass

        seen_ids = set()
        participants = []
        for _p_attempt in range(3):
            try:
                async for _u in userbot.iter_participants(entity, aggressive=True):
                    if _u.id not in seen_ids:
                        seen_ids.add(_u.id)
                        participants.append(_u)
                break  # muvaffaqiyatli tugadi
            except FloodWaitError as e:
                log_flood("iter_participants", e.seconds)
                _record_flood(e.seconds)
                await asyncio.sleep(min(e.seconds + 5, 300))
            except Exception:
                break

        phase1_count = len(participants)

        # --- 2-bosqich: xabar tarixi orqali qo'shimcha userlar ---
        try:
            await status_msg.edit(
                f"📨 **2-bosqich:** Xabarlar tarixidan qo'shimcha userlar izlanmoqda...\n"
                f"(1-bosqichda: {phase1_count} ta topildi)"
            )
        except Exception:
            pass

        _msg2_count = 0
        _cache_batch = []
        _src_str = str(target_group)
        _iter_offset_id = 0
        _iter_done = False
        while not _iter_done:
            try:
                async for msg in userbot.iter_messages(
                    entity, limit=None,
                    offset_id=_iter_offset_id, reverse=False
                ):
                    if not msg.sender_id or msg.sender_id <= 0:
                        _iter_offset_id = msg.id
                        continue
                    _msg2_count += 1
                    _iter_offset_id = msg.id

                    if msg.sender_id not in seen_ids:
                        sender = msg.sender
                        if sender and not getattr(sender, 'bot', False):
                            seen_ids.add(sender.id)
                            participants.append(sender)

                            # Topilgan zahoti to'liq ma'lumot olish va yozish
                            _fn  = sender.first_name or ""
                            _ln  = sender.last_name  or ""
                            _un  = ("@" + sender.username) if sender.username else ""
                            _ph  = sender.phone or ""
                            _bio = ""
                            _shaxsiy = ""
                            _maxfiy  = ""
                            _ochiq   = ""
                            try:
                                await asyncio.sleep(0.7)
                                fi = await asyncio.wait_for(
                                    userbot(GetFullUserRequest(sender.id)), timeout=20
                                )
                                fu   = fi.full_user
                                _bio = fu.about or ""
                                inv  = extract_invite_links(_bio)
                                if inv:
                                    _maxfiy = ", ".join(inv)
                                al  = extract_bio_links(_bio)
                                oc  = [l for l in al if l.startswith('@') or
                                       ('t.me/' in l and '/+' not in l and 'joinchat' not in l)]
                                _ochiq = ", ".join(oc) if oc else ""
                                pc  = getattr(fu, 'personal_channel_id', None)
                                if pc:
                                    _shaxsiy = f"tg://resolve?domain=c{pc}"
                                if inv:
                                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
                                        for lnk in inv:
                                            await _db.execute(
                                                "INSERT OR IGNORE INTO hidden_channel_knocker "
                                                "(channel_id, creator_id, source_group) VALUES (?,?,?)",
                                                (lnk, sender.id, str(target_group))
                                            )
                                        await _db.commit()
                            except FloodWaitError as e:
                                _record_flood(e.seconds)
                                log_flood("phase2_full_user", e.seconds)
                                await asyncio.sleep(min(e.seconds + 2, 120))
                            except Exception:
                                pass

                            if _fn or _ln or _un or _bio:
                                _purl  = (f"https://t.me/{sender.username}" if sender.username
                                          else f"tg://user?id={sender.id}")
                                _bdate = extract_exact_birth_date(_bio)
                                count += 1
                                ws.append([
                                    count, _fn, _ln, _un, sender.id,
                                    ("+" + _ph) if _ph else "",
                                    "❌",
                                    "✅" if getattr(sender, "premium", False) else "❌",
                                    _bio, _shaxsiy, _maxfiy, _ochiq, _purl
                                ])
                                _has_db = _shaxsiy if _shaxsiy else (_maxfiy if _maxfiy else "❌")
                                await db_mod.save_user_to_bank(
                                    sender.id, str(target_group), _fn, _ln, _un,
                                    _ph, _bdate, _bio,
                                    ", ".join(extract_bio_links(_bio)) or "Yo'q",
                                    _has_db
                                )

                    # Matnli xabarlarni keshga yig'ish
                    if msg.text and len(msg.text) > 2:
                        sender = msg.sender
                        s_id = getattr(sender, 'id', msg.sender_id) if sender else msg.sender_id
                        s_name = ""
                        s_un   = ""
                        if sender and hasattr(sender, 'first_name'):
                            s_name = ((sender.first_name or "") + " " + (sender.last_name or "")).strip()
                            s_un   = getattr(sender, 'username', '') or ""
                        msg_dt = msg.date.strftime("%Y-%m-%d %H:%M") if msg.date else ""
                        _cache_batch.append((msg.id, _src_str, s_id, s_name, s_un, msg.text[:500], msg_dt))

                    # Har 1000 xabarda batch-insert
                    if len(_cache_batch) >= 1000:
                        try:
                            async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                                await _db.executemany(
                                    "INSERT OR IGNORE INTO messages_cache "
                                    "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                                    "VALUES (?,?,?,?,?,?,?)",
                                    _cache_batch
                                )
                                await _db.commit()
                        except Exception:
                            pass
                        _cache_batch = []

                    if _msg2_count % 2000 == 0:
                        try:
                            await status_msg.edit(
                                f"📨 **2-bosqich:** `{_msg2_count}` xabar ko'rildi\n"
                                f"👥 Yangi topilgan: `{len(participants) - phase1_count}` ta..."
                            )
                        except Exception:
                            pass
                _iter_done = True  # barcha xabarlar muvaffaqiyatli o'qildi
            except FloodWaitError as e:
                log_flood("iter_messages_scan", e.seconds)
                await asyncio.sleep(min(e.seconds + 5, 300))
                # flood dan keyin davom etamiz (while loop qayta ishlaydi)
            except Exception as _msg_err:
                print(f"[deep_scan] iter_messages xatosi: {_msg_err}")
                _iter_done = True  # boshqa xatoda to'xtatamiz

        # Qolgan kesh batchni saqlash
        if _cache_batch:
            try:
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                    await _db.executemany(
                        "INSERT OR IGNORE INTO messages_cache "
                        "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                        "VALUES (?,?,?,?,?,?,?)",
                        _cache_batch
                    )
                    await _db.commit()
                asyncio.create_task(_check_batch_alerts(_cache_batch))
            except Exception:
                pass

        total = len(participants)
        _cached_msg_count = _msg2_count  # nechta xabar keshga yig'ildi

        try:
            await status_msg.edit(
                f"📨 **2-bosqich tugadi:** `{_cached_msg_count}` xabar keshga saqlandi\n"
                f"👥 Jami **{total}** ta unikal foydalanuvchi topildi\n"
                f"(a'zolar: {phase1_count} + xabar tarixi: {total - phase1_count})\n"
                f"🔍 Profillar tahlil qilinmoqda..."
            )
        except Exception:
            pass

        # Ketma-ket oddiy loop — faqat Phase 1 a'zolari (Phase 2 da yozildi)
        work_list = [u for i, u in enumerate(participants[:phase1_count]) if i >= resume_offset]
        for _wi, user in enumerate(work_list):
            while SCANNER_PAUSED:
                await asyncio.sleep(1)

            uid = user.id
            bio = ""
            shaxsiy = ""
            maxfiy = ""
            ochiq = ""

            try:
                await asyncio.sleep(0.7)   # flood oldini olish — 30 tez+20s to'xtash o'rniga silliq
                fi = await asyncio.wait_for(
                    userbot(GetFullUserRequest(uid)), timeout=20
                )
                fu  = fi.full_user
                bio = fu.about or ""
                inv = extract_invite_links(bio)
                if inv:
                    maxfiy = ", ".join(inv)
                al = extract_bio_links(bio)
                oc = [l for l in al if l.startswith('@') or
                      ('t.me/' in l and '/+' not in l and 'joinchat' not in l)]
                ochiq = ", ".join(oc) if oc else ""
                pc = getattr(fu, 'personal_channel_id', None)
                if pc:
                    shaxsiy = f"tg://resolve?domain=c{pc}"
                if inv:
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
                        for lnk in inv:
                            await _db.execute(
                                "INSERT OR IGNORE INTO hidden_channel_knocker "
                                "(channel_id, creator_id, source_group) VALUES (?,?,?)",
                                (lnk, uid, str(target_group))
                            )
                        await _db.commit()
            except FloodWaitError as e:
                _record_flood(e.seconds)
                log_flood("deep_scan_group", e.seconds)
                await asyncio.sleep(min(e.seconds + 2, 120))
            except Exception:
                pass

            first_name = user.first_name or ""
            last_name  = user.last_name  or ""
            uname      = ("@" + user.username) if user.username else ""
            phone      = user.phone or ""
            is_bot     = "✅" if user.bot else "❌"
            is_premium = "✅" if getattr(user, "premium", False) else "❌"

            if not first_name and not last_name and not uname and not bio:
                continue

            profile_url = (f"https://t.me/{user.username}" if user.username
                           else f"tg://user?id={uid}")
            b_date = extract_exact_birth_date(bio)

            count += 1
            ws.append([
                count, first_name, last_name, uname, uid,
                ("+" + phone) if phone else "",
                is_bot, is_premium, bio,
                shaxsiy, maxfiy, ochiq, profile_url
            ])

            has_db = shaxsiy if shaxsiy else (maxfiy if maxfiy else "❌")
            await db_mod.save_user_to_bank(
                uid, str(target_group), first_name, last_name, uname,
                phone, b_date, bio,
                ", ".join(extract_bio_links(bio)) or "Yo'q",
                has_db
            )

            # Har SAVE_EVERY da Excel ga yozish + progress saqlash
            if count > 0 and count % SAVE_EVERY == 0:
                _sv_loop = asyncio.get_event_loop()
                await _sv_loop.run_in_executor(None, wb.save, output_path)
                await db_mod.update_scan_progress(scan_id, resume_offset + _wi + 1, count)

            # Status yangilash — har 100 ta profilda
            if count % 100 == 0:
                try:
                    await status_msg.edit(
                        f"🔍 **Skanerlamoqda:** `{count}/{total}` ta profil..."
                    )
                except Exception:
                    pass

    except Exception as e:
        await db_mod.finish_scan_session(scan_id, status='error')
        raise
    else:
        await db_mod.finish_scan_session(scan_id, status='done')
    finally:
        # Eng oxirgi saqlash — har doim bajariladi
        try:
            _fl = asyncio.get_event_loop()
            await _fl.run_in_executor(None, apply_excel_styles, ws, count)
            await _fl.run_in_executor(None, wb.save, output_path)
        except Exception:
            pass
        _SCAN_COUNT -= 1
        if _SCAN_COUNT <= 0:
            _SCAN_COUNT = 0
            MONITORING_PAUSED = False
        resource_stop("Ochiq Guruh Skanerlash")

    return count


# ─────────────────────────────────────────────────────────────────────
# RESUME
# ─────────────────────────────────────────────────────────────────────

async def resume_pending_scans(userbot, bot, admin_id):
    pending = await db_mod.get_pending_scans()
    if not pending:
        return
    for scan_id, target_group, output_path, last_offset, total_count, sender_id in pending:
        # scan type ni aniqlash
        scan_type = 'group'  # default
        try:
            notify_id  = sender_id or admin_id
            status_msg = await bot.send_message(
                notify_id,
                f"♻️ **Tugallanmagan skanerlash davom ettirilmoqda!**\n"
                f"🏢 Guruh: `{target_group}`\n"
                f"📊 Oldindan: `{total_count}` ta\n"
                f"⏩ `{last_offset}` pozitsiyadan davom..."
            )
            asyncio.create_task(
                _resume_scan_task(userbot, bot, notify_id, target_group,
                                   output_path, last_offset, total_count,
                                   scan_id, status_msg)
            )
        except Exception as e:
            print(f"Resume xatosi: {e}")
            await db_mod.finish_scan_session(scan_id, status='error')


async def _resume_scan_task(userbot, bot, sender_id, target_group,
                             output_path, last_offset, total_count,
                             scan_id, status_msg):
    try:
        count = await deep_scan_group(
            userbot, target_group, output_path, status_msg,
            resume_offset=last_offset, resume_count=total_count, scan_id=scan_id
        )
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT INTO archive_bin (file_name, file_path, created_date) VALUES (?, ?, ?)",
                (os.path.basename(output_path), output_path,
                 datetime.now().strftime("%Y-%m-%d"))
            )
            await db.commit()
        await bot.send_file(
            sender_id, output_path,
            caption=f"✅ Resume yakunlandi! Jami `{count}` ta profil."
        )
    except Exception as e:
        await bot.send_message(sender_id, f"❌ Resume xatolik: {e}")
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────
# 24 SOATLIK MAXFIY KANAL MONITOR
# Faqat bio dagi t.me/+XXXX invite linklariga so'rovnoma yuboradi
# ─────────────────────────────────────────────────────────────────────

# Kunlik so'rovnoma hisoblagich
_daily_join_count = 0
_daily_join_date  = ""
MAX_DAILY_JOINS   = 288  # Kuniga maksimal so'rovnomalar soni


async def hidden_channel_24h_knocker(userbot, bot, admin_id):
    """
    Har 5 daqiqada pending invite linklarni tekshiradi.
    Userbot dialog ro'yxatida kanal paydo bo'lsa — kirish ochildi.
    24 soatda bir marta yana so'rovnoma yuboradi.
    Kuniga maksimal 25 ta so'rovnoma — flood himoyasi.
    """
    global _daily_join_count, _daily_join_date

    while True:
        if MONITORING_PAUSED:
            await asyncio.sleep(10)
            continue
        try:
            current_time = datetime.now()
            today_str    = current_time.strftime("%Y-%m-%d")

            # Yangi kun bo'lsa hisoblagichni nolga tushirish
            if _daily_join_date != today_str:
                _daily_join_date  = today_str
                _daily_join_count = 0

            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                async with db.execute(
                    "SELECT channel_id, creator_id, source_group, last_request_time "
                    "FROM hidden_channel_knocker WHERE status='pending'"
                ) as cur:
                    tasks = await cur.fetchall()

            if not tasks:
                await asyncio.sleep(300)
                continue

            # Dialog ro'yxatini bir marta yuklab olish
            dialog_map = {}
            try:
                async for dlg in userbot.iter_dialogs():
                    d_id  = abs(dlg.id)
                    d_str = str(d_id)
                    norm  = d_str[3:] if d_str.startswith('100') and len(d_str) > 10 else d_str
                    dialog_map[d_str] = dlg
                    dialog_map[norm]  = dlg
                    if hasattr(dlg, 'entity'):
                        ent = dlg.entity
                        if getattr(ent, 'username', None):
                            dialog_map[f"https://t.me/{ent.username}"] = dlg
            except Exception as e:
                print(f"iter_dialogs xatosi: {e}")

            for ch_id_str, creator_id, source_group, last_req_str in tasks:
                if MONITORING_PAUSED:
                    break

                is_invite = "t.me/+" in ch_id_str or "t.me/joinchat/" in ch_id_str

                # 1. Dialog ro'yxatida borligini tekshirish
                found = dialog_map.get(ch_id_str)
                if not found and not is_invite:
                    raw  = ch_id_str.lstrip('-')
                    norm = raw[3:] if raw.startswith('100') and len(raw) > 10 else raw
                    found = dialog_map.get(norm) or dialog_map.get(raw)

                # 2. Dialog da topilmasa — to'g'ridan get_entity urinib ko'rish
                if not found and is_invite:
                    try:
                        from telethon.tl.functions.messages import CheckChatInviteRequest
                        hash_part = ch_id_str.split("/+")[-1] if "/+" in ch_id_str else ch_id_str.split("joinchat/")[-1]
                        hash_part = hash_part.rstrip("/")
                        invite_info = await userbot(CheckChatInviteRequest(hash=hash_part))
                        # Agar already_member bo'lsa — kirish ochildi
                        if hasattr(invite_info, 'chat'):
                            ch_entity = invite_info.chat
                            # Dialog_map ga qo'shish
                            class FakeDlg:
                                entity = ch_entity
                            found = FakeDlg()
                    except Exception as e:
                        err = str(e).lower()
                        if "already" in err or "member" in err:
                            # A'zo bo'lib qolgan — entity ni olish
                            try:
                                ch_entity = await userbot.get_entity(ch_id_str)
                                class FakeDlg2:
                                    entity = ch_entity
                                found = FakeDlg2()
                            except Exception:
                                pass

                if found:
                    # ✅ Kirish ochildi
                    ent     = found.entity
                    ch_link = ch_id_str  # invite link o'zini saqlaymiz
                    if getattr(ent, 'username', None):
                        ch_link = f"https://t.me/{ent.username}"
                    else:
                        try:
                            full = await userbot(GetFullChannelRequest(ent))
                            inv  = getattr(full.full_chat, 'exported_invite', None)
                            if inv and getattr(inv, 'link', None):
                                ch_link = inv.link
                        except Exception:
                            pass

                    await bot.send_message(
                        admin_id,
                        f"🔓 **MAXFIY KANALGA KIRISH OCHILDI!**\n\n"
                        f"👤 Egasi ID: `{creator_id}`\n"
                        f"🔗 Kanal linki: {ch_link}\n"
                        f"🏢 Manba: `{source_group}`"
                    )
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                        await db.execute(
                            "UPDATE hidden_channel_knocker SET status='joined' WHERE channel_id=?",
                            (ch_id_str,)
                        )
                        if creator_id:
                            await db.execute(
                                "UPDATE users_memory_bank SET has_hidden=? WHERE user_id=?",
                                (ch_link, creator_id)
                            )
                        await db.commit()

                else:
                    # Hali tasdiqlanmadi — 24 soat o'tgan bo'lsa yana so'rovnoma
                    try:
                        last_time = datetime.strptime(last_req_str, "%Y-%m-%d %H:%M")
                    except Exception:
                        last_time = current_time
                    elapsed = (current_time - last_time).total_seconds()

                    if elapsed >= 86400 + random.randint(-1800, 3600):
                        # Kunlik limit tekshirish
                        if _daily_join_count >= MAX_DAILY_JOINS:
                            # Limit to'ldi — ertaga yuboriladi
                            continue

                        if is_invite:
                            sent = await send_join_request(userbot, ch_id_str)
                        else:
                            try:
                                ch_val = int(ch_id_str) if ch_id_str.lstrip('-').isdigit() else ch_id_str
                                await userbot.get_entity(ch_val)
                                sent = True
                            except Exception:
                                sent = False

                        if sent:
                            _daily_join_count += 1
                            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                                await db.execute(
                                    "UPDATE hidden_channel_knocker SET last_request_time=? WHERE channel_id=?",
                                    (current_time.strftime("%Y-%m-%d %H:%M"), ch_id_str)
                                )
                                await db.commit()
                            # So'rovnomalar orasida 5 daqiqa pauza — flood himoyasi
                            await asyncio.sleep(300)

                await asyncio.sleep(random.uniform(3, 8))

        except Exception as e:
            print(f"hidden_channel_24h_knocker: {e}")
        await asyncio.sleep(300)


# ─────────────────────────────────────────────────────────────────────
# BACKGROUND PROFIL TRACKER
# ─────────────────────────────────────────────────────────────────────

async def background_profile_tracker(userbot):
    batch_size = 50

    # Kanal musiqa skanerlash tugaguncha kutish
    global _CHANNEL_MUSIC_DONE
    print("[PROFIL] Kanal musiqalari tugashini kutmoqda...")
    while not _CHANNEL_MUSIC_DONE:
        await asyncio.sleep(30)
    print("[PROFIL] Kanal musiqalari tugadi — profil musiqasi boshlanadi")

    # Bazadan oxirgi offsetni olish
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        await db.execute(
            "CREATE TABLE IF NOT EXISTS tracker_state "
            "(key TEXT PRIMARY KEY, value TEXT)"
        )
        await db.commit()
        async with db.execute(
            "SELECT value FROM tracker_state WHERE key='profile_offset'"
        ) as cur:
            row = await cur.fetchone()
            offset = int(row[0]) if row else 0

    while True:
        if MONITORING_PAUSED:
            await asyncio.sleep(15)
            continue
        try:
            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                async with db.execute(
                    "SELECT DISTINCT user_id FROM users_memory_bank ORDER BY user_id LIMIT ? OFFSET ?",
                    (batch_size, offset)
                ) as cur:
                    users = await cur.fetchall()

            if not users:
                offset = 0
                # Offset nolga tushirildi — bazaga saqlash
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                    await db.execute(
                        "INSERT OR REPLACE INTO tracker_state (key, value) VALUES ('profile_offset', '0')"
                    )
                    await db.commit()
                await asyncio.sleep(1800)
                continue

            now_str    = datetime.now().strftime("%Y-%m-%d %H:%M")
            # GetFullUserRequest: Telegram ~80 req/min limit
            # 2 parallel + 1.5s sleep = ~1.3 req/s = 78 req/min — xavfsiz chegara
            _api_sem   = asyncio.Semaphore(2)
            # i3/i5 uchun: 3 parallel — CPU thrashing oldini oladi
            _dl_sem    = asyncio.Semaphore(3)

            async def _do_one_profile(uid):
                # Rate-limiting uyqu SEMAPHORE TASHQARISIDA — boshqa tasklar bloklanmaydi
                await asyncio.sleep(random.uniform(1.2, 1.8))
                try:
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                        async with db.execute(
                            "SELECT bio, has_hidden FROM users_memory_bank WHERE user_id=? LIMIT 1",
                            (uid,)
                        ) as cur:
                            old_row = await cur.fetchone()
                    old_bio        = (old_row[0] or "") if old_row else ""
                    old_has_hidden = (old_row[1] or "") if old_row else ""

                    # Semaphore FAQAT API chaqiruvi atrofida
                    async with _api_sem:
                        fi = await _safe_api_call(lambda: userbot(GetFullUserRequest(uid)))
                    if fi is None:
                        return
                    bio = fi.full_user.about or ""

                    # Profil musiqasi — barcha 4 field, semaphore tashqarisida
                    try:
                        music_docs = []
                        for field in ['saved_music', 'profile_song', 'profile_songs', 'music']:
                            val = getattr(fi.full_user, field, None)
                            if val is None:
                                continue
                            if isinstance(val, list):
                                music_docs.extend(val)
                            else:
                                music_docs.append(val)

                        async def _proc_doc(idx, doc):
                            if hasattr(doc, 'document'):
                                doc = doc.document
                            if not hasattr(doc, 'id'):
                                return
                            tmp_music = os.path.join(
                                os.path.dirname(os.path.abspath(__file__)),
                                f"tmp_profile_{uid}_{idx}.ogg"
                            )
                            try:
                                async with _dl_sem:
                                    await userbot.download_media(doc, file=tmp_music)
                                if os.path.exists(tmp_music):
                                    fp, duration = await music_mod.get_fingerprint_async(tmp_music)
                                    if fp:
                                        await music_mod.init_music_db()
                                        already = await music_mod.is_profile_music_saved(uid, fp)
                                        if not already:
                                            await music_mod.save_fingerprint(
                                                str(uid), f"Profil: {uid}",
                                                f"profile_{uid}_{idx}", fp, duration or 0
                                            )
                                        hits = await music_mod.check_against_watch_list(fp)
                                        for hit in hits:
                                            _WATCH_ALERTS.put_nowait({
                                                'admin_id':    hit['admin_id'],
                                                'watch_name':  hit['watch_name'],
                                                'score':       hit['score'],
                                                'source_name': f"Profil: {uid}",
                                                'source_id':   str(uid),
                                                'source_type': 'profil'
                                            })
                            except Exception:
                                pass
                            finally:
                                if os.path.exists(tmp_music):
                                    os.remove(tmp_music)

                        if music_docs:
                            await asyncio.gather(
                                *[_proc_doc(i, d) for i, d in enumerate(music_docs)],
                                return_exceptions=True
                            )
                    except Exception:
                        pass

                    # has_hidden hisoblash
                    invite_links = extract_invite_links(bio)
                    has_hidden   = "❌"
                    if invite_links:
                        has_hidden = ", ".join(invite_links)
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                            for inv_link in invite_links:
                                await db.execute(
                                    "INSERT OR IGNORE INTO hidden_channel_knocker "
                                    "(channel_id, creator_id, source_group, last_request_time) "
                                    "VALUES (?, ?, ?, ?)",
                                    (inv_link, uid, "Monitoring", now_str)
                                )
                            await db.commit()
                    elif getattr(fi.full_user, 'personal_channel_id', None):
                        ch_id = fi.full_user.personal_channel_id
                        try:
                            ch_ent  = await userbot.get_entity(ch_id)
                            ch_user = getattr(ch_ent, 'username', None)
                            has_hidden = f"https://t.me/{ch_user}" if ch_user else str(ch_id)
                        except ChannelPrivateError:
                            has_hidden = f"🔒 Maxfiy (ID:{ch_id})"
                        except Exception:
                            pass

                    open_ch = ", ".join(extract_bio_links(bio)) or "Yo'q"

                    bio_changed    = bio.strip() != old_bio.strip()
                    hidden_changed = has_hidden != old_has_hidden

                    if bio_changed or hidden_changed:
                        await db_mod.update_user_changes(uid, bio, open_ch, has_hidden)
                    else:
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                            await db.execute(
                                "UPDATE users_memory_bank SET bio=?, open_channels=? WHERE user_id=?",
                                (bio, open_ch, uid)
                            )
                            await db.commit()

                except Exception:
                    pass
                finally:
                    # Post-processing uyqu SEMAPHORE TASHQARISIDA — to'g'ri tezlik
                    if _RESOURCE['profile_slow']:
                        await asyncio.sleep(random.uniform(3, 6))
                    else:
                        await asyncio.sleep(random.uniform(1.5, 3))

            # 5 ta profil parallel skanerlash
            await asyncio.gather(
                *[asyncio.create_task(_do_one_profile(uid)) for (uid,) in users],
                return_exceptions=True
            )

            offset += batch_size
            # Offsetni bazaga saqlash
            try:
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                    await db.execute(
                        "INSERT OR REPLACE INTO tracker_state (key, value) VALUES ('profile_offset', ?)",
                        (str(offset),)
                    )
                    await db.commit()
            except Exception:
                pass
        except Exception as e:
            print(f"background_profile_tracker: {e}")
        await asyncio.sleep(60)


# ─────────────────────────────────────────────────────────────────────
# XABAR SKANERLASH
# Guruh/kanal xabarlaridan yozgan odamlarni topadi
# A'zo bo'lmasa ham ochiq guruhda ishlaydi
# ─────────────────────────────────────────────────────────────────────

async def scan_messages(userbot, target, output_path, status_msg, days=None,
                        resume_offset=0, resume_count=0, scan_id=None):
    """
    Guruh/kanal xabarlarini o'qib, yozgan foydalanuvchilarni skanerLaydi.
    """
    global _SCAN_COUNT, MONITORING_PAUSED, _FLOOD_PENALTY
    _SCAN_COUNT += 1
    MONITORING_PAUSED = True
    _FLOOD_PENALTY = 0.0
    _RESOURCE['music_paused'] = True
    _RESOURCE['profile_slow'] = True
    _RESOURCE['heavy_scan'] = True
    _RESOURCE['current_task'] = "Yopiq Guruh Skanerlash"
    _RESOURCE['task_start'] = datetime.now()

    if scan_id is None:
        sender_id_tmp = getattr(status_msg, 'sender_id', 0) or getattr(status_msg, 'chat_id', 0)
        scan_id = await db_mod.create_scan_session(str(target), output_path, sender_id_tmp)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    try:
        # Guruhga ulanish — a'zo bo'lmasa join qilish
        try:
            entity = await userbot.get_entity(target)
        except Exception:
            try:
                await userbot(JoinChannelRequest(target))
                await asyncio.sleep(1)
                entity = await userbot.get_entity(target)
            except Exception as e:
                raise Exception(f"Guruhga ulanib bo'lmadi: {e}")

        try:
            await status_msg.edit("📨 Xabarlar o'qilmoqda, foydalanuvchilar aniqlanmoqda...")
        except Exception:
            pass

        # Xabar yozgan unikal foydalanuvchilarni yig'ish
        unique_users = {}  # user_id → user object

        # Sana filtri
        from datetime import timezone
        offset_date = None
        if days:
            from datetime import timedelta
            offset_date = datetime.now(timezone.utc) - timedelta(days=days)

        # Xabarlardan foydalanuvchilarni yig'ish + keshga saqlash
        unique_ids = set()
        msg_count_tmp = 0
        _cache_batch = []   # Kesh uchun batch
        _src_str = str(target)

        async for msg in userbot.iter_messages(entity, limit=None, offset_date=offset_date):
            if msg.sender_id and msg.sender_id > 0:
                sender = msg.sender
                if sender and not getattr(sender, 'bot', False) and hasattr(sender, 'first_name'):
                    if msg.sender_id not in unique_users:
                        unique_users[msg.sender_id] = sender
                else:
                    unique_ids.add(msg.sender_id)

            # Matnli xabarlarni keshga yig'ish (batch)
            if msg.text and len(msg.text) > 2:
                sender = msg.sender
                s_id = getattr(sender, 'id', msg.sender_id or 0) if sender else (msg.sender_id or 0)
                s_name = ""
                s_un   = ""
                if sender and hasattr(sender, 'first_name'):
                    s_name = ((sender.first_name or "") + " " + (sender.last_name or "")).strip()
                    s_un   = getattr(sender, 'username', '') or ""
                msg_dt = msg.date.strftime("%Y-%m-%d %H:%M") if msg.date else ""
                _cache_batch.append((msg.id, _src_str, s_id, s_name, s_un, msg.text[:500], msg_dt))

            # Har 300 xabarda bir marta batch-insert
            if len(_cache_batch) >= 300:
                try:
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                        await _db.executemany(
                            "INSERT OR IGNORE INTO messages_cache "
                            "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                            "VALUES (?,?,?,?,?,?,?)",
                            _cache_batch
                        )
                        await _db.commit()
                    # Alert tekshiruvi (fon taskda)
                    asyncio.create_task(_check_batch_alerts(_cache_batch))
                except Exception:
                    pass
                _cache_batch = []

            msg_count_tmp += 1
            if msg_count_tmp % 1000 == 0:
                try:
                    await status_msg.edit(
                        f"📨 `{msg_count_tmp}` ta xabar o'qildi | "
                        f"👥 `{len(unique_users) + len(unique_ids)}` ta unikal..."
                    )
                except Exception:
                    pass

        # Qolgan batch ni saqlash
        if _cache_batch:
            try:
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                    await _db.executemany(
                        "INSERT OR IGNORE INTO messages_cache "
                        "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                        "VALUES (?,?,?,?,?,?,?)",
                        _cache_batch
                    )
                    await _db.commit()
                asyncio.create_task(_check_batch_alerts(_cache_batch))
            except Exception:
                pass

        # Sender None bo'lganlarni get_entity bilan olish
        missing = unique_ids - set(unique_users.keys())
        if missing:
            try:
                await status_msg.edit(
                    f"👥 `{len(missing)}` ta profil ma'lumoti olinmoqda..."
                )
            except Exception:
                pass
            for uid in missing:
                try:
                    user = await userbot.get_entity(uid)
                    if user and not getattr(user, 'bot', False):
                        unique_users[uid] = user
                except FloodWaitError as e:
                    log_flood("scan_messages_get_entity", e.seconds)
                    await asyncio.sleep(min(e.seconds + 3, 300))
                    try:
                        user = await userbot.get_entity(uid)
                        if user and not getattr(user, 'bot', False):
                            unique_users[uid] = user
                    except Exception:
                        # Ma'lumot olib bo'lmasa ham ID bilan yozish
                        class MinimalUser:
                            def __init__(self, user_id):
                                self.id = user_id
                                self.first_name = ""
                                self.last_name = ""
                                self.username = None
                                self.phone = None
                                self.bot = False
                        unique_users[uid] = MinimalUser(uid)
                except Exception:
                    # Ma'lumot olib bo'lmasa ham ID bilan yozish
                    class MinimalUser:
                        def __init__(self, user_id):
                            self.id = user_id
                            self.first_name = ""
                            self.last_name = ""
                            self.username = None
                            self.phone = None
                            self.bot = False
                    unique_users[uid] = MinimalUser(uid)

        total = len(unique_users)
        try:
            await status_msg.edit(f"👥 {total} ta unikal foydalanuvchi topildi. Profillar tahlil qilinmoqda...")
        except Exception:
            pass

        if not unique_users:
            raise Exception("Xabar yozgan foydalanuvchi topilmadi.")

        # Excel tayyorlash
        wb    = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Xabar Skaneri"
        sheet.append([
            "№", "Ism", "Familiya", "Username", "Telegram ID",
            "Telefon", "Bio", "Bio Linklar",
            "Shaxsiy Kanal Linki", "Maxfiy Kanal Linki", "Profil havolasi"
        ])

        count    = 0
        SAVE_EVERY = 50

        for uid, user in unique_users.items():
            while SCANNER_PAUSED:
                await asyncio.sleep(1)

            count += 1
            if count % 20 == 0:
                try:
                    await status_msg.edit(
                        f"🔍 **Tahlil qilinmoqda:** `{count}/{total}` ta profil..."
                    )
                except Exception:
                    pass

            f_name = user.first_name or ""
            l_name = user.last_name  or ""
            uname  = ("@" + user.username) if user.username else ""
            phone  = getattr(user, 'phone', '') or ""
            bio      = ""
            shaxsiy  = ""
            maxfiy   = ""

            await asyncio.sleep(0.35)
            try:
                fi = await asyncio.wait_for(
                    userbot(GetFullUserRequest(uid)), timeout=20
                )
                bio = fi.full_user.about or ""
                inv_links = extract_invite_links(bio)
                if inv_links:
                    maxfiy = ", ".join(inv_links)
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                        for inv in inv_links:
                            await db.execute(
                                "INSERT OR IGNORE INTO hidden_channel_knocker "
                                "(channel_id, creator_id, source_group) "
                                "VALUES (?, ?, ?)",
                                (inv, uid, str(target))
                            )
                        await db.commit()
                ch_id = getattr(fi.full_user, 'personal_channel_id', None)
                if ch_id:
                    shaxsiy = f"tg://resolve?domain=c{ch_id}"
            except FloodWaitError as e:
                _record_flood(e.seconds)
                log_flood("scan_messages_user", e.seconds)
                await asyncio.sleep(min(e.seconds + 2, 120))
            except Exception:
                pass

            # O'chirilgan hisob → o'tkazib yuborish
            if not f_name and not l_name and not uname and not bio:
                continue

            bio_links_str = ", ".join(extract_bio_links(bio)) if bio else "Yo'q"
            p_link  = ("https://t.me/" + user.username) if user.username else f"tg://user?id={uid}"
            b_date  = extract_exact_birth_date(bio)
            has_db  = maxfiy if maxfiy else (shaxsiy if shaxsiy else "❌")

            sheet.append([
                count, f_name, l_name, uname, uid,
                ("+" + phone) if phone else "",
                bio, bio_links_str,
                shaxsiy, maxfiy, p_link
            ])

            # Bazaga saqlash
            await db_mod.save_user_to_bank(
                uid, str(target), f_name, l_name, uname,
                phone, b_date, bio, bio_links_str, has_db
            )

            if count % SAVE_EVERY == 0:
                apply_excel_styles(sheet, count)
                wb.save(output_path)

    except Exception as e:
        await db_mod.finish_scan_session(scan_id, status='error')
        raise e
    finally:
        _SCAN_COUNT -= 1
        if _SCAN_COUNT <= 0:
            _SCAN_COUNT = 0
            MONITORING_PAUSED = False
            _RESOURCE['music_paused'] = False
            _RESOURCE['profile_slow'] = False
            _RESOURCE['heavy_scan'] = False
            _RESOURCE['current_task'] = None
            _RESOURCE['task_start'] = None

    apply_excel_styles(sheet, count)
    wb.save(output_path)
    await db_mod.finish_scan_session(scan_id, status='done')
    return count



# ─────────────────────────────────────────────────────────────────────
# KANAL COMMENT SKANERLASH
# Kanal postlaridagi commentariyalardan foydalanuvchilarni topadi
# ─────────────────────────────────────────────────────────────────────

async def scan_channel_comments(userbot, target, output_path, status_msg,
                                resume_offset=0, resume_count=0, scan_id=None):
    """
    Kanal postlarining comment qismidan (linked discussion guruh)
    yozgan foydalanuvchilarni topib skanerLaydi.

    Mantiq:
      1. Kanal linked_chat (discussion guruh) ni topadi
      2. Discussion guruhdagi barcha xabarlarni o'qiydi
      3. Unikal foydalanuvchilarni yig'adi
      4. Har birini GetFullUserRequest bilan skanerLaydi
    """
    global _SCAN_COUNT, MONITORING_PAUSED, _FLOOD_PENALTY
    _SCAN_COUNT += 1
    MONITORING_PAUSED = True
    _FLOOD_PENALTY = 0.0
    _RESOURCE['heavy_scan'] = True
    _RESOURCE['current_task'] = "Kanal Comment Skanerlash"
    _RESOURCE['task_start'] = datetime.now()

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    if scan_id is None:
        sender_id = getattr(status_msg, 'chat_id', 0)
        scan_id = await db_mod.create_scan_session(str(target), output_path, sender_id)

    try:
        # 1. Kanalga ulanish (a\'zo bo\'lmasdan)
        try:
            channel = await userbot.get_entity(target)
        except Exception as e:
            raise Exception(f"Kanalga ulanib bo\'lmadi: {e}")

        try:
            await status_msg.edit("🔍 Kanalning discussion guruhi qidirilmoqda...")
        except Exception:
            pass

        # 2. Linked discussion guruhni topish
        discussion_group = None
        try:
            full_ch = await userbot(GetFullChannelRequest(channel))
            linked_id = getattr(full_ch.full_chat, 'linked_chat_id', None)
            if linked_id:
                discussion_group = await userbot.get_entity(linked_id)
        except Exception:
            pass

        if not discussion_group:
            raise Exception(
                "Bu kanalda comment bo'limi (discussion guruh) topilmadi.\n"
                "Kanal postlari ostida comment yozish imkoniyati yoqilmagan bo'lishi mumkin."
            )

        ch_title = getattr(channel, 'title', str(target))
        gr_title = getattr(discussion_group, 'title', 'Discussion')

        try:
            await status_msg.edit(
                f"✅ Discussion guruh topildi: **{gr_title}**\n"
                f"📨 Commentariyalar o'qilmoqda..."
            )
        except Exception:
            pass

        # 3. Discussion guruhdan xabar yozganlarni yig'ish
        unique_users = {}  # user_id → user object
        _msg_count = 0

        async for msg in userbot.iter_messages(discussion_group, limit=None):
            if not msg.sender_id:
                continue
            if msg.sender_id < 0:
                continue  # Kanal/guruh xabarlarini o'tkazib yuborish
            _msg_count += 1
            if msg.sender_id not in unique_users:
                sender = msg.sender
                if sender and not getattr(sender, 'bot', False):
                    unique_users[msg.sender_id] = sender
            # Har 500 xabarda progress ko'rsatish — muzlab qolmagan
            if _msg_count % 500 == 0:
                try:
                    await status_msg.edit(
                        f"📨 **Xabarlar o'qilmoqda:** `{_msg_count}` ta ko'rildi\n"
                        f"👥 Hozircha topilgan: `{len(unique_users)}` ta unikal foydalanuvchi..."
                    )
                except Exception:
                    pass

        total = len(unique_users)
        if not total:
            raise Exception("Comment yozgan foydalanuvchi topilmadi.")

        try:
            await status_msg.edit(
                f"👥 **{total}** ta unikal foydalanuvchi topildi.\n"
                f"🔍 Profillar tahlil qilinmoqda..."
            )
        except Exception:
            pass

        # 4. Parallel profil skanerlash (2 parallel API + flood xavfsiz)
        wb    = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Comment Skaneri"
        sheet.append([
            "№", "Ism", "Familiya", "Username", "Telegram ID",
            "Telefon", "Bio", "Bio Linklar",
            "Shaxsiy Kanal Linki", "Maxfiy Kanal Linki", "Profil havolasi"
        ])

        count      = 0
        SAVE_EVERY = 50

        # Natijalarni yig'ib, keyin tartib bilan yozish
        _cmt_sem  = asyncio.Semaphore(4)   # 4 parallel — _safe_api_call flood boshqaradi
        _progress = [0]                    # shared counter
        collected = {}                     # uid → row_data

        async def _scan_one(uid, user):
            while SCANNER_PAUSED:
                await asyncio.sleep(1)

            f_name = user.first_name or ""
            l_name = user.last_name  or ""
            uname  = ("@" + user.username) if user.username else ""
            phone  = getattr(user, 'phone', '') or ""
            bio    = ""
            shaxsiy = ""
            maxfiy  = ""

            async with _cmt_sem:
                await asyncio.sleep(0.35)
                try:
                    fi = await asyncio.wait_for(
                        userbot(GetFullUserRequest(uid)), timeout=20
                    )
                    bio = fi.full_user.about or ""
                    inv_links = extract_invite_links(bio)
                    if inv_links:
                        maxfiy = ", ".join(inv_links)
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                            for inv in inv_links:
                                await db.execute(
                                    "INSERT OR IGNORE INTO hidden_channel_knocker "
                                    "(channel_id, creator_id, source_group) "
                                    "VALUES (?, ?, ?)",
                                    (inv, uid, str(target))
                                )
                            await db.commit()
                    ch_id = getattr(fi.full_user, 'personal_channel_id', None)
                    if ch_id:
                        shaxsiy = f"tg://resolve?domain=c{ch_id}"
                except FloodWaitError as e:
                    _record_flood(e.seconds)
                    log_flood("scan_channel_comments", e.seconds)
                    await asyncio.sleep(min(e.seconds + 2, 120))
                except Exception:
                    pass

            # O'chirilgan hisob → saqlamay o'tkazib yuborish
            if not f_name and not l_name and not uname and not bio:
                return

            bio_links_str = ", ".join(extract_bio_links(bio)) if bio else "Yo'q"
            p_link  = (f"https://t.me/{user.username}" if user.username
                       else f"tg://user?id={uid}")
            b_date  = extract_exact_birth_date(bio)
            has_db  = maxfiy if maxfiy else (shaxsiy if shaxsiy else "❌")

            collected[uid] = (f_name, l_name, uname, phone, bio,
                              bio_links_str, shaxsiy, maxfiy, p_link, b_date, has_db)

            _progress[0] += 1
            if _progress[0] % 20 == 0:
                try:
                    await status_msg.edit(
                        f"🔍 **Tahlil:** `{_progress[0]}/{total}` ta profil..."
                    )
                except Exception:
                    pass

        # Barcha userlarni parallel ishga tushirish
        await asyncio.gather(
            *[asyncio.create_task(_scan_one(uid, user))
              for uid, user in unique_users.items()],
            return_exceptions=True
        )

        # Natijalarni tartib bilan Excelga yozish
        for uid, user in unique_users.items():
            if uid not in collected:
                continue
            f_name, l_name, uname, phone, bio, bio_links_str, shaxsiy, maxfiy, p_link, b_date, has_db = collected[uid]
            count += 1
            sheet.append([
                count, f_name, l_name, uname, uid,
                ("+" + phone) if phone else "",
                bio, bio_links_str, shaxsiy, maxfiy, p_link
            ])
            await db_mod.save_user_to_bank(
                uid, str(target), f_name, l_name, uname,
                phone, b_date, bio, bio_links_str, has_db
            )
            if count % SAVE_EVERY == 0:
                apply_excel_styles(sheet, count)
                wb.save(output_path)

    except Exception as e:
        await db_mod.finish_scan_session(scan_id, status='error')
        raise e
    else:
        await db_mod.finish_scan_session(scan_id, status='done')
    finally:
        # Oxirgi saqlash — har doim bajariladi
        try:
            _fl2 = asyncio.get_event_loop()
            await _fl2.run_in_executor(None, apply_excel_styles, sheet, count)
            await _fl2.run_in_executor(None, wb.save, output_path)
        except Exception:
            pass
        _SCAN_COUNT -= 1
        if _SCAN_COUNT <= 0:
            _SCAN_COUNT = 0
            MONITORING_PAUSED = False
        _RESOURCE['heavy_scan'] = False
        _RESOURCE['music_paused'] = False
        _RESOURCE['profile_slow'] = False
        _RESOURCE['current_task'] = None
        _RESOURCE['task_start'] = None

    return count, ch_title


# ─────────────────────────────────────────────────────────────────────
# KALIT SO'Z QIDIRUV
# Guruh/kanal xabarlaridan kalit so'z bo'yicha qidiradi
# ─────────────────────────────────────────────────────────────────────

async def search_keywords(userbot, target, keywords_str, status_msg, days=None):
    """
    Guruh/kanal xabarlaridan kalit so'zlarni qidiradi.
    keywords_str: "sotaman, telefon, uy" — vergul bilan ajratilgan
    Qaytaradi: natijalar ro'yxati [{name, username, user_id, date, text, source}]
    """
    # Kalit so'zlarni tayyorlash
    keywords = [k.strip().lower() for k in keywords_str.split(',') if k.strip()]
    if not keywords:
        return []

    results = []

    try:
        # Manba ga ulanish — safe_get_entity bilan flood himoyasi
        entity = await safe_get_entity(userbot, target)
        if entity is None:
            return []  # Kirish imkoni yo'q yoki flood — o'tkazib yuborish

        title = getattr(entity, 'title', str(target))

        try:
            await status_msg.edit(
                f"🔎 **`{', '.join(keywords)}`** qidirilmoqda...\n"
                f"📍 `{title}` xabarlari o'qilmoqda..."
            )
        except Exception:
            pass

        msg_count = 0

        # Sana filtri
        from datetime import timezone
        offset_date = None
        if days:
            from datetime import timedelta
            offset_date = datetime.now(timezone.utc) - timedelta(days=days)

        async for msg in userbot.iter_messages(
            entity, limit=None,
            offset_date=offset_date
        ):
            # Matn yo'q VA audio/musiqa ham yo'q — o'tkazib yuborish
            has_text  = bool(msg.text)
            has_audio = bool(msg.audio or msg.voice)
            if not has_text and not has_audio:
                continue

            # Sana filtri — offset_date dan eski bo'lsa to'xtatish
            if offset_date and msg.date:
                msg_dt = msg.date.replace(tzinfo=timezone.utc) if msg.date.tzinfo is None else msg.date
                if msg_dt < offset_date:
                    break

            msg_count += 1
            if msg_count % 500 == 0:
                try:
                    await status_msg.edit(
                        f"🔎 `{msg_count}` ta xabar ko'rildi | "
                        f"✅ Topildi: `{len(results)}` ta..."
                    )
                except Exception:
                    pass

            # Qidiruv matni: caption + audio title + performer + fayl nomi
            search_parts = []
            if msg.text:
                search_parts.append(msg.text)
            if has_audio:
                audio = msg.audio or msg.voice
                if hasattr(audio, 'title') and audio.title:
                    search_parts.append(audio.title)
                if hasattr(audio, 'performer') and audio.performer:
                    search_parts.append(audio.performer)
                if msg.file and msg.file.name:
                    search_parts.append(msg.file.name)

            if not search_parts:
                continue

            search_text = " ".join(search_parts).lower()
            matched = [kw for kw in keywords if kw in search_text]
            if not matched:
                continue

            # Kim yozgan?

            sender = msg.sender
            if sender is None:
                try:
                    sender = await msg.get_sender()
                except Exception:
                    continue
            if not sender:
                continue
            if getattr(sender, 'bot', False):
                continue

            name     = ""
            username = ""
            uid      = getattr(sender, 'id', 0)

            if hasattr(sender, 'first_name'):
                name = (sender.first_name or "") + " " + (sender.last_name or "")
                name = name.strip()
                username = sender.username or ""
            elif hasattr(sender, 'title'):
                # Kanal/guruh nomi
                name = sender.title or ""

            # Natija matni: audio bo'lsa sarlavha + ijrochi, aks holda caption
            if has_audio:
                audio = msg.audio or msg.voice
                a_title = getattr(audio, 'title', '') or ''
                a_perf  = getattr(audio, 'performer', '') or ''
                f_name  = (msg.file.name if msg.file and msg.file.name else '')
                display = ""
                if a_title:
                    display = f"🎵 {a_title}"
                    if a_perf:
                        display += f" — {a_perf}"
                elif f_name:
                    display = f"🎵 {f_name}"
                else:
                    display = "🎵 Audio fayl"
                if msg.text:
                    display += f"\n💬 {msg.text[:100]}"
                text = display
            else:
                text = msg.text or ""
                if len(text) > 250:
                    text = text[:247] + "..."

            results.append({
                'name':     name,
                'username': username,
                'user_id':  uid,
                'date':     msg.date.strftime("%Y-%m-%d %H:%M") if msg.date else "",
                'text':     text,
                'source':   title,
                'matched':  ", ".join(matched)
            })

        try:
            await status_msg.edit(
                f"✅ Qidiruv yakunlandi!\n"
                f"📊 Ko'rilgan xabarlar: `{msg_count}` ta\n"
                f"🎯 Topildi: `{len(results)}` ta"
            )
        except Exception:
            pass

    except Exception as e:
        raise e

    return results


# ─────────────────────────────────────────────────────────────────────
# KANAL MUSIQA TRACKER
# Monitoring kanallaridagi audio xabarlarni skanerLaydi
# ─────────────────────────────────────────────────────────────────────

async def music_channel_tracker(userbot):
    """
    Monitoring kanallaridagi barcha audio xabarlarni yuklab,
    fingerprint oladi va saqlaydi. Audio keyin o'chiriladi.
    Faqat yangi xabarlarni tekshiradi (oxirgi ID saqlanadi).
    """
    await music_mod.init_music_db()

    while True:
        try:
            # Barcha manbalarni olish
            sources = await music_mod.get_all_sources()

            for source in sources:
                # Og'ir skanerlash ishlayotgan bo'lsa kutish
                while _RESOURCE['music_paused']:
                    await asyncio.sleep(10)
                if MONITORING_PAUSED:
                    await asyncio.sleep(5)
                    continue

                await asyncio.sleep(1)  # Har kanal oldida pauza — flood himoyasi
                try:
                    entity = await safe_get_entity(userbot, source)
                    if entity is None:
                        continue
                except Exception:
                    # Kanalga kira olmadik — hidden_channel_knocker ga qo'shish
                    try:
                        is_invite = "t.me/+" in str(source) or "t.me/joinchat/" in str(source)
                        if is_invite:
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                                await db.execute(
                                    "INSERT OR IGNORE INTO hidden_channel_knocker "
                                    "(channel_id, creator_id, source_group, last_request_time) "
                                    "VALUES (?, 0, 'Musiqa Tracker', ?)",
                                    (source, now_str)
                                )
                                await db.commit()
                    except Exception:
                        pass
                    continue

                channel_name = getattr(entity, 'title', str(source))
                channel_id   = str(entity.id)

                # Oxirgi skanerlangan xabar ID ni olish
                last_msg_id = 0
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                    try:
                        async with db.execute(
                            "SELECT last_msg_id FROM music_channel_progress WHERE channel_id=?",
                            (channel_id,)
                        ) as cur:
                            row = await cur.fetchone()
                            if row:
                                last_msg_id = row[0] or 0
                    except Exception:
                        # Jadval yo'q bo'lsa yaratish
                        await db.execute(
                            "CREATE TABLE IF NOT EXISTS music_channel_progress "
                            "(channel_id TEXT PRIMARY KEY, last_msg_id INTEGER)"
                        )
                        await db.commit()

                # Yangi audio xabarlarni olish
                new_last_id = last_msg_id
                audio_count = 0

                # Birinchi marta (last_msg_id=0) — barcha xabarlar
                # Keyingi safar — faqat yangilari
                iter_kwargs = {"limit": None}
                if last_msg_id > 0:
                    iter_kwargs["min_id"] = last_msg_id

                BASE_DIR_LOCAL = os.path.dirname(os.path.abspath(__file__))

                # i3/i5 uchun: 3 parallel — 2 yadroli CPU uchun optimal
                CONCURRENCY = 3
                _ch_sem   = asyncio.Semaphore(CONCURRENCY)
                _ch_tasks = set()

                async def _pipeline(m):
                    nonlocal audio_count
                    tmp_path = os.path.join(BASE_DIR_LOCAL, f"tmp_ch_{channel_id}_{m.id}.ogg")
                    async with _ch_sem:
                        # Yuklab olish (3 urinish)
                        ok = False
                        for attempt in range(3):
                            try:
                                await m.download_media(file=tmp_path)
                                if os.path.exists(tmp_path) and os.path.getsize(tmp_path) > 0:
                                    ok = True
                                    break
                                if os.path.exists(tmp_path):
                                    os.remove(tmp_path)
                            except Exception:
                                if os.path.exists(tmp_path):
                                    os.remove(tmp_path)
                                if attempt < 2:
                                    await asyncio.sleep(1)
                        if not ok:
                            return
                        try:
                            fp, duration = await music_mod.get_fingerprint_async(tmp_path)
                            if fp:
                                await music_mod.save_fingerprint(
                                    channel_id, channel_name,
                                    f"msg_{m.id}", fp, duration or 0
                                )
                                audio_count += 1
                                hits = await music_mod.check_against_watch_list(fp)
                                for hit in hits:
                                    _WATCH_ALERTS.put_nowait({
                                        'admin_id':    hit['admin_id'],
                                        'watch_name':  hit['watch_name'],
                                        'score':       hit['score'],
                                        'source_name': channel_name,
                                        'source_id':   channel_id,
                                        'source_type': 'kanal'
                                    })
                        except Exception:
                            pass
                        finally:
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)

                _cache_batch = []
                _cache_src   = str(source)

                async for msg in userbot.iter_messages(entity, **iter_kwargs):
                    if is_music_file(msg):
                        if msg.id > new_last_id:
                            new_last_id = msg.id
                        t = asyncio.create_task(_pipeline(msg))
                        _ch_tasks.add(t)
                        t.add_done_callback(_ch_tasks.discard)

                    # Matnli xabarlarni keshga yig'ish
                    if msg.text and len(msg.text) > 2:
                        sender = msg.sender
                        s_id   = getattr(sender, 'id', msg.sender_id or 0) if sender else (msg.sender_id or 0)
                        s_name = ""
                        s_un   = ""
                        if sender and hasattr(sender, 'first_name'):
                            s_name = ((sender.first_name or "") + " " + (sender.last_name or "")).strip()
                            s_un   = getattr(sender, 'username', '') or ""
                        msg_dt = msg.date.strftime("%Y-%m-%d %H:%M") if msg.date else ""
                        _cache_batch.append((msg.id, _cache_src, s_id, s_name, s_un, msg.text[:500], msg_dt))

                    if len(_cache_batch) >= 300:
                        try:
                            async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                                await _db.executemany(
                                    "INSERT OR IGNORE INTO messages_cache "
                                    "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                                    "VALUES (?,?,?,?,?,?,?)",
                                    _cache_batch
                                )
                                await _db.commit()
                            # Musiqa skanerida alert tekshiruv yo'q — parallel yuk kamaytirish
                        except Exception:
                            pass
                        _cache_batch = []

                # Qolgan batchni saqlash
                if _cache_batch:
                    try:
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _db:
                            await _db.executemany(
                                "INSERT OR IGNORE INTO messages_cache "
                                "(msg_id,source,sender_id,sender_name,sender_username,text,msg_date) "
                                "VALUES (?,?,?,?,?,?,?)",
                                _cache_batch
                            )
                            await _db.commit()
                        # Alert tekshiruv yo'q — real-time monitoring qiladi
                    except Exception:
                        pass

                if _ch_tasks:
                    await asyncio.gather(*_ch_tasks, return_exceptions=True)

                # Oxirgi xabar ID ni yangilash - xatolik bo'lsa ham saqlash
                if new_last_id > last_msg_id:
                    try:
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                            await db.execute(
                                "CREATE TABLE IF NOT EXISTS music_channel_progress "
                                "(channel_id TEXT PRIMARY KEY, last_msg_id INTEGER)"
                            )
                            await db.execute(
                                "INSERT OR REPLACE INTO music_channel_progress "
                                "(channel_id, last_msg_id) VALUES (?, ?)",
                                (channel_id, new_last_id)
                            )
                            await db.commit()
                    except Exception as e:
                        print(f"Progress saqlash xatosi: {e}")

                await asyncio.sleep(random.uniform(0.5, 1.5))

        except Exception as e:
            if 'FloodWait' in str(type(e).__name__):
                wait = getattr(e, 'seconds', 60)
                log_flood("music_channel_tracker", wait)
            print(f"music_channel_tracker xatosi: {e}")

        # Barcha kanallar tekshirildi
        global _CHANNEL_MUSIC_DONE
        _CHANNEL_MUSIC_DONE = True
        print("[MUSIQA] Barcha kanal musiqalari skanerlandi — profil musiqasiga o'tiladi")

        # 1 soat kutish, keyin qaytadan
        await asyncio.sleep(3600)
        _CHANNEL_MUSIC_DONE = False


# ─────────────────────────────────────────────────────────────────────
# PARALLEL MUSIQA SKANERLASH — har a'zo uchun
# ─────────────────────────────────────────────────────────────────────

def is_music_file(msg):
    """Musiqa fayllarini tekshiradi (voice xabarlar emas)"""
    if msg.voice:
        return False  # Voice xabar — o'tkazib yuborish
    if msg.audio:
        mime = getattr(msg.audio, 'mime_type', '') or ''
        mime = mime.lower()
        # Mime bo'sh bo'lsa ham audio bo'lsa olish
        if not mime:
            return True
        # Voice/video formatlarini o'tkazib yuborish
        skip = ['video/', 'image/', 'application/']
        if any(s in mime for s in skip):
            return False
        return True  # Barcha audio formatlarini olish
    return False

async def _scan_user_music(userbot, uid, name, channel_link, full_info=None):
    """
    Har a'zo skanerlanganda parallel ishga tushadigan funksiya:
    1. Profil musiqasi → fingerprint → o'chiradi
    2. Shaxsiy kanal musiqalari → fingerprint → o'chiradi
    full_info: allaqachon olingan GetFullUserRequest natijasi (agar bor bo'lsa)
    """
    await music_mod.init_music_db()
    BASE_DIR_LOCAL = os.path.dirname(os.path.abspath(__file__))

    try:
        # Agar full_info berilmagan bo'lsa — yangi so'rov yuborish
        if full_info is not None:
            fi = full_info
        else:
            try:
                fi = await userbot(GetFullUserRequest(uid))
            except FloodWaitError as e:
                log_flood("_scan_user_music", e.seconds)
                await asyncio.sleep(min(e.seconds + 3, 600))
                fi = await userbot(GetFullUserRequest(uid))
        fu = fi.full_user

        # 1. Profil musiqalari
        music_docs = []
        for field in ['saved_music', 'profile_song', 'profile_songs', 'music']:
            val = getattr(fu, field, None)
            if val is None:
                continue
            if isinstance(val, list):
                music_docs.extend(val)
            else:
                music_docs.append(val)

        for idx, doc in enumerate(music_docs):
            if hasattr(doc, 'document'):
                doc = doc.document
            if not hasattr(doc, 'id'):
                continue
            tmp_path = os.path.join(BASE_DIR_LOCAL, f"tmp_profile_{uid}_{idx}.ogg")
            try:
                await userbot.download_media(doc, file=tmp_path)
                if os.path.exists(tmp_path):
                    fp, duration = await music_mod.get_fingerprint_async(tmp_path)
                    if fp:
                        await music_mod.save_fingerprint(
                            str(uid), f"Profil: {name}",
                            f"profile_{uid}_{idx}", fp, duration or 0
                        )
            except Exception:
                pass
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

    except Exception:
        pass

    # 2. Shaxsiy kanal musiqalari
    if channel_link and channel_link.startswith('http'):
        try:
            entity = await userbot.get_entity(channel_link)
            channel_id   = str(entity.id)
            channel_name = getattr(entity, 'title', channel_link)

            audio_msgs = []
            async for msg in userbot.iter_messages(entity, limit=None):
                if is_music_file(msg):
                    audio_msgs.append(msg)

            async def _dl2(m):
                tmp = os.path.join(BASE_DIR_LOCAL, f"tmp_ch_{channel_id}_{m.id}.ogg")
                for attempt in range(3):
                    try:
                        await m.download_media(file=tmp)
                        if os.path.exists(tmp) and os.path.getsize(tmp) > 0:
                            return (m.id, tmp)
                        if os.path.exists(tmp):
                            os.remove(tmp)
                    except Exception:
                        if os.path.exists(tmp):
                            os.remove(tmp)
                return (m.id, None)

            BATCH = 3  # i3/i5 uchun: 3 parallel yuklab olish
            for i in range(0, len(audio_msgs), BATCH):
                batch = audio_msgs[i:i+BATCH]
                dl_results = await asyncio.gather(*[_dl2(m) for m in batch])
                for msg_id, tmp in dl_results:
                    if tmp and os.path.exists(tmp):
                        try:
                            fp, dur = await music_mod.get_fingerprint_async(tmp)
                            if fp:
                                await music_mod.save_fingerprint(
                                    channel_id, channel_name,
                                    f"msg_{msg_id}", fp, dur or 0
                                )
                        except Exception:
                            pass
                        finally:
                            if os.path.exists(tmp):
                                os.remove(tmp)

                await asyncio.sleep(1)

        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────
# YANGI MAXFIY KANAL MONITOR
# Kuniga 50 ta so'rovnoma, har 28 daqiqada 1 ta
# Har 28 daqiqada kirish tekshiriladi
# Kirish ochildi → musiqa skanerlash
# ─────────────────────────────────────────────────────────────────────

_daily_knock_count = 0
_daily_knock_date  = ""
MAX_DAILY_KNOCKS   = 50   # Kuniga maksimal so'rovnomalar
KNOCK_INTERVAL     = 15 * 60  # 15 daqiqa (sekund)


async def smart_channel_knocker(userbot, bot, admin_id):
    """
    Har 28 daqiqada:
    1. Pending kanallardan biriga so'rovnoma yuboradi (kuniga max 50)
    2. Allaqachon so'rovnoma yuborilgan kanallar kirilganmi tekshiradi
    3. Kirish ochildi → musiqa skanerlash + admin ga xabar
    """
    global _daily_knock_count, _daily_knock_date

    # Bot yonganda bazadan bugungi knock countni yuklash
    today = datetime.now().strftime("%Y-%m-%d")
    try:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "CREATE TABLE IF NOT EXISTS knock_state "
                "(key TEXT PRIMARY KEY, value TEXT)"
            )
            await db.commit()
            async with db.execute(
                "SELECT value FROM knock_state WHERE key='knock_date'"
            ) as cur:
                row = await cur.fetchone()
            if row and row[0] == today:
                async with db.execute(
                    "SELECT value FROM knock_state WHERE key='knock_count'"
                ) as cur:
                    cnt = await cur.fetchone()
                if cnt:
                    _daily_knock_count = int(cnt[0])
                    _daily_knock_date = today
                    print(f"[KNOCKER] Bugungi count yuklandi: {_daily_knock_count}/{MAX_DAILY_KNOCKS}")
    except Exception as e:
        print(f"[KNOCKER] Count yuklashda xato: {e}")

    # Bot yonganda 5 daqiqa kutish
    await asyncio.sleep(300)

    while True:
        await asyncio.sleep(KNOCK_INTERVAL)

        if MONITORING_PAUSED:
            continue

        try:
            today = datetime.now().strftime("%Y-%m-%d")
            if _daily_knock_date != today:
                _daily_knock_date  = today
                _daily_knock_count = 0
                # Yangi kun — bazani tozalash
                try:
                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                        await db.execute(
                            "INSERT OR REPLACE INTO knock_state (key, value) VALUES (?, ?)",
                            ("knock_date", today)
                        )
                        await db.execute(
                            "INSERT OR REPLACE INTO knock_state (key, value) VALUES (?, ?)",
                            ("knock_count", "0")
                        )
                        await db.commit()
                except Exception:
                    pass

            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                # Faqat 1 ta kanal — navbat bilan
                async with db.execute(
                    "SELECT channel_id, creator_id, source_group, last_request_time "
                    "FROM hidden_channel_knocker WHERE status='pending' "
                    "ORDER BY last_request_time ASC LIMIT 1"
                ) as cur:
                    row = await cur.fetchone()

            if not row:
                continue
            tasks = [row]

            for ch_id_str, creator_id, source_group, last_req in tasks:
                if MONITORING_PAUSED:
                    break

                # 1. Kirish ochildimi tekshirish
                joined = False
                try:
                    entity = await userbot.get_entity(ch_id_str)
                    # Kirish mumkin → ochildi
                    joined = True
                except Exception:
                    pass

                if not joined and ("/+" in ch_id_str or "joinchat/" in ch_id_str):
                    # Invite link tekshirish
                    try:
                        from telethon.tl.functions.messages import CheckChatInviteRequest
                        if "/+" in ch_id_str:
                            hash_part = ch_id_str.split("/+")[-1].rstrip("/")
                        else:
                            hash_part = ch_id_str.split("joinchat/")[-1].rstrip("/")
                        invite_info = await userbot(CheckChatInviteRequest(hash=hash_part))
                        if hasattr(invite_info, 'chat'):
                            joined = True
                            entity = invite_info.chat
                    except Exception as e:
                        err = str(e).lower()
                        if "already" in err or "member" in err:
                            joined = True
                            try:
                                entity = await userbot.get_entity(ch_id_str)
                            except Exception:
                                joined = False

                if joined:
                    # Kirish ochildi!
                    ch_link = ch_id_str
                    ch_name = getattr(entity, 'title', ch_id_str) if hasattr(entity, 'title') else ch_id_str
                    # Haqiqiy raqamli ID — Excel "Kanal ID" ustuni uchun
                    numeric_id_str = ""
                    if hasattr(entity, 'id') and entity.id:
                        _eid = str(entity.id).lstrip('-')
                        numeric_id_str = f"-100{_eid}" if not str(entity.id).startswith('-100') else str(entity.id)

                    await bot.send_message(
                        admin_id,
                        f"🔓 **MAXFIY KANALGA KIRISH OCHILDI!**\n\n"
                        f"📢 Kanal: `{ch_name}`\n"
                        f"🔗 Link: {ch_link}\n"
                        f"🆔 ID: `{numeric_id_str}`\n"
                        f"🏢 Manba: `{source_group}`\n\n"
                        f"🎵 Kanal musiqalari skanerlanmoqda..."
                    )

                    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                        await db.execute(
                            "UPDATE hidden_channel_knocker SET status='joined', numeric_id=? WHERE channel_id=?",
                            (numeric_id_str, ch_id_str)
                        )
                        if creator_id:
                            await db.execute(
                                "UPDATE users_memory_bank SET has_hidden=? WHERE user_id=?",
                                (ch_link, creator_id)
                            )
                        await db.commit()

                    # Kanal musiqalarini skanerlash
                    asyncio.create_task(
                        _scan_channel_music_after_join(userbot, bot, admin_id, entity, ch_link)
                    )
                    continue

                # 2. So'rovnoma yuborish — limit yo'q, davomiy

                last_dt = None
                try:
                    last_dt = datetime.strptime(last_req, "%Y-%m-%d %H:%M")
                except Exception:
                    pass

                elapsed = (datetime.now() - last_dt).total_seconds() if last_dt else 99999

                if elapsed >= 86400:  # 24 soat o'tgan
                    sent = await send_join_request(userbot, ch_id_str)
                    if sent:
                        _daily_knock_count += 1
                        # Bazaga saqlash
                        try:
                            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                                await db.execute(
                                    "INSERT OR REPLACE INTO knock_state (key, value) VALUES (?, ?)",
                                    ("knock_date", today)
                                )
                                await db.execute(
                                    "INSERT OR REPLACE INTO knock_state (key, value) VALUES (?, ?)",
                                    ("knock_count", str(_daily_knock_count))
                                )
                                await db.commit()
                        except Exception:
                            pass
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                            await db.execute(
                                "UPDATE hidden_channel_knocker SET last_request_time=? WHERE channel_id=?",
                                (now_str, ch_id_str)
                            )
                            await db.commit()

        except Exception as e:
            if 'FloodWait' in str(type(e).__name__):
                wait = getattr(e, 'seconds', 60)
                log_flood("smart_channel_knocker", wait)
            print(f"smart_channel_knocker xatosi: {e}")


async def _scan_channel_music_after_join(userbot, bot, admin_id, entity, ch_link):
    """Kanal ochildi — ichidagi barcha musiqalarni skanerla."""
    await music_mod.init_music_db()
    BASE_DIR_LOCAL = os.path.dirname(os.path.abspath(__file__))
    channel_id   = str(entity.id) if hasattr(entity, 'id') else str(entity)
    channel_name = getattr(entity, 'title', ch_link)
    count = 0

    try:
        audio_msgs = []
        async for msg in userbot.iter_messages(entity, limit=None):
            if is_music_file(msg):
                audio_msgs.append(msg)

        async def _dl_join(m):
            """3 marta urinish bilan yuklab olish"""
            tmp = os.path.join(BASE_DIR_LOCAL, f"tmp_join_{channel_id}_{m.id}.ogg")
            for attempt in range(3):
                try:
                    await m.download_media(file=tmp)
                    if os.path.exists(tmp) and os.path.getsize(tmp) > 0:
                        return (m.id, tmp)
                    if os.path.exists(tmp):
                        os.remove(tmp)
                except Exception:
                    if os.path.exists(tmp):
                        os.remove(tmp)
            return (m.id, None)

        BATCH = 3  # i3/i5 uchun: 3 parallel yuklab olish
        for i in range(0, len(audio_msgs), BATCH):
            batch = audio_msgs[i:i+BATCH]
            dl_results = await asyncio.gather(*[_dl_join(m) for m in batch])
            for msg_id, tmp in dl_results:
                if tmp and os.path.exists(tmp):
                    try:
                        fp, dur = await music_mod.get_fingerprint_async(tmp)
                        if fp:
                            await music_mod.save_fingerprint(
                                channel_id, channel_name,
                                f"msg_{msg_id}", fp, dur or 0
                            )
                            count += 1
                            # Kuzatiladigan musiqa tekshirish
                            hits = await music_mod.check_against_watch_list(fp)
                            for hit in hits:
                                _WATCH_ALERTS.put_nowait({
                                    'admin_id':    hit['admin_id'],
                                    'watch_name':  hit['watch_name'],
                                    'score':       hit['score'],
                                    'source_name': channel_name,
                                    'source_id':   channel_id,
                                    'source_type': 'kanal'
                                })
                    except Exception:
                        pass
                    finally:
                        if os.path.exists(tmp):
                            os.remove(tmp)

        await bot.send_message(
            admin_id,
            f"✅ **`{channel_name}`** musiqa skanerlash yakunlandi!\n"
            f"🎵 Topildi: `{count}` ta fingerprint"
        )

    except Exception as e:
        print(f"_scan_channel_music_after_join xatosi: {e}")


# ─────────────────────────────────────────────────────────────────────
# LOKAL XABARLAR KESHI — kalit so'z qidirish uchun
# ─────────────────────────────────────────────────────────────────────

async def sync_source_messages(userbot, source: str, limit_days: int = 90):
    """
    Bir manbaning yangi xabarlarini lokal bazaga saqlaydi.
    Faqat oxirgi sinxronlashdan keyingi xabarlarni oladi.
    """
    import database as db_mod
    from datetime import timezone, timedelta

    entity = await safe_get_entity(userbot, source)
    if entity is None:
        return 0

    # Oxirgi saqlangan xabar ID sini olish
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT last_msg_id FROM source_sync_state WHERE source=?", (source,)
        ) as cur:
            row = await cur.fetchone()
        last_id = row[0] if row else 0

    # Sana chegarasi — birinchi marta bo'lsa limit_days ga qadar
    offset_date = None
    if last_id == 0:
        offset_date = datetime.now(timezone.utc) - timedelta(days=limit_days)

    saved = 0
    new_last_id = last_id
    msg_counter = 0

    try:
        async for msg in userbot.iter_messages(
            entity,
            limit=None,
            min_id=last_id,
            offset_date=offset_date,
            reverse=True
        ):
            if not msg or not msg.id:
                continue

            # Faqat matnli / audio xabarlar
            text = ""
            if msg.text:
                text = msg.text
            elif msg.audio or msg.voice:
                parts = []
                audio = msg.audio or msg.voice
                if hasattr(audio, 'title') and audio.title:
                    parts.append(audio.title)
                if hasattr(audio, 'performer') and audio.performer:
                    parts.append(audio.performer)
                if msg.file and msg.file.name:
                    parts.append(msg.file.name)
                if msg.message:
                    parts.append(msg.message)
                text = " | ".join(parts)

            if not text:
                continue

            sender = msg.sender
            s_id   = getattr(sender, 'id', 0) if sender else 0
            s_name = ""
            s_un   = ""
            if sender and hasattr(sender, 'first_name'):
                s_name = ((sender.first_name or "") + " " + (sender.last_name or "")).strip()
                s_un   = sender.username or ""
            elif sender and hasattr(sender, 'title'):
                s_name = sender.title or ""

            msg_date = msg.date.strftime("%Y-%m-%d %H:%M") if msg.date else ""

            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                try:
                    await db.execute(
                        "INSERT OR IGNORE INTO messages_cache "
                        "(msg_id, source, sender_id, sender_name, sender_username, text, msg_date) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (msg.id, source, s_id, s_name, s_un, text, msg_date)
                    )
                    await db.commit()
                    saved += 1
                except Exception:
                    pass

            if msg.id > new_last_id:
                new_last_id = msg.id

            msg_counter += 1
            # Har 100 xabarda 2 sekund nafas — flood oldini olish
            if msg_counter % 100 == 0:
                await asyncio.sleep(2)
            else:
                await asyncio.sleep(0.03)

    except FloodWaitError as e:
        # Flood bo'lsa — holatni saqlash va chiqish (keyingi sinxronda davom etadi)
        log_flood("sync_source_messages", e.seconds)
        if e.seconds <= 120:
            await asyncio.sleep(e.seconds + 2)
    except Exception:
        pass

    # Sinxron holatini yangilash
    if new_last_id > last_id:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT OR REPLACE INTO source_sync_state (source, last_msg_id, last_synced) "
                "VALUES (?, ?, ?)",
                (source, new_last_id, now_str)
            )
            await db.commit()

    return saved


async def search_keywords_local(keyword_str: str, days: int = None):
    """
    Lokal messages_cache dan kalit so'z qidiradi.
    Telegram API ga murojaat qilmaydi.
    Qaytaradi: natijalar ro'yxati [{name, username, user_id, date, text, source, matched}]
    """
    import database as db_mod
    from datetime import timezone, timedelta

    keywords = [k.strip().lower() for k in keyword_str.split(',') if k.strip()]
    if not keywords:
        return []

    # SQL LIKE shartlari
    like_clauses = " OR ".join(["LOWER(text) LIKE ?" for _ in keywords])
    like_params  = [f"%{kw}%" for kw in keywords]

    date_clause = ""
    date_param  = []
    if days:
        cutoff = (datetime.now() - __import__('datetime').timedelta(days=days)).strftime("%Y-%m-%d")
        date_clause = "AND msg_date >= ?"
        date_param  = [cutoff]

    query = f"""
        SELECT sender_id, sender_name, sender_username, text, msg_date, source
        FROM messages_cache
        WHERE ({like_clauses})
        {date_clause}
        ORDER BY msg_date DESC
        LIMIT 500
    """

    results = []
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(query, like_params + date_param) as cur:
            rows = await cur.fetchall()

    for (s_id, s_name, s_un, text, msg_date, source) in rows:
        search_text = (text or "").lower()
        matched = [kw for kw in keywords if kw in search_text]
        if not matched:
            continue
        display = text or ""
        if len(display) > 250:
            display = display[:247] + "..."
        results.append({
            'name':     s_name or "",
            'username': s_un   or "",
            'user_id':  s_id   or 0,
            'date':     msg_date or "",
            'text':     display,
            'source':   source or "",
            'matched':  ", ".join(matched)
        })

    return results


async def get_cache_stats():
    """Kesh statistikasi."""
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=15) as db:
        async with db.execute("SELECT COUNT(*) FROM messages_cache") as cur:
            total = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(DISTINCT source) FROM messages_cache") as cur:
            sources = (await cur.fetchone())[0]
        async with db.execute(
            "SELECT MAX(last_synced) FROM source_sync_state"
        ) as cur:
            last_sync = (await cur.fetchone())[0]
    return total, sources, last_sync


# ─────────────────────────────────────────────────────────────────────
# ID BO'YICHA QIDIRUV
# ─────────────────────────────────────────────────────────────────────

def _make_msg_link(source: str, msg_id: int) -> str:
    """Xabar havolasini yaratish."""
    if not source or not msg_id:
        return ""
    s = source.strip()
    # t.me/username/msg_id
    if 't.me/' in s and '+' not in s and 'joinchat' not in s:
        uname = s.split('t.me/')[-1].split('/')[0].rstrip('/')
        if uname:
            return f"https://t.me/{uname}/{msg_id}"
    if s.startswith('@'):
        return f"https://t.me/{s[1:]}/{msg_id}"
    # Numeric channel ID
    if s.lstrip('-').isdigit():
        num = abs(int(s))
        if num > 1000000000:
            return f"https://t.me/c/{num}/{msg_id}"
    return ""


async def lookup_channel_by_id(userbot, channel_id: int):
    """
    Kanal ID si bo'yicha kanal ma'lumotlari va havola qaytaradi.
    Avval DB dan, keyin Telegram API dan qidiradi.
    """
    import database as db_mod

    # 1. Lokal DB dan qidirish
    abs_id = str(abs(channel_id))
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=15) as db:
        async with db.execute(
            "SELECT channel_id, numeric_id FROM hidden_channel_knocker "
            "WHERE numeric_id=? OR channel_id=?",
            (abs_id, str(channel_id))
        ) as cur:
            row = await cur.fetchone()

    if row:
        ch_link = row[0]
        return {'title': ch_link, 'username': None, 'members': None}, ch_link

    # 2. resolved_channel_ids dan
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=15) as db:
        async with db.execute(
            "SELECT channel_link FROM resolved_channel_ids WHERE numeric_id=?",
            (abs_id,)
        ) as cur:
            row2 = await cur.fetchone()
    if row2:
        return {'title': row2[0], 'username': None, 'members': None}, row2[0]

    # 3. Telegram API dan (bir marta so'rov)
    try:
        entity = await safe_get_entity(userbot, channel_id)
        if entity is None:
            return None, None
        title   = getattr(entity, 'title', str(channel_id))
        uname   = getattr(entity, 'username', None)
        members = getattr(getattr(entity, 'participants_count', None), '__int__', lambda: None)()
        if uname:
            link = f"https://t.me/{uname}"
        else:
            link = f"https://t.me/c/{abs(channel_id)}"
        return {'title': title, 'username': uname, 'members': members}, link
    except Exception:
        return None, None


async def lookup_user_by_id(user_id: int):
    """
    Profil ID bo'yicha to'liq qidiruv:
    1. users_memory_bank — a'zo guruhlar + profil ma'lumotlari (xabar yozmagan bo'lsa ham)
    2. messages_cache    — yozgan xabarlar + havolalar
    """
    import database as db_mod

    # 1. Profil va guruhlar
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT MAX(first_name), MAX(last_name), MAX(username), MAX(phone),
                   MAX(bio), MAX(open_channels), MAX(has_hidden),
                   GROUP_CONCAT(DISTINCT group_link), MIN(added_date), MAX(last_updated)
            FROM users_memory_bank WHERE user_id=?
            """,
            (user_id,)
        ) as cur:
            pr = await cur.fetchone()

    # 2. Xabarlar
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT msg_id, source, sender_name, sender_username, text, msg_date "
            "FROM messages_cache WHERE sender_id=? ORDER BY msg_date DESC",
            (user_id,)
        ) as cur:
            msg_rows = await cur.fetchall()

    profile = None
    if pr and any(pr):
        fn, ln, un, ph, bio, oc, hh, grps, added, updated = pr
        profile = {
            'first_name':    fn or "",
            'last_name':     ln or "",
            'username':      un or "",
            'phone':         ph or "",
            'bio':           bio or "",
            'open_channels': oc or "",
            'has_hidden':    hh or "",
            'groups':        [g.strip() for g in (grps or "").split(',') if g.strip()],
            'added_date':    added or "",
            'last_updated':  updated or "",
        }

    messages = []
    for msg_id, source, name, username, text, msg_date in msg_rows:
        link = _make_msg_link(source, msg_id)
        src_title = source or ""
        if 't.me/' in src_title:
            src_title = src_title.split('t.me/')[-1].rstrip('/')
        elif src_title.startswith('@'):
            src_title = src_title[1:]
        messages.append({
            'msg_id':       msg_id,
            'source':       source,
            'source_title': src_title,
            'name':         name or "",
            'username':     username or "",
            'text':         (text or "")[:300],
            'date':         msg_date or "",
            'link':         link,
        })

    return profile, messages


async def lookup_user_messages(user_id: int):
    _, messages = await lookup_user_by_id(user_id)
    return messages


# ═════════════════════════════════════════════════════════════════════
# YANGI FUNKSIYALAR — TERGOV VA SCAM ANIQLASH
# ═════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────
# #7 TRUST SCORE
# ─────────────────────────────────────────────────────────────────────

def calculate_trust_score(profile: dict, msg_count: int = 0) -> tuple:
    score = 0
    reasons = []
    if profile.get('phone'):
        score += 25; reasons.append("+25 telefon")
    if profile.get('bio') and len(profile['bio']) > 3:
        score += 15; reasons.append("+15 bio")
    if profile.get('username'):
        score += 10; reasons.append("+10 username")
    if profile.get('open_channels') and profile['open_channels'] not in ('', "Yo'q", 'Yoq'):
        score += 10; reasons.append("+10 ochiq kanal")
    if len(profile.get('groups', [])) >= 3:
        score += 10; reasons.append("+10 3+ guruh")
    if msg_count > 0:
        score += 10; reasons.append("+10 xabar yozgan")
    if msg_count > 50:
        score += 5; reasons.append("+5 faol")
    if profile.get('has_hidden') and profile['has_hidden'] not in ('', '❌', 'Yoq', "Yo'q"):
        score -= 20; reasons.append("-20 maxfiy kanal")
    if profile.get('added_date'):
        try:
            added = datetime.strptime(profile['added_date'][:10], '%Y-%m-%d')
            days = (datetime.now() - added).days
            if days > 365:
                score += 15; reasons.append("+15 eski akkaunt")
            elif days > 90:
                score += 8; reasons.append("+8 90+ kun")
        except Exception:
            pass
    score = max(0, min(100, score))
    if score >= 70:
        label = "🟢 Ishonchli"
    elif score >= 40:
        label = "🟡 O'rta"
    else:
        label = "🔴 Shubhali"
    return score, label, reasons


# ─────────────────────────────────────────────────────────────────────
# #10 USERNAME / TELEFON BO'YICHA QIDIRUV
# ─────────────────────────────────────────────────────────────────────

async def search_by_username(username: str) -> list:
    import database as db_mod
    uname = username.lstrip('@').lower().strip()
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT DISTINCT user_id, MAX(first_name), MAX(last_name), username,
                   MAX(phone), MAX(bio), GROUP_CONCAT(DISTINCT group_link), MAX(added_date)
            FROM users_memory_bank
            WHERE LOWER(username)=?
            GROUP BY user_id
            LIMIT 20
            """,
            (uname,)
        ) as cur:
            rows = await cur.fetchall()
    results = []
    for r in rows:
        results.append({
            'user_id':    r[0],
            'first_name': r[1] or '',
            'last_name':  r[2] or '',
            'username':   r[3] or '',
            'phone':      r[4] or '',
            'bio':        r[5] or '',
            'groups':     [g.strip() for g in (r[6] or '').split(',') if g.strip()],
            'added_date': r[7] or '',
        })
    return results

async def search_by_phone(userbot, phone: str) -> dict:
    import database as db_mod
    from telethon.tl.functions.contacts import ImportContactsRequest, DeleteContactsRequest
    from telethon.tl.types import InputPhoneContact

    # Avval bazadan qidirish
    clean = phone.strip().replace(' ', '').replace('-', '')
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT DISTINCT user_id, MAX(first_name), MAX(last_name), MAX(username),
                   phone, MAX(bio), GROUP_CONCAT(DISTINCT group_link), MAX(added_date)
            FROM users_memory_bank
            WHERE phone LIKE ?
            GROUP BY user_id LIMIT 5
            """,
            (f'%{clean[-9:]}%',)
        ) as cur:
            db_rows = await cur.fetchall()

    db_result = None
    if db_rows:
        r = db_rows[0]
        db_result = {
            'user_id':    r[0],
            'first_name': r[1] or '',
            'last_name':  r[2] or '',
            'username':   r[3] or '',
            'phone':      r[4] or '',
            'bio':        r[5] or '',
            'groups':     [g.strip() for g in (r[6] or '').split(',') if g.strip()],
            'added_date': r[7] or '',
            'source':     'baza',
        }
        return db_result

    # Telegram ImportContacts orqali qidirish
    try:
        result = await userbot(ImportContactsRequest([
            InputPhoneContact(client_id=0, phone=clean, first_name='X', last_name='')
        ]))
        if result.users:
            u = result.users[0]
            try:
                await userbot(DeleteContactsRequest(id=[u]))
            except Exception:
                pass
            return {
                'user_id':    u.id,
                'first_name': u.first_name or '',
                'last_name':  u.last_name or '',
                'username':   u.username or '',
                'phone':      getattr(u, 'phone', phone) or phone,
                'bio':        '',
                'groups':     [],
                'added_date': '',
                'source':     'telegram',
            }
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────
# #12 FOYDALANUVCHI FAOLLIK TIMELINE
# ─────────────────────────────────────────────────────────────────────

async def get_user_timeline(user_id: int, days: int = 30) -> dict:
    import database as db_mod
    from datetime import timedelta
    since = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT DATE(msg_date) as day, COUNT(*) as cnt, GROUP_CONCAT(DISTINCT source)
            FROM messages_cache
            WHERE sender_id=? AND msg_date >= ?
            GROUP BY day ORDER BY day
            """,
            (user_id, since)
        ) as cur:
            rows = await cur.fetchall()
        async with db.execute(
            "SELECT COUNT(*) FROM messages_cache WHERE sender_id=?", (user_id,)
        ) as cur:
            total = (await cur.fetchone())[0]
        async with db.execute(
            "SELECT MIN(msg_date), MAX(msg_date) FROM messages_cache WHERE sender_id=?",
            (user_id,)
        ) as cur:
            span = await cur.fetchone()
    timeline = [{'date': r[0], 'count': r[1], 'sources': r[2]} for r in rows]
    return {'timeline': timeline, 'total': total,
            'first_msg': span[0] if span else None,
            'last_msg':  span[1] if span else None}


# ─────────────────────────────────────────────────────────────────────
# #15 CROSS-GROUP TAHLIL — UMUMIY A'ZOLAR
# ─────────────────────────────────────────────────────────────────────

async def get_common_members(group1: str, group2: str) -> list:
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT a.user_id, a.first_name, a.last_name, a.username, a.phone
            FROM users_memory_bank a
            JOIN users_memory_bank b ON a.user_id = b.user_id
            WHERE a.group_link=? AND b.group_link=?
            ORDER BY a.first_name
            """,
            (group1, group2)
        ) as cur:
            rows = await cur.fetchall()
    return [{'user_id': r[0], 'first_name': r[1] or '', 'last_name': r[2] or '',
             'username': r[3] or '', 'phone': r[4] or ''} for r in rows]


# ─────────────────────────────────────────────────────────────────────
# #16 KOORDINATSIYALANGAN XATTI-HARAKAT
# ─────────────────────────────────────────────────────────────────────

async def detect_coordinated_behavior(group_link: str, hours: int = 48) -> dict:
    import database as db_mod
    from datetime import timedelta

    # 1. Bir vaqtda qo'shilgan akkauntlar (added_date bo'yicha)
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT user_id, first_name, username, phone, bio, added_date
            FROM users_memory_bank
            WHERE group_link=? AND added_date IS NOT NULL
            ORDER BY added_date
            """,
            (group_link,)
        ) as cur:
            members = await cur.fetchall()

    # Vaqt oynasida guruhlash
    clusters = []
    if members:
        window_secs = hours * 3600
        i = 0
        while i < len(members):
            cluster = [members[i]]
            j = i + 1
            try:
                t0 = datetime.strptime(members[i][5][:16], '%Y-%m-%d %H:%M')
            except Exception:
                i += 1
                continue
            while j < len(members):
                try:
                    tj = datetime.strptime(members[j][5][:16], '%Y-%m-%d %H:%M')
                    if (tj - t0).total_seconds() <= window_secs:
                        cluster.append(members[j])
                        j += 1
                    else:
                        break
                except Exception:
                    j += 1
            if len(cluster) >= 5:
                clusters.append(cluster)
            i = j if j > i else i + 1

    # 2. Bio o'xshashligi — bo'sh bio li akkauntlar
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT COUNT(*) FROM users_memory_bank
            WHERE group_link=? AND (bio IS NULL OR bio='')
            """,
            (group_link,)
        ) as cur:
            no_bio_count = (await cur.fetchone())[0]
        async with db.execute(
            "SELECT COUNT(*) FROM users_memory_bank WHERE group_link=?", (group_link,)
        ) as cur:
            total_count = (await cur.fetchone())[0]
        async with db.execute(
            """
            SELECT COUNT(*) FROM users_memory_bank
            WHERE group_link=? AND (phone IS NULL OR phone='')
            """,
            (group_link,)
        ) as cur:
            no_phone_count = (await cur.fetchone())[0]

    no_bio_pct   = round(no_bio_count * 100 / max(total_count, 1))
    no_phone_pct = round(no_phone_count * 100 / max(total_count, 1))
    risk_score = 0
    if no_bio_pct > 70:   risk_score += 30
    if no_phone_pct > 80: risk_score += 20
    if clusters:          risk_score += min(len(clusters) * 10, 50)
    risk_score = min(risk_score, 100)

    return {
        'total':         total_count,
        'clusters':      clusters[:5],
        'cluster_count': len(clusters),
        'no_bio_pct':    no_bio_pct,
        'no_phone_pct':  no_phone_pct,
        'risk_score':    risk_score,
    }


# ─────────────────────────────────────────────────────────────────────
# #17 YOZUV USLUBI TAHLILI
# ─────────────────────────────────────────────────────────────────────

async def analyze_writing_style(user_id: int) -> dict:
    import database as db_mod
    import re as _re
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT text FROM messages_cache WHERE sender_id=? AND is_deleted=0 "
            "AND text IS NOT NULL LIMIT 500",
            (user_id,)
        ) as cur:
            rows = await cur.fetchall()
    if not rows:
        return None
    texts = [r[0] for r in rows if r[0]]
    total_chars  = sum(len(t) for t in texts)
    total_words  = sum(len(t.split()) for t in texts)
    avg_len      = round(total_chars / max(len(texts), 1))
    avg_words    = round(total_words / max(len(texts), 1))
    emoji_count  = sum(len(_re.findall(r'[\U0001F000-\U0001FFFF]', t)) for t in texts)
    question_cnt = sum(t.count('?') for t in texts)
    exclaim_cnt  = sum(t.count('!') for t in texts)
    caps_ratio   = round(sum(1 for t in texts for c in t if c.isupper()) /
                         max(total_chars, 1) * 100)
    # Top so'zlar
    all_words = []
    for t in texts:
        for w in t.lower().split():
            w = _re.sub(r'[^\w]', '', w)
            if len(w) > 3:
                all_words.append(w)
    word_freq = {}
    for w in all_words:
        word_freq[w] = word_freq.get(w, 0) + 1
    top_words = sorted(word_freq.items(), key=lambda x: -x[1])[:10]
    return {
        'msg_count':   len(texts),
        'avg_len':     avg_len,
        'avg_words':   avg_words,
        'emoji_ratio': round(emoji_count / max(len(texts), 1), 1),
        'question_pct': round(question_cnt * 100 / max(len(texts), 1)),
        'exclaim_pct':  round(exclaim_cnt * 100 / max(len(texts), 1)),
        'caps_ratio':   caps_ratio,
        'top_words':    top_words,
    }

async def compare_writing_styles(user_id1: int, user_id2: int) -> int:
    s1 = await analyze_writing_style(user_id1)
    s2 = await analyze_writing_style(user_id2)
    if not s1 or not s2:
        return 0
    score = 100
    diff_len   = abs(s1['avg_len']   - s2['avg_len'])
    diff_words = abs(s1['avg_words'] - s2['avg_words'])
    diff_emoji = abs(s1['emoji_ratio'] - s2['emoji_ratio'])
    diff_caps  = abs(s1['caps_ratio'] - s2['caps_ratio'])
    score -= min(diff_len // 5, 25)
    score -= min(diff_words * 3, 20)
    score -= min(int(diff_emoji * 10), 15)
    score -= min(diff_caps // 2, 15)
    # Umumiy so'zlar
    words1 = {w for w, _ in s1['top_words']}
    words2 = {w for w, _ in s2['top_words']}
    common = len(words1 & words2)
    score += common * 3
    return max(0, min(100, score))


# ─────────────────────────────────────────────────────────────────────
# #18 VAQT KORRELYATSIYASI
# ─────────────────────────────────────────────────────────────────────

async def get_temporal_correlations(min_overlap: int = 5, limit: int = 20) -> list:
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        # Har bir foydalanuvchi qaysi soatlarda faol
        async with db.execute(
            """
            SELECT sender_id, CAST(strftime('%H', msg_date) AS INTEGER) as hour, COUNT(*) as cnt
            FROM messages_cache
            WHERE sender_id IS NOT NULL AND is_deleted=0
            GROUP BY sender_id, hour
            HAVING cnt >= 2
            """,
        ) as cur:
            rows = await cur.fetchall()
    # {user_id: set of active hours}
    user_hours = {}
    for sender_id, hour, _ in rows:
        if sender_id not in user_hours:
            user_hours[sender_id] = set()
        user_hours[sender_id].add(hour)
    # Juftliklar
    users = list(user_hours.keys())
    pairs = []
    for i in range(len(users)):
        for j in range(i + 1, len(users)):
            overlap = len(user_hours[users[i]] & user_hours[users[j]])
            if overlap >= min_overlap:
                pairs.append((users[i], users[j], overlap))
    pairs.sort(key=lambda x: -x[2])
    return pairs[:limit]


# ─────────────────────────────────────────────────────────────────────
# #19 O'CHIRILGAN XABARLARNI KUZATISH
# ─────────────────────────────────────────────────────────────────────

async def get_deleted_messages(source: str, limit: int = 50) -> list:
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            """
            SELECT msg_id, sender_id, sender_name, sender_username, text, msg_date
            FROM messages_cache
            WHERE source=? AND is_deleted=1
            ORDER BY msg_date DESC LIMIT ?
            """,
            (source, limit)
        ) as cur:
            rows = await cur.fetchall()
    return [{'msg_id': r[0], 'sender_id': r[1], 'name': r[2] or '',
             'username': r[3] or '', 'text': (r[4] or '')[:300],
             'date': r[5] or '', 'link': _make_msg_link(source, r[0])} for r in rows]

async def mark_deleted_messages(source: str, current_msg_ids: set):
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT msg_id FROM messages_cache WHERE source=? AND is_deleted=0",
            (source,)
        ) as cur:
            cached_ids = {r[0] for r in await cur.fetchall()}
    deleted_ids = cached_ids - current_msg_ids
    if deleted_ids:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.executemany(
                "UPDATE messages_cache SET is_deleted=1 WHERE msg_id=? AND source=?",
                [(mid, source) for mid in deleted_ids]
            )
            await db.commit()
    return len(deleted_ids)


# ─────────────────────────────────────────────────────────────────────
# #20 AKKAUNT HAYOTI TAHLILI
# ─────────────────────────────────────────────────────────────────────

async def get_account_lifecycle(user_id: int) -> dict:
    import database as db_mod
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        # Oylik faollik
        async with db.execute(
            """
            SELECT strftime('%Y-%m', msg_date) as month, COUNT(*) as cnt
            FROM messages_cache
            WHERE sender_id=? AND is_deleted=0
            GROUP BY month ORDER BY month
            """,
            (user_id,)
        ) as cur:
            monthly = await cur.fetchall()
        # Soatlik faollik
        async with db.execute(
            """
            SELECT CAST(strftime('%H', msg_date) AS INTEGER) as hour, COUNT(*) as cnt
            FROM messages_cache
            WHERE sender_id=?
            GROUP BY hour ORDER BY hour
            """,
            (user_id,)
        ) as cur:
            hourly = await cur.fetchall()
        # O'zgarishlar soni
        async with db.execute(
            "SELECT COUNT(*) FROM user_change_log WHERE user_id=?", (user_id,)
        ) as cur:
            change_count = (await cur.fetchone())[0]
        # Birinchi ko'rinish
        async with db.execute(
            "SELECT MIN(added_date) FROM users_memory_bank WHERE user_id=?", (user_id,)
        ) as cur:
            first_seen = (await cur.fetchone())[0]
        # Birinchi va oxirgi xabar
        async with db.execute(
            "SELECT MIN(msg_date), MAX(msg_date) FROM messages_cache WHERE sender_id=? AND is_deleted=0",
            (user_id,)
        ) as cur:
            span = await cur.fetchone()
        first_msg = span[0] if span and span[0] else None
        last_msg  = span[1] if span and span[1] else None

    if not monthly:
        peak_hour = None
        peak_month = None
    else:
        peak_month = max(monthly, key=lambda x: x[1])[0] if monthly else None
        peak_hour  = max(hourly,  key=lambda x: x[1])[0] if hourly  else None

    return {
        'monthly':      [{'month': r[0], 'count': r[1]} for r in monthly],
        'hourly':       [{'hour': r[0],  'count': r[1]} for r in hourly],
        'peak_month':   peak_month,
        'peak_hour':    peak_hour,
        'change_count': change_count,
        'first_seen':   first_seen or '',
        'first_msg':    first_msg,
        'last_msg':     last_msg,
    }


# ─────────────────────────────────────────────────────────────────────
# #24 PROFIL RASMI
# ─────────────────────────────────────────────────────────────────────

async def get_profile_photo(userbot, user_id: int) -> str:
    try:
        from telethon.tl.functions.photos import GetUserPhotosRequest
        photos = await userbot(GetUserPhotosRequest(
            user_id=user_id, offset=0, max_id=0, limit=5
        ))
        if photos.photos:
            path = f"/tmp/tg_photo_{user_id}.jpg"
            await userbot.download_media(photos.photos[0], file=path)
            return path
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────
# #14 EVIDENCE PAKETI
# ─────────────────────────────────────────────────────────────────────

async def generate_evidence_report(user_id: int, userbot=None) -> str:
    import database as db_mod
    profile, messages = await lookup_user_by_id(user_id)
    changes          = await db_mod.get_user_change_log(user_id)
    lifecycle        = await get_account_lifecycle(user_id)
    style            = await analyze_writing_style(user_id)
    score, label, _  = calculate_trust_score(profile or {}, len(messages))

    lines = []
    sep = "=" * 60
    lines.append(sep)
    lines.append("TERGOV HISOBOTI — KIBER-STANSIYA OSINT PRO")
    lines.append(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"Profil ID: {user_id}")
    lines.append(sep)
    lines.append("")

    if profile:
        full_name = (profile['first_name'] + " " + profile['last_name']).strip() or "Nomsiz"
        lines.append("[ PROFIL MA'LUMOTI ]")
        lines.append(f"Ism:             {full_name}")
        lines.append(f"Username:        @{profile['username']}" if profile['username'] else "Username:        —")
        lines.append(f"Telefon:         {profile['phone']}" if profile['phone'] else "Telefon:         Yopiq")
        lines.append(f"Bio:             {profile['bio'][:200]}" if profile['bio'] else "Bio:             —")
        lines.append(f"Ochiq kanal:     {profile['open_channels']}" if profile['open_channels'] else "Ochiq kanal:     —")
        lines.append(f"Maxfiy kanal:    {profile['has_hidden']}" if profile['has_hidden'] and profile['has_hidden'] not in ('❌', '') else "Maxfiy kanal:    Yo'q")
        lines.append(f"Birinchi ko'rilgan: {profile['added_date']}")
        lines.append(f"So'nggi yangilanish: {profile['last_updated']}")
        lines.append("")
        lines.append(f"[ A'ZO GURUHLAR — {len(profile['groups'])} ta ]")
        for g in profile['groups']:
            lines.append(f"  • {g}")
        lines.append("")
    else:
        lines.append("[ PROFIL: bazada topilmadi ]")
        lines.append("")

    lines.append(f"[ TRUST SCORE: {score}/100 — {label} ]")
    lines.append("")

    if changes:
        lines.append(f"[ O'ZGARISHLAR TARIXI — {len(changes)} ta ]")
        for ch in changes:
            lines.append(f"  {ch[3]}  {ch[0]}: '{ch[1]}' → '{ch[2]}'")
        lines.append("")

    if lifecycle['monthly']:
        lines.append("[ OYLIK FAOLLIK ]")
        for m in lifecycle['monthly']:
            bar = "█" * min(m['count'], 40)
            lines.append(f"  {m['month']}: {bar} ({m['count']})")
        if lifecycle['peak_hour'] is not None:
            lines.append(f"  Eng faol soat: {lifecycle['peak_hour']}:00")
        lines.append("")

    if style:
        lines.append("[ YOZUV USLUBI ]")
        lines.append(f"  O'rtacha xabar uzunligi: {style['avg_len']} belgi, {style['avg_words']} so'z")
        lines.append(f"  Emoji/xabar: {style['emoji_ratio']}")
        lines.append(f"  Savol (%): {style['question_pct']}  Undov (%): {style['exclaim_pct']}")
        lines.append(f"  Katta harf (%): {style['caps_ratio']}")
        top = ", ".join(f"{w}({c})" for w, c in style['top_words'][:5])
        lines.append(f"  Tez-tez ishlatiladigan so'zlar: {top}")
        lines.append("")

    if messages:
        lines.append(f"[ XABARLAR — {len(messages)} ta ]")
        for i, m in enumerate(messages):
            lines.append(f"  [{i+1}] {m['date']} | {m['source_title']}")
            lines.append(f"      {m['text'][:150]}")
            if m.get('link'):
                lines.append(f"      Havola: {m['link']}")
        lines.append("")

    lines.append(sep)
    lines.append("HISOBOT TUGADI")
    lines.append(sep)

    report_text = "\n".join(lines)
    path = f"/tmp/evidence_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
    with open(path, 'w', encoding='utf-8') as f:
        f.write(report_text)
    return path


# ─────────────────────────────────────────────────────────────────────
# #22 TERGOV — HISOBOT
# ─────────────────────────────────────────────────────────────────────

async def generate_investigation_report(inv_id: int) -> str:
    import database as db_mod
    invs = await db_mod.get_investigations()
    inv  = next((i for i in invs if i[0] == inv_id), None)
    if not inv:
        return None
    targets = await db_mod.get_investigation_targets(inv_id)

    lines = []
    sep = "=" * 60
    lines.append(sep)
    lines.append(f"TERGOV: {inv[1]}")
    lines.append(f"Yaratilgan: {inv[4]}")
    if inv[3]:
        lines.append(f"Izohlar: {inv[3]}")
    lines.append(sep)
    lines.append("")

    user_ids = [int(t[2]) for t in targets if t[1] == 'user' and str(t[2]).lstrip('-').isdigit()]
    channel_ids = [t[2] for t in targets if t[1] == 'channel']

    for uid in user_ids:
        profile, messages = await lookup_user_by_id(uid)
        score, label, _   = calculate_trust_score(profile or {}, len(messages))
        lines.append(f"[ SHAXS ID: {uid} ]")
        if profile:
            full = (profile['first_name'] + " " + profile['last_name']).strip()
            lines.append(f"  Ism: {full}")
            if profile['username']:
                lines.append(f"  Username: @{profile['username']}")
            if profile['phone']:
                lines.append(f"  Telefon: {profile['phone']}")
            lines.append(f"  Guruhlar: {', '.join(profile['groups'][:5])}")
        lines.append(f"  Trust Score: {score}/100 — {label}")
        lines.append(f"  Xabarlar: {len(messages)} ta")
        # Umumiy guruhlar boshqa shaxslar bilan
        if len(user_ids) > 1:
            profile_groups = set(profile['groups']) if profile else set()
            for uid2 in user_ids:
                if uid2 == uid:
                    continue
                p2, _ = await lookup_user_by_id(uid2)
                if p2:
                    common = profile_groups & set(p2['groups'])
                    if common:
                        lines.append(f"  Umumiy guruh {uid2} bilan: {', '.join(list(common)[:3])}")
        lines.append("")

    for ch in channel_ids:
        lines.append(f"[ KANAL: {ch} ]")
        lines.append("")

    lines.append(sep)
    path = f"/tmp/investigation_{inv_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
    with open(path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
    return path


# ─────────────────────────────────────────────────────────────────────
# #23 TARMOQ TOPOLOGIYASI (HTML)
# ─────────────────────────────────────────────────────────────────────

async def generate_network_map(user_ids: list) -> str:
    import database as db_mod
    import json as _json

    nodes = []
    edges = []
    seen_nodes = set()
    group_nodes = set()

    for uid in user_ids[:30]:
        profile, _ = await lookup_user_by_id(uid)
        if not profile:
            continue
        label = (profile['first_name'] or str(uid))[:20]
        score, s_label, _ = calculate_trust_score(profile)
        color = '#e74c3c' if score < 40 else ('#f39c12' if score < 70 else '#2ecc71')
        if uid not in seen_nodes:
            nodes.append({'id': str(uid), 'label': label, 'color': color,
                          'title': f"ID:{uid} | Score:{score}"})
            seen_nodes.add(uid)
        for g in profile['groups']:
            if g not in group_nodes:
                nodes.append({'id': g, 'label': g[:20], 'color': '#3498db',
                              'shape': 'diamond', 'title': g})
                group_nodes.add(g)
            edges.append({'from': str(uid), 'to': g})

    nodes_json = _json.dumps(nodes)
    edges_json = _json.dumps(edges)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Network Map — OSINT Pro</title>
<script src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
<style>body{{margin:0;background:#1a1a2e}}
#net{{width:100%;height:100vh}}
</style></head><body>
<div id="net"></div>
<script>
var nodes=new vis.DataSet({nodes_json});
var edges=new vis.DataSet({edges_json});
var container=document.getElementById('net');
var options={{
  nodes:{{font:{{color:'#fff'}},size:20}},
  edges:{{color:'#555',arrows:'to'}},
  physics:{{stabilization:true}},
  background:{{color:'#1a1a2e'}}
}};
new vis.Network(container,{{nodes:nodes,edges:edges}},options);
</script></body></html>"""

    path = f"/tmp/network_map_{datetime.now().strftime('%Y%m%d_%H%M')}.html"
    with open(path, 'w', encoding='utf-8') as f:
        f.write(html)
    return path


# ─────────────────────────────────────────────────────────────────────
# ALERT TEKSHIRUVI — xabar cache ga qo'shilganda
# ─────────────────────────────────────────────────────────────────────

async def check_message_alerts(msg_id: int, source: str, sender_id: int,
                                sender_name: str, text: str, msg_date: str) -> list:
    import database as db_mod
    if not text:
        return []
    text_lower = text.lower()
    alerts = await db_mod.get_active_alerts()
    hits = []
    for alert_id, admin_id, keyword, target_groups in alerts:
        if keyword not in text_lower:
            continue
        if target_groups:
            allowed = [g.strip() for g in target_groups.split(',')]
            if not any(g in source for g in allowed):
                continue
        is_new = await db_mod.check_and_record_alert_hit(alert_id, msg_id, source, sender_id or 0)
        if is_new:
            hits.append({
                'admin_id':    admin_id,
                'keyword':     keyword,
                'msg_id':      msg_id,
                'source':      source,
                'sender_id':   sender_id,
                'sender_name': sender_name or '',
                'text':        text[:300],
                'date':        msg_date,
                'link':        _make_msg_link(source, msg_id),
            })
    return hits


# ═════════════════════════════════════════════════════════════════════
# TERGOV MA'LUMOTI — TO'LIQ PDF HISOBOT
# ═════════════════════════════════════════════════════════════════════

async def resolve_identifier_to_uid(userbot, identifier: str):
    """
    Telefon, @username yoki ID dan user_id ni aniqlash.
    Returns: (user_id, source_info)
    """
    import database as db_mod
    ident = identifier.strip()

    # Numeric ID (musbat)
    if ident.lstrip('-').isdigit():
        num = int(ident)
        if num > 0:
            return num, f"ID: {num}"
        else:
            return None, "Kanal ID (manfiy)"

    # @username
    if ident.startswith('@'):
        uname = ident[1:].lower()
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                "SELECT DISTINCT user_id FROM users_memory_bank WHERE LOWER(username)=? LIMIT 1",
                (uname,)
            ) as cur:
                row = await cur.fetchone()
        if row:
            return row[0], f"@{uname} (bazadan)"
        # Telegram dan qidirish
        try:
            entity = await userbot.get_entity(ident)
            return entity.id, f"@{uname} (Telegramdan)"
        except Exception:
            return None, "Username topilmadi"

    # Telefon raqam
    if ident.startswith('+') or (len(ident) >= 9 and ident[0].isdigit()):
        result = await search_by_phone(userbot, ident)
        if result:
            return result['user_id'], f"Telefon: {ident}"
        return None, "Telefon topilmadi"

    return None, "Noma'lum format"


async def generate_tergov_pdf(userbot, identifier: str) -> str:
    """
    Telefon / @username / ID bo'yicha to'liq PDF tergov hisoboti.
    """
    import database as db_mod

    BASE = os.path.dirname(os.path.abspath(__file__))
    FONT_PATH = os.path.join(BASE, 'DejaVuSans.ttf')

    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, Image, HRFlowable, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics

    def _valid_img(path):
        """Rasmni PIL bilan tekshiradi — yaroqsiz bo'lsa False qaytaradi."""
        try:
            from PIL import Image as PILImage
            if not path or not os.path.exists(path) or os.path.getsize(path) < 100:
                return False
            with PILImage.open(path) as im:
                im.verify()
            return True
        except Exception:
            return False
    from reportlab.pdfbase.ttfonts import TTFont

    # Font ro'yxatdan o'tkazish
    if 'DejaVu' not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont('DejaVu', FONT_PATH))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', FONT_PATH))

    # ── 1. Identifikatordan user_id ni aniqlash ──────────────────────
    user_id, id_source = await resolve_identifier_to_uid(userbot, identifier)
    if not user_id:
        return None, id_source

    # ── 2. Ma'lumotlar to'plash ──────────────────────────────────────
    profile, messages  = await lookup_user_by_id(user_id)
    changes            = await db_mod.get_user_change_log(user_id)
    lifecycle          = await get_account_lifecycle(user_id)
    style_data         = await analyze_writing_style(user_id)
    score, slabel, reasons = calculate_trust_score(profile or {}, len(messages))

    # Profil rasmlari
    photo_paths = []
    try:
        from telethon.tl.functions.photos import GetUserPhotosRequest
        photos_result = await userbot(GetUserPhotosRequest(
            user_id=user_id, offset=0, max_id=0, limit=10
        ))
        for i, ph in enumerate(photos_result.photos[:5]):
            p = f"/tmp/pdf_photo_{user_id}_{i}.jpg"
            try:
                await userbot.download_media(ph, file=p)
                if os.path.exists(p):
                    photo_paths.append(p)
            except Exception:
                pass
    except Exception:
        pass

    # Musiqa xabarlari
    music_msgs = []
    try:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                """
                SELECT msg_id, source, text, msg_date
                FROM messages_cache
                WHERE sender_id=? AND (
                    text LIKE '%🎵%' OR text LIKE '%🎶%' OR text LIKE '%mp3%'
                    OR text LIKE '%musiqa%' OR text LIKE '%music%' OR text LIKE '%audio%'
                    OR text LIKE '%song%' OR text LIKE '%track%'
                )
                ORDER BY msg_date DESC LIMIT 30
                """,
                (user_id,)
            ) as cur:
                music_msgs = await cur.fetchall()
    except Exception:
        pass

    # Guruh statistikasi
    group_stats = {}
    for m in messages:
        src = m['source_title'] or m['source']
        group_stats[src] = group_stats.get(src, 0) + 1

    # ── 3. PDF yaratish ──────────────────────────────────────────────
    pdf_path = f"/tmp/tergov_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    _now_str  = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── Sahifa footer (faqat pastki chiziq + qizil matn) ─────────────
    def _draw_page(canv, doc_obj):
        canv.saveState()
        pw, ph = A4
        canv.setStrokeColor(colors.HexColor('#dddddd'))
        canv.setLineWidth(0.5)
        canv.line(15*mm, 14*mm, pw - 15*mm, 14*mm)
        canv.restoreState()

    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        rightMargin=18*mm, leftMargin=18*mm,
        topMargin=18*mm, bottomMargin=22*mm,
        onFirstPage=_draw_page,
        onLaterPages=_draw_page,
    )

    styles = getSampleStyleSheet()

    def S(text, size=10, bold=False, color=colors.black, align='LEFT'):
        style = ParagraphStyle(
            name=f's{size}{bold}{align}',
            fontName='DejaVu',
            fontSize=size,
            textColor=color,
            alignment={'LEFT': 0, 'CENTER': 1, 'RIGHT': 2, 'JUSTIFY': 4}.get(align, 0),
            leading=size * 1.4,
            spaceAfter=2,
        )
        safe = str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return Paragraph(safe, style)

    def HR():
        return HRFlowable(width="100%", thickness=0.5, color=colors.grey, spaceAfter=4)

    import re as _re

    _tc_style = ParagraphStyle(
        name='tc',
        fontName='DejaVu',
        fontSize=8,
        leading=10,
        splitLongWords=True,
        wordWrap='LTR',
    )
    _tc_hdr = ParagraphStyle(
        name='tc_hdr',
        fontName='DejaVu',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        splitLongWords=True,
    )

    def TC(text, header=False):
        """Table Cell — matn katakda wrap bo'ladi."""
        safe = str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return Paragraph(safe, _tc_hdr if header else _tc_style)

    def _plain(cell):
        """TC yoki string dan oddiy matnni olish."""
        if hasattr(cell, 'text'):
            return _re.sub(r'<[^>]+>', '', cell.text)
        return str(cell) if cell is not None else ''

    def _auto_col_widths(data, total_w):
        """
        Har bir ustun kengligi: o'sha ustundagi eng uzun mazmun (50 belgi)
        asosida o'lchanadi. Katta ustunlar 40% dan oshmasligi ta'minlanadi.
        Jami aniq total_w ga to'g'rilanadi.
        """
        if not data:
            return []
        n = max(len(row) for row in data)
        PAD = 10  # pt — chap/o'ng padding

        natural = []
        for j in range(n):
            col_max = PAD * 2 + 6  # minimal
            for row in data:
                if j < len(row):
                    txt = _plain(row[j])
                    # Birinchi qator (ko'p qatorli matnda sarlavhadan kenglikni olish)
                    first = txt.split('\n')[0][:60]
                    w = pdfmetrics.stringWidth(first, 'DejaVu', 8) + PAD * 2
                    col_max = max(col_max, w)
            natural.append(col_max)

        # Har bir ustun 42% dan ko'p joy olmasin
        cap = total_w * 0.42
        capped = [min(w, cap) for w in natural]

        # Jami kenglikka proportsional moslashtirish
        total = sum(capped)
        scale = total_w / total if total > 0 else 1
        return [w * scale for w in capped]

    def TBL(data, col_widths=None, header_bg=colors.HexColor('#2c3e50')):
        widths = col_widths if col_widths is not None else _auto_col_widths(data, W)
        t = Table(data, colWidths=widths, repeatRows=1)
        style = TableStyle([
            ('FONTNAME',    (0,0), (-1,-1), 'DejaVu'),
            ('FONTSIZE',    (0,0), (-1,-1), 8),
            ('BACKGROUND',  (0,0), (-1,0),  header_bg),
            ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
            ('FONTSIZE',    (0,0), (-1,0),  9),
            ('ALIGN',       (0,0), (-1,-1), 'LEFT'),
            ('VALIGN',      (0,0), (-1,-1), 'TOP'),
            ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#f8f9fa'), colors.white]),
            ('TOPPADDING',  (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ])
        t.setStyle(style)
        return t

    story = []
    W = 174*mm  # usable width (A4 - 2*18mm margins)

    # ── MUQOVA ───────────────────────────────────────────────────────
    story.append(Spacer(1, 10*mm))
    # Profil rasmi (agar bor bo'lsa)
    if photo_paths and _valid_img(photo_paths[0]):
        try:
            img = Image(photo_paths[0], width=40*mm, height=40*mm)
            img.hAlign = 'CENTER'
            story.append(img)
            story.append(Spacer(1, 3*mm))
        except Exception:
            pass

    story.append(S("HISOBOT", size=22, align='CENTER', color=colors.HexColor('#2c3e50')))
    story.append(Spacer(1, 3*mm))
    story.append(HR())

    full_name = ""
    if profile:
        full_name = (profile['first_name'] + " " + profile['last_name']).strip() or "Nomsiz"
    story.append(S(f"Shaxs: {full_name or identifier}", size=16, align='CENTER', color=colors.HexColor('#e74c3c')))
    story.append(S(f"ID: {user_id}  |  {id_source}", size=10, align='CENTER', color=colors.HexColor('#555')))
    story.append(S(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                   f"Trust Score: {score}/100 — {slabel}", size=10, align='CENTER'))
    story.append(HR())
    story.append(Spacer(1, 5*mm))

    # ── 1. PROFIL MA'LUMOTI ───────────────────────────────────────────
    story.append(S("1. PROFIL MA'LUMOTI", size=13, color=colors.HexColor('#2c3e50')))
    story.append(HR())
    if profile:
        data = [[TC("Maydon", header=True), TC("Qiymat", header=True)]]
        rows = [
            (TC("To'liq ism"),          TC(full_name)),
            (TC("Username"),            TC(f"@{profile['username']}" if profile['username'] else "—")),
            (TC("Telefon"),             TC(profile['phone'] or "Yopiq")),
            (TC("Bio"),                 TC(profile['bio'][:200] if profile['bio'] else "—")),
            (TC("Ochiq kanal"),         TC(profile['open_channels'] or "—")),
            (TC("Maxfiy kanal"),        TC(profile['has_hidden'] if profile['has_hidden'] not in ('', '❌') else "Yo'q")),
            (TC("Birinchi ko'rilgan"),  TC(profile['added_date'] or "—")),
            (TC("So'nggi yangilangan"), TC(profile['last_updated'] or "—")),
            (TC("A'zo guruhlar"),       TC(f"{len(profile['groups'])} ta")),
            (TC("Yozilgan xabarlar"),   TC(f"{len(messages)} ta")),
        ]
        data.extend(rows)
        story.append(TBL(data))
    else:
        story.append(S("Profil bazada topilmadi.", color=colors.red))
    story.append(Spacer(1, 5*mm))

    # ── 2. TRUST SCORE ────────────────────────────────────────────────
    story.append(S("2. TRUST SCORE (ISHONCHLILIK BALI)", size=13, color=colors.HexColor('#2c3e50')))
    story.append(HR())
    score_color = colors.HexColor('#27ae60') if score >= 70 else (
        colors.HexColor('#f39c12') if score >= 40 else colors.HexColor('#e74c3c'))
    story.append(S(f"Ball: {score}/100 — {slabel}", size=14, color=score_color))
    if reasons:
        for r in reasons:
            story.append(S(f"  • {r}", size=9, color=colors.HexColor('#555')))
    story.append(Spacer(1, 5*mm))

    # ── 3. A'ZO GURUHLAR / KANALLAR ───────────────────────────────────
    if profile and profile['groups']:
        story.append(S("3. A'ZO GURUHLAR / KANALLAR", size=13, color=colors.HexColor('#2c3e50')))
        story.append(HR())
        story.append(S(f"Jami: {len(profile['groups'])} ta"))
        data = [[TC("#", header=True), TC("Guruh / Kanal", header=True), TC("Xabarlar soni", header=True)]]
        for i, g in enumerate(profile['groups'], 1):
            cnt = group_stats.get(g.split('t.me/')[-1].rstrip('/'), 0)
            if cnt == 0:
                cnt = group_stats.get(g, 0)
            data.append([TC(str(i)), TC(g), TC(str(cnt) if cnt else "—")])
        story.append(TBL(data))
        story.append(Spacer(1, 5*mm))

    # ── 4. XABAR STATISTIKASI ─────────────────────────────────────────
    story.append(S("4. XABAR STATISTIKASI", size=13, color=colors.HexColor('#2c3e50')))
    story.append(HR())
    if lifecycle['monthly']:
        story.append(S(f"Birinchi xabar: {lifecycle['first_msg'] or '—'}"))
        story.append(S(f"Oxirgi xabar:  {lifecycle['last_msg'] or '—'}"))
        if lifecycle['peak_month']:
            story.append(S(f"Eng faol oy:   {lifecycle['peak_month']}"))
        if lifecycle['peak_hour'] is not None:
            story.append(S(f"Eng faol soat: {lifecycle['peak_hour']}:00"))
        story.append(Spacer(1, 3*mm))
        # Oylik grafik (matn ko'rinishida)
        data = [[TC("Oy", header=True), TC("Xabarlar", header=True), TC("Grafik", header=True)]]
        for m in lifecycle['monthly']:
            bar = "█" * min(m['count'] // max(1, max(x['count'] for x in lifecycle['monthly']) // 20), 20)
            data.append([TC(m['month']), TC(str(m['count'])), TC(bar)])
        story.append(TBL(data))
    else:
        story.append(S("Kesh da xabar statistikasi yo'q."))
    story.append(Spacer(1, 5*mm))

    # ── 5. YOZUV USLUBI ───────────────────────────────────────────────
    if style_data:
        story.append(S("5. YOZUV USLUBI TAHLILI", size=13, color=colors.HexColor('#2c3e50')))
        story.append(HR())
        data = [[TC("Ko'rsatkich", header=True), TC("Qiymat", header=True)]]
        data.extend([
            (TC("Tahlil qilingan xabarlar"), TC(str(style_data['msg_count']))),
            (TC("O'rtacha uzunlik (belgi)"), TC(str(style_data['avg_len']))),
            (TC("O'rtacha so'zlar soni"),   TC(str(style_data['avg_words']))),
            (TC("Emoji / xabar"),           TC(str(style_data['emoji_ratio']))),
            (TC("Savol xabarlari (%)"),     TC(str(style_data['question_pct']))),
            (TC("Undov xabarlari (%)"),     TC(str(style_data['exclaim_pct']))),
            (TC("Katta harf (%)"),          TC(str(style_data['caps_ratio']))),
            (TC("Top so'zlar"),             TC(", ".join(f"{w}({c})" for w, c in style_data['top_words'][:8]))),
        ])
        story.append(TBL(data))
        story.append(Spacer(1, 5*mm))

    # ── 6. O'ZGARISHLAR TARIXI ────────────────────────────────────────
    if changes:
        story.append(S("6. O'ZGARISHLAR TARIXI", size=13, color=colors.HexColor('#2c3e50')))
        story.append(HR())
        field_names = {'first_name': 'Ism', 'last_name': 'Familiya',
                       'username': 'Username', 'phone': 'Telefon', 'bio': 'Bio'}
        data = [[TC("Sana", header=True), TC("Maydon", header=True),
                 TC("Eski", header=True), TC("Yangi", header=True)]]
        for ch in changes[:30]:
            data.append([TC(ch[3][:16]), TC(field_names.get(ch[0], ch[0])),
                         TC((ch[1] or '—')[:50]), TC((ch[2] or '—')[:50])])
        story.append(TBL(data))
        story.append(Spacer(1, 5*mm))

    # ── 7. MUSIQA XABARLARI ───────────────────────────────────────────
    if music_msgs:
        story.append(S("7. MUSIQA / MEDIA XABARLARI", size=13, color=colors.HexColor('#2c3e50')))
        story.append(HR())
        data = [[TC("Sana", header=True), TC("Manba", header=True), TC("Xabar", header=True)]]
        for mm_row in music_msgs[:20]:
            src_title = (mm_row[1] or '').split('t.me/')[-1].rstrip('/')[:30]
            data.append([
                TC((mm_row[3] or '')[:16]),
                TC(src_title),
                TC((mm_row[2] or '')[:100])
            ])
        story.append(TBL(data))
        story.append(Spacer(1, 5*mm))

    # ── 8. BARCHA XABARLAR ────────────────────────────────────────────
    story.append(PageBreak())
    story.append(S("8. BARCHA YOZILGAN XABARLAR", size=13, color=colors.HexColor('#2c3e50')))
    story.append(HR())
    if messages:
        story.append(S(f"Jami: {len(messages)} ta xabar — barchasi qo'shilgan"))
        story.append(Spacer(1, 3*mm))
        data = [[TC("#", header=True), TC("Sana", header=True), TC("Manba", header=True),
                 TC("Xabar matni", header=True), TC("Havola", header=True)]]
        for i, m in enumerate(messages, 1):
            link_val = m.get('link') or '—'
            if link_val != '—' and len(link_val) > 45:
                link_val = link_val[:45] + '…'
            data.append([
                TC(str(i)),
                TC((m['date'] or '')[:16]),
                TC((m['source_title'] or '')[:30]),
                TC((m['text'] or '')[:120]),
                TC(link_val),
            ])
        # Jami: 7+25+32+72+38 = 174mm
        story.append(TBL(data))
    else:
        story.append(S("Bu foydalanuvchi skanerlangan guruhlarda matnli xabar yozmagan yoki xabar tarixi mavjud emas."))
    story.append(Spacer(1, 5*mm))

    # ── 9. PROFIL RASMLARI ────────────────────────────────────────────
    if photo_paths:
        story.append(PageBreak())
        story.append(S("9. PROFIL RASMLARI", size=13, color=colors.HexColor('#2c3e50')))
        story.append(HR())
        story.append(S(f"Topilgan rasmlar: {len(photo_paths)} ta"))
        story.append(Spacer(1, 5*mm))
        # Rasmlarni 2 ustun qilib joylash
        img_row = []
        for i, ph in enumerate(photo_paths):
            if not _valid_img(ph):
                continue
            try:
                img = Image(ph, width=80*mm, height=80*mm)
                img_row.append(img)
                if len(img_row) == 2:
                    t = Table([img_row], colWidths=[90*mm, 90*mm])
                    story.append(t)
                    story.append(Spacer(1, 3*mm))
                    img_row = []
            except Exception:
                pass
        if img_row:
            t = Table([img_row + [""]], colWidths=[90*mm, 90*mm])
            story.append(t)

    # ── 10. YAKUNIY XULOSA ────────────────────────────────────────────
    story.append(PageBreak())
    story.append(S("10. YAKUNIY XULOSA", size=13, color=colors.HexColor('#2c3e50')))
    story.append(HR())
    data = [[TC("Ko'rsatkich", header=True), TC("Qiymat", header=True)]]
    data.extend([
        (TC("Tekshirilgan shaxs"),      TC(full_name or identifier)),
        (TC("Telegram ID"),             TC(str(user_id))),
        (TC("Trust Score"),             TC(f"{score}/100 — {slabel}")),
        (TC("Jami guruhlar"),           TC(f"{len(profile['groups']) if profile else 0} ta")),
        (TC("Jami xabarlar (keshda)"),  TC(f"{len(messages)} ta")),
        (TC("O'zgarishlar soni"),       TC(f"{len(changes)} ta")),
        (TC("Musiqa xabarlari"),        TC(f"{len(music_msgs)} ta")),
        (TC("Profil rasmlari"),         TC(f"{len(photo_paths)} ta")),
        (TC("Hisobot yaratildi"),       TC(datetime.now().strftime("%Y-%m-%d %H:%M"))),
    ])
    story.append(TBL(data))
    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor('#bbbbbb'),
                             spaceAfter=6, spaceBefore=4))
    # PDF ni saqlash
    doc.build(story)

    # Vaqtinchalik rasmlarni tozalash
    for ph in photo_paths:
        try:
            os.remove(ph)
        except Exception:
            pass

    return pdf_path, f"{full_name or identifier} — ID {user_id}"
