import logging
import re
import time

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("bot_errors.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# main.py
import os
import sys
import json
import random
import asyncio
from datetime import datetime
import openpyxl
import aiosqlite
from telethon import TelegramClient, events, Button
from telethon.tl.functions.users import GetFullUserRequest
from telethon.errors import FloodWaitError
from dotenv import load_dotenv
import database as db_mod
import tg_scrapers as engine
import music_scanner as music

# ── Clod-AI: AI xulosa moduli ────────────────────────────────────────────────
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    import modules.ai_summary as ai_mod
    _AI_ENABLED = bool(os.environ.get("ANTHROPIC_API_KEY"))
except ImportError:
    ai_mod = None
    _AI_ENABLED = False
# ─────────────────────────────────────────────────────────────────────────────

load_dotenv()

def _pm(event, group: int = 1) -> str:
    """pattern_match.group(N) yoki tv_input_handler fake event uchun message.text fallback."""
    pm = getattr(event, 'pattern_match', None)
    if pm is not None:
        v = pm.group(group)
        return v.strip() if isinstance(v, str) else (v or '')
    # Fake event: /cmd [subverb] arg1 arg2 ...
    txt = (getattr(getattr(event, 'message', None), 'text', '') or '').strip()
    parts = txt.split()
    # Command prefix: skip '/' words and lowercase subcommands (add, new, del, report...)
    skip = 0
    for p in parts:
        if p.startswith('/') or (p.isalpha() and p.islower() and len(p) <= 12):
            skip += 1
        else:
            break
    args = parts[skip:]
    if not args:
        return ''
    if group == 1:
        return args[0]
    # group 2+: join remaining so multiword values work
    return ' '.join(args[group - 1:]) if len(args) >= group else ''

API_ID         = int(os.getenv("API_ID"))
API_HASH       = os.getenv("API_HASH").strip()
BOT_TOKEN      = os.getenv("BOT_TOKEN").strip()

# ── 1-TUZATMA: SUPER_ADMIN_ID va ADMIN_IDS .env dan o'qiladi ────────
SUPER_ADMIN_ID = int(os.getenv("SUPER_ADMIN_ID", "0"))

# ADMIN_IDS=123456,789012,345678 ko'rinishida yoziladi
_admin_ids_raw = os.getenv("ADMIN_IDS", "")
EXTRA_ADMIN_IDS = set()
for _x in _admin_ids_raw.split(","):
    _x = _x.strip()
    if _x.isdigit():
        EXTRA_ADMIN_IDS.add(int(_x))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Doimiy Excel fayl — har safar yangi yaratmaydi, ustiga qo'shib boradi
EXCEL_PERSIST_PATH = os.path.join(BASE_DIR, "OSINT_Monitoring.xlsx")
EXCEL_IDS_PATH     = os.path.join(BASE_DIR, "excel_exported_ids.json")

def _load_excel_ids():
    if os.path.exists(EXCEL_IDS_PATH):
        with open(EXCEL_IDS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data.get("tc", [])), set(data.get("all", []))
    return set(), set()

def _save_excel_ids(tc_ids: set, all_ids: set):
    with open(EXCEL_IDS_PATH, "w", encoding="utf-8") as f:
        json.dump({"tc": list(tc_ids), "all": list(all_ids)}, f)

userbot = TelegramClient(os.path.join(BASE_DIR, 'userbot_session'), API_ID, API_HASH)
bot     = TelegramClient(os.path.join(BASE_DIR, 'bot_session'),     API_ID, API_HASH)

MAIN_KEYBOARD = [
    [Button.text("🔍 Skanerlash")],
    [Button.text("🔎 Kalit So'z Qidiruv")],
    [Button.text("📊 Kuzatuv Holati (Status)"),  Button.text("📥 Monitoringdan Fayl Olish")],
    [Button.text("📁 Arxiv / Savatcha"),         Button.text("🔒 Maxfiy Kanal Qo'shish")],
    [Button.text("👥 Adminlar Ro'yxati"),     Button.text("📋 Manbalar Ro'yxati")],
    [Button.text("🎼 Ko'p Musiqa Qidirish"),  Button.text("🔔 Kuzatiladigan Musiqalar")],
    [Button.text("💬 Kamentariya Xisoboti"),  Button.text("🔄 Botni Qayta Yuklash")],
    [Button.text("🤖 AI Xulosa")]
]

USER_STATES = {}

BUTTON_TEXTS = {
    "🔍 Skanerlash",
    "🔎 Kalit So'z Qidiruv",
    "📊 Kuzatuv Holati (Status)", "📥 Monitoringdan Fayl Olish",
    "📁 Arxiv / Savatcha", "🔒 Maxfiy Kanal Qo'shish",
    "👥 Adminlar Ro'yxati", "📋 Manbalar Ro'yxati",
    "🎼 Ko'p Musiqa Qidirish", "🔔 Kuzatiladigan Musiqalar",

    "💬 Kamentariya Xisoboti",
    "🔄 Botni Qayta Yuklash",
    "🤖 AI Xulosa",
}

# ─────────────────────────────────────────────────────────────────────
# ADMIN TEKSHIRUVI
# Ustuvorlik: SUPER_ADMIN_ID > EXTRA_ADMIN_IDS (.env) > DB adminlar
# ─────────────────────────────────────────────────────────────────────

async def is_admin(uid):
    if uid == SUPER_ADMIN_ID:
        return True
    if uid in EXTRA_ADMIN_IDS:
        return True
    return await db_mod.is_admin(uid, SUPER_ADMIN_ID)

# ─────────────────────────────────────────────────────────────────────
# /start
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern='/start'))
async def start_handler(event):
    if not await is_admin(event.sender_id):
        await event.respond("🔒 **Tizimga kirish taqiqlangan!**")
        return
    USER_STATES[event.sender_id] = None
    badge = "👑 Super Admin" if event.sender_id == SUPER_ADMIN_ID else "👤 Admin"
    await event.respond(
        f"🛰 **Kiber-Stansiya OSINT Pro ishga tushdi**\n"
        f"Sizning darajangiz: {badge}",
        buttons=MAIN_KEYBOARD
    )

# ─────────────────────────────────────────────────────────────────────
# ADMIN BOSHQARUVI (faqat Super Admin)
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/add_admin (\d+)'))
async def add_admin_handler(event):
    if event.sender_id != SUPER_ADMIN_ID:
        return
    new_id = int(_pm(event, 1))
    await db_mod.add_admin(new_id)
    await event.respond(f"✅ `{new_id}` adminlar ro'yxatiga qo'shildi.")

@bot.on(events.NewMessage(pattern=r'/del_admin (\d+)'))
async def del_admin_handler(event):
    if event.sender_id != SUPER_ADMIN_ID:
        return
    del_id = int(_pm(event, 1))
    await db_mod.remove_admin(del_id)
    await event.respond(f"❌ `{del_id}` o'chirildi.")

# ─────────────────────────────────────────────────────────────────────
# TUGMA: SKANERLASH (guruh / kanal / yopiq — avtomatik aniqlanadi)
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="🔍 Skanerlash"))
async def btn_scan(event):
    if not await is_admin(event.sender_id):
        return
    USER_STATES[event.sender_id] = 'waiting_auto_scan_link'
    await event.respond(
        "🔍 **Skanerlash linkini kiriting:**\n"
        "• Ochiq guruh: `@username` yoki `t.me/username`\n"
        "• Yopiq guruh: `t.me/+inviteHash`\n"
        "• Kanal: `@channel` yoki `t.me/channel`"
    )


# ─────────────────────────────────────────────────────────────────────
# TUGMA: KALIT SO'Z QIDIRUV
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="🔎 Kalit So'z Qidiruv"))
async def btn_keyword_search(event):
    if not await is_admin(event.sender_id):
        return
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT COUNT(DISTINCT group_link) FROM users_memory_bank "
            "WHERE group_link IS NOT NULL AND group_link != ''"
        ) as cur:
            source_count = (await cur.fetchone())[0]
    USER_STATES[event.sender_id] = 'waiting_keyword_word'
    await event.respond(
        f"🔎 **Kalit so'zni kiriting:**\n"
        f"_Bazada `{source_count}` ta manba mavjud — hammasidan qidiriladi_\n\n"
        "Bir nechta so'z bo'lsa vergul bilan ajrating:\n"
        "`sotaman, telefon, uy`"
    )

# ─────────────────────────────────────────────────────────────────────
# TUGMA: KUZATUV HOLATI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r"📊 Kuzatuv Holati \(Status\)"))
async def btn_status(event):
    if not await is_admin(event.sender_id):
        return

    # Asosiy statistika
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as db:
        total   = (await (await db.execute("SELECT COUNT(DISTINCT user_id) FROM users_memory_bank")).fetchone())[0]
        pending = (await (await db.execute("SELECT COUNT(*) FROM hidden_channel_knocker WHERE status='pending'")).fetchone())[0]
        joined  = (await (await db.execute("SELECT COUNT(*) FROM hidden_channel_knocker WHERE status='joined'")).fetchone())[0]
        running = (await (await db.execute("SELECT COUNT(*) FROM scan_resume WHERE status='running'")).fetchone())[0]
        # Barcha kanallar soni (users_memory_bank + hidden_channel_knocker joined)
        all_sources = (await (await db.execute(
            "SELECT COUNT(*) FROM ("
            "  SELECT DISTINCT group_link as src FROM users_memory_bank WHERE group_link IS NOT NULL AND group_link != ''"
            "  UNION"
            "  SELECT channel_id as src FROM hidden_channel_knocker WHERE status='joined'"
            ")"
        )).fetchone())[0]

    # Musiqa statistika
    total_fp, total_ch = await music.get_stats()

    # Profillardan nechta musiqa skanerlandi
    import aiosqlite as _aio
    async with _aio.connect(music.MUSIC_DB, timeout=10) as mdb:
        profile_scanned = (await (await mdb.execute(
            "SELECT COUNT(DISTINCT channel_id) FROM music_fingerprints "
            "WHERE file_name LIKE 'profile_%'"
        )).fetchone())[0]

        channel_scanned = (await (await mdb.execute(
            "SELECT COUNT(DISTINCT channel_id) FROM music_fingerprints "
            "WHERE file_name LIKE 'msg_%'"
        )).fetchone())[0]

        channel_progress = 0

    # music_channel_progress asosiy DBda saqlanadi (music DBda emas)
    try:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=10) as _mdb2:
            await _mdb2.execute(
                "CREATE TABLE IF NOT EXISTS music_channel_progress "
                "(channel_id TEXT PRIMARY KEY, last_msg_id INTEGER)"
            )
            await _mdb2.commit()
            channel_progress = (await (await _mdb2.execute(
                "SELECT COUNT(*) FROM music_channel_progress"
            )).fetchone())[0]
    except Exception:
        channel_progress = 0

    # Qolgan profillar (musiqasi tekshirilmagan)
    profiles_left   = total - profile_scanned
    channels_left   = all_sources - channel_scanned

    # Navbat holati
    queue_info = ""
    if _SCAN_QUEUE and not _SCAN_QUEUE.empty():
        queue_info = f"\n📋 Navbatda: `{_SCAN_QUEUE.qsize()}` ta topshiriq"

    await event.respond(
        f"🛰 **KIBER-STANSIYA STATUSI:**\n\n"
        f"⚙️ **{engine.resource_status()}**\n"
        f"👥 Monitoringdagi profillar: `{total}` ta\n"
        f"📥 Kutilayotgan maxfiy kanallar: `{pending}` ta\n"
        f"🔓 Kirilgan maxfiy kanallar: `{joined}` ta\n"
        f"📨 Bugungi so\'rovnomalar: `{engine._daily_knock_count}/{engine.MAX_DAILY_KNOCKS}` ta\n"
        f"⚡️ Skaner: **{'⏸ PAUZADA' if engine.SCANNER_PAUSED else '▶️ ISHLAMOQDA'}**"
        + queue_info +
        f"\n\n🎵 **MUSIQA SKANERLASH:**\n"
        f"👤 Profil musiqasi skanerlandi: `{profile_scanned}` ta | Qoldi: `{profiles_left}` ta\n"
        f"📢 Kanal musiqasi skanerlandi: `{channel_scanned}` ta | Qoldi: `{max(0, channels_left)}` ta\n"
        f"🎶 Jami fingerprint: `{total_fp}` ta"
        + (f"\n♻️ Davom ettiriladigan: `{running}` ta" if running else "")
    )

# ─────────────────────────────────────────────────────────────────────
# TUGMA: MONITORING DAN FAYL OLISH
# ── 3-TUZATMA: has_hidden da invite link bo'lsa Maxfiy Kanal ustuniga
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="📥 Monitoringdan Fayl Olish"))
async def btn_fayl(event):
    if not await is_admin(event.sender_id):
        return
    status_msg = await event.respond("⏳ Bazadan yangi ma'lumotlar tekshirilmoqda...")

    # ── Avvalgi eksport ID larini yuklash ────────────────────────────
    tc_exported, all_exported = _load_excel_ids()
    file_exists = os.path.exists(EXCEL_PERSIST_PATH)

    # ── Excel faylni yuklash yoki yangi yaratish ─────────────────────
    _loop = asyncio.get_event_loop()
    if file_exists:
        wb = await _loop.run_in_executor(None, openpyxl.load_workbook, EXCEL_PERSIST_PATH)
        sh  = wb["Target Channels"] if "Target Channels" in wb.sheetnames else wb.active
        sh2 = wb["Лист1"] if "Лист1" in wb.sheetnames else wb.create_sheet("Лист1")
        tc_start_row  = sh.max_row  # keyingi satr davom etadi
        all_start_row = sh2.max_row
    else:
        wb  = openpyxl.Workbook()
        sh  = wb.active
        sh.title = "Target Channels"
        sh.append([
            "№", "Telegram ID", "Kanal ID", "Ism", "Username", "Telefon",
            "Bio Linklar", "Shaxsiy Kanal Linki", "Maxfiy Kanal Linki",
            "Bio", "Manba Guruh", "O'zgargan Sana"
        ])
        sh2 = wb.create_sheet(title="Лист1")
        sh2.append([
            "№", "Telegram ID", "Kanal ID", "Ism", "Familya", "Username",
            "Telefon", "Bio", "Ochiq Kanallar", "Maxfiy Kanal",
            "Manba Guruh/Kanal", "Qo'shilgan Sana", "Yangilangan Sana"
        ])
        tc_start_row  = 1
        all_start_row = 1

    async with aiosqlite.connect(db_mod.DB_NAME, timeout=15) as db:
        async with db.execute(
            "SELECT user_id, first_name, username, phone, open_channels, "
            "has_hidden, bio, group_link, MAX(last_updated) as last_updated "
            "FROM users_memory_bank "
            "WHERE (has_hidden IS NOT NULL AND has_hidden != '' AND has_hidden != '❌') "
            "   OR (open_channels IS NOT NULL AND open_channels != '' AND open_channels != 'Yo''q') "
            "GROUP BY user_id "
            "ORDER BY last_updated DESC"
        ) as cur:
            rows = await cur.fetchall()

    # Yangi foydalanuvchilar — allaqachon eksport qilinmaganlar
    seen_uids = set()
    new_rows = []
    for row in rows:
        uid = row[0]
        if uid not in seen_uids and uid not in tc_exported:
            seen_uids.add(uid)
            new_rows.append(row)
    rows = new_rows

    # Yangi ma'lumot yo'q — faylni qayta yuborish
    if not rows and file_exists:
        await status_msg.edit("✅ Yangi ma'lumot yo'q — fayl allaqachon yangilangan.")
        await bot.send_file(
            event.sender_id, EXCEL_PERSIST_PATH,
            caption=f"📊 **Monitoring fayli** (yangilanmadi)\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        return
    elif not rows and not file_exists:
        await status_msg.edit("📭 Hozircha kanal aniqlangan profil yo'q. Avval guruh skanerlang.")
        return

    await status_msg.edit(f"⏳ {len(rows)} ta yangi profil qo'shilmoqda...")

    # Barcha kanal ID larni olish (N+1 dan qochish)
    uid_list = [r[0] for r in rows]
    uid_to_kanal = {}

    # 1. hidden_channel_knocker — maxfiy kanallar (joined bo'lganda numeric_id saqlanadi)
    if uid_list:
        placeholders = ",".join("?" * len(uid_list))
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                f"SELECT creator_id, channel_id, numeric_id FROM hidden_channel_knocker "
                f"WHERE creator_id IN ({placeholders})",
                uid_list
            ) as cur:
                for _creator_id, _channel_id, _numeric_id in await cur.fetchall():
                    if _creator_id not in uid_to_kanal:
                        if _numeric_id and str(_numeric_id).strip():
                            uid_to_kanal[_creator_id] = str(_numeric_id).strip()
                        else:
                            _raw = str(_channel_id).strip()
                            if _raw.lstrip('-').isdigit():
                                _cid = _raw.lstrip('-')
                                uid_to_kanal[_creator_id] = (
                                    f"-100{_cid}" if not _raw.startswith('-100') else _raw
                                )
                            else:
                                uid_to_kanal[_creator_id] = _raw

    # 2. resolved_channel_ids — shaxsiy/ochiq kanallar
    #    (skanerlash paytida resolve_personal_channel tomonidan saqlanadi)
    all_links = set()
    for _uid, _fn, _un, _ph, _oc, _hh, _bio, _gl, _lu in rows:
        if _oc and _oc not in ("", "Yo'q"):
            for lnk in _oc.split(','):
                lnk = lnk.strip()
                if lnk:
                    all_links.add(lnk)
        if _hh and _hh not in ("", "❌") and "Maxfiy" not in (_hh or ""):
            for lnk in _hh.split(','):
                lnk = lnk.strip()
                if lnk:
                    all_links.add(lnk)

    link_to_numid = {}
    if all_links:
        _links_list = list(all_links)
        _ph = ",".join("?" * len(_links_list))
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                f"SELECT channel_link, numeric_id FROM resolved_channel_ids "
                f"WHERE channel_link IN ({_ph})",
                _links_list
            ) as cur:
                for _cl, _nid in await cur.fetchall():
                    if _nid:
                        link_to_numid[_cl] = _nid

    # uid → kanal_id: knocker dan topilmasa, open_channels/has_hidden linkini keshdan qidirish
    for _uid, _fn, _un, _ph, _oc, _hh, _bio, _gl, _lu in rows:
        if _uid in uid_to_kanal:
            continue
        # open_channels dan birinchi linki tekshirish
        if _oc and _oc not in ("", "Yo'q"):
            for lnk in _oc.split(','):
                lnk = lnk.strip()
                if lnk and lnk in link_to_numid:
                    uid_to_kanal[_uid] = link_to_numid[lnk]
                    break
        if _uid in uid_to_kanal:
            continue
        # has_hidden dan tekshirish
        if _hh and _hh not in ("", "❌") and "Maxfiy" not in (_hh or ""):
            for lnk in _hh.split(','):
                lnk = lnk.strip()
                if lnk and lnk in link_to_numid:
                    uid_to_kanal[_uid] = link_to_numid[lnk]
                    break

    for i, (uid, fname, uname, phone, open_ch, has_hidden, bio, grp_link, last_updated) in enumerate(rows, tc_start_row):
        bio_links_list = engine.extract_bio_links(bio) if bio else []
        bio_links_str  = ", ".join(bio_links_list) if bio_links_list else ""
        phone_str      = ("+" + phone) if phone else ""

        shaxsiy = ""
        maxfiy  = ""
        if has_hidden:
            if "t.me/+" in has_hidden or "t.me/joinchat/" in has_hidden:
                maxfiy = has_hidden
            elif has_hidden.startswith("http"):
                shaxsiy = has_hidden
            elif "Maxfiy" in has_hidden:
                inv_links = engine.extract_invite_links(bio) if bio else []
                maxfiy = ", ".join(inv_links) if inv_links else "🔒 Maxfiy (link yo'q)"
        if not maxfiy and bio:
            inv_links = engine.extract_invite_links(bio)
            if inv_links:
                maxfiy = ", ".join(inv_links)

        kanal_id = uid_to_kanal.get(uid, "")

        sh.append([
            i, uid, kanal_id, fname or "", uname or "", phone_str,
            bio_links_str, shaxsiy, maxfiy,
            bio or "", grp_link or "", last_updated or ""
        ])

    # ── ЛИСТ1: faqat yangi skanerlangan foydalanuvchilar ─────────────
    await status_msg.edit("⏳ Лист1 ma'lumotlari yuklanmoqda...")

    async with aiosqlite.connect(db_mod.DB_NAME, timeout=60) as db:
        async with db.execute(
            """
            SELECT
                user_id,
                MAX(first_name)   AS first_name,
                MAX(last_name)    AS last_name,
                MAX(username)     AS username,
                MAX(phone)        AS phone,
                MAX(bio)          AS bio,
                MAX(open_channels) AS open_channels,
                MAX(has_hidden)   AS has_hidden,
                GROUP_CONCAT(DISTINCT group_link) AS all_groups,
                MIN(added_date)   AS added_date,
                MAX(last_updated) AS last_updated
            FROM users_memory_bank
            GROUP BY user_id
            ORDER BY last_updated DESC
            """
        ) as cur:
            all_users_raw = await cur.fetchall()

    # Faqat yangi foydalanuvchilar
    new_all_users = [u for u in all_users_raw if u[0] not in all_exported]
    await status_msg.edit(f"⏳ Лист1 yozilmoqda... ({len(new_all_users)} ta yangi profil)")

    for j, (uid2, fn2, ln2, un2, ph2, bio2,
            oc2, hh2, grp2, added2, upd2) in enumerate(new_all_users, all_start_row):
        ph2_str  = ("+" + ph2) if ph2 else ""
        un2_str  = ("@" + un2) if un2 else ""
        kanal_id2 = uid_to_kanal.get(uid2, "")
        maxfiy2  = ""
        shaxsiy2 = ""
        if hh2 and hh2 not in ("", "❌"):
            if "t.me/+" in hh2 or "t.me/joinchat/" in hh2:
                maxfiy2 = hh2
            elif hh2.startswith("http"):
                shaxsiy2 = hh2
            elif "Maxfiy" in hh2:
                maxfiy2 = "🔒 Maxfiy"
        sh2.append([
            j, uid2, kanal_id2, fn2 or "", ln2 or "", un2_str, ph2_str,
            bio2 or "", shaxsiy2 or oc2 or "", maxfiy2,
            grp2 or "", added2 or "", upd2 or ""
        ])

    _loop = asyncio.get_event_loop()
    await status_msg.edit("⏳ Excel formatlash...")
    await _loop.run_in_executor(None, engine.apply_excel_styles, sh, sh.max_row - 1)
    # Лист1 katta bo'lsa yengil styling — per-cell loop yo'q, tezroq
    if sh2.max_row > 1000:
        await _loop.run_in_executor(None, engine.apply_excel_styles_light, sh2)
    else:
        await _loop.run_in_executor(None, engine.apply_excel_styles, sh2, sh2.max_row - 1)
    await status_msg.edit("⏳ Fayl saqlanmoqda...")

    # Faylni saqlash
    await _loop.run_in_executor(None, wb.save, EXCEL_PERSIST_PATH)

    # Eksport qilingan ID larni yangilash
    tc_exported.update(r[0] for r in rows)
    all_exported.update(u[0] for u in new_all_users)
    _save_excel_ids(tc_exported, all_exported)

    # Umumiy statistika
    total_tc  = sh.max_row - 1
    total_all = sh2.max_row - 1

    await bot.send_file(
        event.sender_id, EXCEL_PERSIST_PATH,
        caption=(
            f"📊 **Monitoring fayli** (yangilandi)\n\n"
            f"📋 **Target Channels** (1-varaq): `{total_tc}` ta — kanali bor profillar\n"
            f"   ➕ Yangi qo'shildi: `{len(rows)}` ta\n"
            f"📋 **Лист1** (2-varaq): `{total_all}` ta — barcha skanerlangan\n"
            f"   ➕ Yangi qo'shildi: `{len(new_all_users)}` ta\n"
            f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
    )
    await status_msg.delete()

# ─────────────────────────────────────────────────────────────────────
# TUGMA: ARXIV
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="📁 Arxiv / Savatcha"))
async def btn_arxiv(event):
    if not await is_admin(event.sender_id):
        return
    USER_STATES[event.sender_id] = 'waiting_archive_date'
    await event.respond("📅 **Arxiv sanasini kiriting (YYYY-MM-DD):**")

# ─────────────────────────────────────────────────────────────────────
# TUGMA: MAXFIY KANAL QO'SHISH
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="🔒 Maxfiy Kanal Qo'shish"))
async def btn_maxfiy(event):
    if not await is_admin(event.sender_id):
        return
    USER_STATES[event.sender_id] = 'waiting_custom_channel'
    await event.respond("🔗 **24 soatlik monitoringga qo'shiladigan maxfiy kanal linkini yuboring:**")

# ─────────────────────────────────────────────────────────────────────
# TUGMA: ADMINLAR RO'YXATI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="👥 Adminlar Ro'yxati"))
async def btn_admins(event):
    if not await is_admin(event.sender_id):
        return
    admins = await db_mod.get_all_admins()
    lines  = [f"👑 **Super Admin:** `{SUPER_ADMIN_ID}`\n"]
    if EXTRA_ADMIN_IDS:
        lines.append("📋 **.env adminlar:**")
        for aid in sorted(EXTRA_ADMIN_IDS):
            lines.append(f"  • `{aid}` (statik)")
    if admins:
        lines.append("📋 **Qo'shilgan adminlar:**")
        for i, (aid,) in enumerate(admins, 1):
            lines.append(f"  {i}. `{aid}`")
    if not EXTRA_ADMIN_IDS and not admins:
        lines.append("📋 Hali qo'shimcha admin qo'shilmagan.")
    if event.sender_id == SUPER_ADMIN_ID:
        lines.append(
            "\n\n📌 **Buyruqlar:**\n"
            "`/add_admin [ID]` — admin qo'shish\n"
            "`/del_admin [ID]` — adminni o'chirish\n\n"
            "💡 Yoki `.env` faylida `ADMIN_IDS=ID1,ID2` qilib yozing"
        )
    await event.respond("\n".join(lines))

# ─────────────────────────────────────────────────────────────────────
# TUGMA: MANBALAR RO'YXATI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="📋 Manbalar Ro'yxati"))
async def btn_sources(event):
    if not await is_admin(event.sender_id):
        return
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT group_link, COUNT(DISTINCT user_id) as cnt "
            "FROM users_memory_bank "
            "WHERE group_link IS NOT NULL AND group_link != '' "
            "GROUP BY group_link ORDER BY cnt DESC"
        ) as cur:
            sources = await cur.fetchall()

    if not sources:
        await event.respond("📭 Hali hech qanday manba skanerlanbagan.")
        return

    lines = [f"📋 **Skanerlangan manbalar:** `{len(sources)}` ta\n"]
    total_users = 0
    for i, (link, cnt) in enumerate(sources, 1):
        lines.append(f"  {i}. `{link}` — `{cnt}` ta profil")
        total_users += cnt

    lines.append(f"\n👥 **Jami profillar:** `{total_users}` ta")

    # Telegram 4096 belgi limitiga bo'lish
    text = "\n".join(lines)
    if len(text) > 3800:
        chunks = []
        chunk  = ""
        for line in lines:
            if len(chunk) + len(line) + 1 > 3800:
                chunks.append(chunk)
                chunk = line
            else:
                chunk += ("\n" if chunk else "") + line
        if chunk:
            chunks.append(chunk)
        for ch in chunks:
            await event.respond(ch)
    else:
        await event.respond(text)


# ─────────────────────────────────────────────────────────────────────
# INLINE CALLBACK: KALIT SO'Z SANA TANLASH
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.CallbackQuery(pattern=b"kw_date_"))
async def keyword_date_callback(event):
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict) or state.get('state') != 'waiting_keyword_date':
        await event.answer("⚠️ Sessiya tugagan. Qaytadan boshlang.")
        return

    keyword  = state['keyword']
    data     = event.data.decode()
    days_map = {
        'kw_date_1':   (1,   "Oxirgi 1 kun"),
        'kw_date_7':   (7,   "Oxirgi 7 kun"),
        'kw_date_30':  (30,  "Oxirgi 1 oy"),
        'kw_date_365': (365, "Oxirgi 1 yil"),
        'kw_date_all': (None,"Hammasi"),
    }
    days, label = days_map.get(data, (None, "Hammasi"))

    USER_STATES[event.sender_id] = None
    await event.answer(f"✅ {label} tanlandi")
    await event.edit(
        f"🔎 **Qidirilmoqda:** `{keyword}`\n"
        f"📅 Davr: **{label}**\n"
        f"📍 Bazadagi barcha manbalardan qidirilmoqda..."
    )
    status = await event.get_message()
    if engine._SCAN_COUNT > 0:
        await event.respond(
            f"⚠️ Hozir skanerlash ishlayapti!\n"
            f"Navbatga qo\'shildi — tugagach boshlanadi."
        )
        _SCAN_QUEUE.put_nowait(('keyword', event.sender_id, (keyword, days, status)))
        return
    asyncio.create_task(run_keyword_search(event.sender_id, keyword, status, days))


# ─────────────────────────────────────────────────────────────────────
# GLOBAL INPUT PROCESSOR
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage)
async def global_input_processor(event):
    if not await is_admin(event.sender_id):
        return
    if event.text and (event.text.strip() in BUTTON_TEXTS or event.text.startswith('/')):
        return
    state = USER_STATES.get(event.sender_id)
    if not state:
        return

    # ── 1. Avtomatik skanerlash (guruh / kanal / yopiq) ─────────────
    if state == 'waiting_auto_scan_link':
        if not event.text:
            return
        ok, target = engine.validate_target(event.text)
        if not ok:
            await event.respond("⚠️ Noto'g'ri format. @username, t.me/link yoki guruh ID kiriting.")
            return
        USER_STATES[event.sender_id] = None
        # Invite link (yopiq guruh) aniqlash
        is_invite = bool(re.search(r't\.me/\+|joinchat', target))
        if engine._SCAN_COUNT > 0:
            await event.respond(
                f"⚠️ Hozir skanerlash ishlayapti!\n"
                f"Navbatga qo\'shildi — tugagach boshlanadi."
            )
            _SCAN_QUEUE.put_nowait(('msg' if is_invite else 'auto', event.sender_id, target))
            return
        if is_invite:
            status = await event.respond("⏳ Yopiq guruh xabarlari o'qilmoqda...")
            fname  = f"Msg_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
            fpath  = os.path.join(BASE_DIR, fname)
            asyncio.create_task(run_msg_scan(event.sender_id, target, fpath, status))
        else:
            status = await event.respond("⏳ Skanerlash boshlandi, tur aniqlanmoqda...")
            asyncio.create_task(run_auto_scan(event.sender_id, target, status))

    # ── 1d. Kalit so'z qidiruv — so'z kiritish ──────────────────────
    elif state == 'waiting_keyword_word':
        if not event.text:
            return
        input_text = event.text.strip()

        # ── ID qidiruvi: sof raqam kiritilgan bo'lsa ────────────────
        if input_text.lstrip('-').isdigit():
            USER_STATES.pop(event.sender_id, None)
            status_msg = await event.respond("🔍 ID qidirilmoqda...")
            asyncio.create_task(run_id_search(event.sender_id, input_text, status_msg))
            return

        ok, keyword = engine.validate_keyword(input_text)
        if not ok:
            await event.respond("⚠️ Kalit so'z 1-200 belgi bo'lishi kerak.")
            return
        USER_STATES[event.sender_id] = {'state': 'waiting_keyword_date', 'keyword': keyword}
        await event.respond(
            f"✅ Kalit so'z: `{keyword}`\n\n"
            "📅 **Qidiruv davrini tanlang:**",
            buttons=[
                [Button.inline("📅 Oxirgi 1 kun",   data="kw_date_1")],
                [Button.inline("📅 Oxirgi 7 kun",   data="kw_date_7")],
                [Button.inline("📅 Oxirgi 1 oy",    data="kw_date_30")],
                [Button.inline("📅 Oxirgi 1 yil",   data="kw_date_365")],
                [Button.inline("📅 Hammasi",         data="kw_date_all")],
            ]
        )


    # ── 3. Arxiv sanasi ──────────────────────────────────────────────
    elif state == 'waiting_archive_date':
        if not event.text:
            return
        USER_STATES[event.sender_id] = None
        sana = event.text.strip()
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                "SELECT file_name, file_path FROM archive_bin WHERE created_date=?", (sana,)
            ) as cur:
                files = await cur.fetchall()
        if files:
            for fn, fp in files:
                if os.path.exists(fp):
                    await bot.send_file(event.sender_id, fp, caption=f"📁 `{fn}`")
        else:
            await event.respond("📭 Bu sanada fayl topilmadi.")

    # ── 4. Maxfiy kanal qo'shish ─────────────────────────────────────
    elif state == 'waiting_tergov_id':
        if not event.text:
            return
        ident = event.text.strip()
        if not ident:
            return
        USER_STATES.pop(event.sender_id, None)
        status_msg = await event.respond(
            f"🕵️ `{ident}` bo'yicha ma'lumot to'planmoqda...\n"
            "📸 Rasmlar, xabarlar, statistika yuklanmoqda..."
        )
        asyncio.create_task(run_tergov_pdf(event.sender_id, ident, status_msg))
        return

    elif state == 'waiting_custom_channel':
        if not event.text:
            return
        # Faqat invite link yoki oddiy link qabul qilish
        ch = event.text.strip()
        _is_valid_invite = (
            ch.startswith("https://t.me/+") or
            ch.startswith("https://t.me/joinchat/") or
            ch.startswith("t.me/+") or
            ch.startswith("t.me/joinchat/")
        )
        if not _is_valid_invite or len(ch) > 512:
            await event.respond("⚠️ Faqat invite link qabul qilinadi: https://t.me/+XXXX")
            return
        USER_STATES[event.sender_id] = None
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT OR IGNORE INTO hidden_channel_knocker "
                "(channel_id, creator_id, source_group, last_request_time) "
                "VALUES (?, 0, 'Qo''lda qo''shilgan', ?)",
                (ch, now_str)
            )
            await db.commit()
        await event.respond(f"📥 `{ch}` 24 soatlik kuzatuvga qo'shildi.")


# ─────────────────────────────────────────────────────────────────────
# RAZVEDKA — BITTA RAQAM
# ── 2-TUZATMA: timeout va to'g'ri xato tutish
# ─────────────────────────────────────────────────────────────────────






def _scan_err(e: Exception) -> str:
    """Xato turini aniqlab, foydalanuvchiga qulay xabar qaytaradi."""
    s = str(e)
    if "not part of" in s or "Guruhga ulanib" in s or "not a member" in s:
        return ("🔒 **Userbot bu guruhga a'zo emas**\n\n"
                "Userbot hisobini guruhga qo'lda qo'shib, qayta urining.\n"
                "Sabab: admin botni chiqarib yuborgan bo'lishi mumkin.")
    if "ChannelPrivate" in s or "private" in s.lower():
        return "🔒 **Guruh/kanal yopiq** — kirish taqiqlangan."
    if "flood" in s.lower() or "FloodWait" in s:
        return "⏳ **Telegram vaqtincha chekladi.** Biroz kutib, qayta urining."
    if "ResolveUsername" in s or "USERNAME" in s.upper():
        return "⏳ **Username so'rovi cheklangan.** 30 daqiqa kutib, qayta urining."
    if "Topilmadi" in s or "not found" in s.lower():
        return "❌ **Guruh/kanal topilmadi.** Linkni tekshirib qayta urining."
    return f"❌ Xatolik: {e}"


async def run_keyword_search(sender_id, keyword, status_msg, days=None):
    try:
        # ── 1. Lokal keshdan qidiruv (Telegram API yo'q) ─────────────
        cache_total, cache_sources, last_sync = await engine.get_cache_stats()

        if cache_total > 0:
            try:
                await status_msg.edit(
                    f"🔎 **`{keyword}`** lokal bazadan qidirilmoqda...\n"
                    f"💾 Kesh: `{cache_total:,}` ta xabar | `{cache_sources}` ta manba\n"
                    f"🕐 Oxirgi sinxron: {last_sync or 'nomalum'}"
                )
            except Exception:
                pass

            all_results = await engine.search_keywords_local(keyword, days=days)

            davr_str = (
                f"Oxirgi {days} kun" if days and days < 365
                else ("Oxirgi 1 yil" if days == 365 else "Hammasi")
            )

            if not all_results:
                await bot.send_message(
                    sender_id,
                    f"🔎 **`{keyword}`** — lokal bazada topilmadi.\n"
                    f"💾 Kesh: `{cache_total:,}` xabar | Davr: {davr_str}\n"
                    f"💡 Kesh yangilanmagan bo'lsa /sync_cache buyrug'ini yuboring."
                )
                return

            header = (
                f"🔎 **Qidiruv natijalari (lokal):** `{keyword}`\n"
                f"📅 Davr: **{davr_str}**\n"
                f"📊 Topildi: `{len(all_results)}` ta\n"
                f"💾 Lokal keshdan — Telegram API ishlatilmadi\n"
                f"{'─' * 30}\n\n"
            )

        else:
            # ── 2. Kesh bo'sh — Telegram dan qidiruv (bir martalik) ──
            try:
                await status_msg.edit(
                    "⚠️ Lokal kesh hali to'ldirilmagan.\n"
                    "📡 Telegram'dan qidirilmoqda (bu oxirgi marta)...\n"
                    "💡 Keyingi qidiruvlar tezkor bo'ladi."
                )
            except Exception:
                pass

            sources_set = set()
            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                async with db.execute(
                    "SELECT DISTINCT group_link FROM users_memory_bank "
                    "WHERE group_link IS NOT NULL AND group_link != ''", ()
                ) as cur:
                    for (link,) in await cur.fetchall():
                        sources_set.add(link)
                async with db.execute(
                    "SELECT channel_id FROM hidden_channel_knocker WHERE status='joined'"
                ) as cur:
                    for (ch_id,) in await cur.fetchall():
                        sources_set.add(ch_id)

            sources = list(sources_set)
            if not sources:
                await bot.send_message(
                    sender_id, "📭 Bazada manba yo'q. Avval guruh skanerlang."
                )
                return

            all_results = []
            for i, source in enumerate(sources):
                try:
                    await status_msg.edit(
                        f"📡 `{source}` ({i+1}/{len(sources)})"
                    )
                except Exception:
                    pass
                try:
                    r = await engine.search_keywords(userbot, source, keyword, status_msg, days=days)
                    all_results.extend(r)
                except Exception:
                    continue
                await asyncio.sleep(1.5)

            davr_str = (
                f"Oxirgi {days} kun" if days and days < 365
                else ("Oxirgi 1 yil" if days == 365 else "Hammasi")
            )
            if not all_results:
                await bot.send_message(
                    sender_id,
                    f"🔎 **`{keyword}`** — `{len(sources)}` ta manbadan topilmadi."
                )
                return

            header = (
                f"🔎 **Qidiruv natijalari:** `{keyword}`\n"
                f"📅 Davr: **{davr_str}**\n"
                f"📊 Topildi: `{len(all_results)}` ta | Manbalar: `{len(sources)}` ta\n"
                f"{'─' * 30}\n\n"
            )

        # ── Natijalarni yuborish ──────────────────────────────────────
        chunk     = header
        chunk_num = 1
        for r in all_results:
            line = (
                f"👤 **{r['name']}**"
                + (f" (@{r['username']})" if r['username'] else "")
                + f"\n🆔 `{r['user_id']}`\n"
                f"📍 {r['source']}\n"
                f"📅 {r['date']}\n"
                f"💬 {r['text']}\n"
                f"{'─' * 25}\n\n"
            )
            if len(chunk) + len(line) > 3800:
                await bot.send_message(sender_id, chunk)
                await asyncio.sleep(0.5)
                chunk      = f"_(davomi {chunk_num + 1})_\n\n" + line
                chunk_num += 1
            else:
                chunk += line
        if chunk.strip():
            await bot.send_message(sender_id, chunk)

        # ── AI XULOSA: kalit so'z qidiruv tahlili ────────────────────
        if _AI_ENABLED and ai_mod and all_results:
            try:
                ai_status = await bot.send_message(sender_id, "🤖 AI xulosa tayyorlanmoqda...")
                summary = await ai_mod.summarize_keywords(keyword, all_results)
                await ai_status.edit(f"🤖 **AI Xulosa — `{keyword}` qidiruvi**\n\n{summary}")
            except Exception:
                pass
        # ─────────────────────────────────────────────────────────────

    except Exception as e:
        await bot.send_message(sender_id, _scan_err(e))
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


async def run_comment_scan(sender_id, target, fpath, status_msg):
    try:
        await status_msg.edit(
            f"💬 Kanal comment skanerlash boshlandi...\n"
            f"📍 Manba: `{target}`"
        )
    except Exception:
        pass
    try:
        result = await engine.scan_channel_comments(userbot, target, fpath, status_msg)
        count, ch_title = result
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT INTO archive_bin (file_name, file_path, created_date) VALUES (?, ?, ?)",
                (os.path.basename(fpath), fpath, datetime.now().strftime("%Y-%m-%d"))
            )
            await db.commit()
        await bot.send_file(
            sender_id, fpath,
            caption=(
                f"✅ **{ch_title}** comment skanerlash yakunlandi!\n"
                f"👥 Jami: `{count}` ta profil yozildi."
            )
        )
    except Exception as e:
        import traceback
        err_detail = traceback.format_exc()
        await bot.send_message(
            sender_id,
            f"❌ Comment skanerlashda xatolik:\n`{e}`\n\n"
            f"Sabab: discussion guruh topilmadi yoki kanal yopiq."
        )
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


async def run_msg_scan(sender_id, target, fpath, status_msg):
    try:
        count = await engine.scan_messages(userbot, target, fpath, status_msg)
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT INTO archive_bin (file_name, file_path, created_date) VALUES (?, ?, ?)",
                (os.path.basename(fpath), fpath, datetime.now().strftime("%Y-%m-%d"))
            )
            await db.commit()
        await bot.send_file(
            sender_id, fpath,
            caption=f"✅ Xabar skanerlash yakunlandi! `{count}` ta profil yozildi."
        )
    except FloodWaitError as e:
        h = e.seconds // 3600
        m = (e.seconds % 3600) // 60
        await bot.send_message(sender_id,
            f"⏳ **Telegram FloodWait**\n⏱ Kutish: **{h} soat {m} daqiqa**\nBu vaqt o'tgach qayta urining.")
    except Exception as e:
        err = str(e)
        if "not part of" in err or "Guruhga ulanib" in err or "not a member" in err:
            await bot.send_message(sender_id,
                "🔒 **Userbot bu guruhga a'zo emas**\n\n"
                "Userbot hisobini guruhga qo'lda qo'shib, qayta urinib ko'ring.\n"
                "Sabab: guruh admini botni chiqarib yuborgan bo'lishi mumkin.")
        elif "private" in err.lower() or "ChannelPrivate" in err:
            await bot.send_message(sender_id,
                "🔒 **Guruh yopiq**\n\nBu guruhga kirish taqiqlangan.")
        elif "flood" in err.lower():
            await bot.send_message(sender_id,
                "⏳ **Telegram vaqtincha chekladi**\nBiroz kutib, qayta urining.")
        else:
            await bot.send_message(sender_id, f"❌ Xabar skanerlashda xatolik: {e}")
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


async def run_auto_scan(sender_id, target, status_msg):
    """Link turini avtomatik aniqlab tegishli skanerlovchini ishga tushiradi."""
    try:
        from telethon.tl.types import Channel
        ent = await userbot.get_entity(target)  # faqat 1 marta resolve
        if isinstance(ent, Channel) and ent.broadcast:
            fname = f"Comment_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
            fpath = os.path.join(BASE_DIR, fname)
            await run_comment_scan(sender_id, target, fpath, status_msg)
        else:
            fname = f"Group_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
            fpath = os.path.join(BASE_DIR, fname)
            # pre_entity: deep_scan_group qayta resolve qilmaydi
            await run_background_scan(sender_id, target, fpath, status_msg, pre_entity=ent)
    except FloodWaitError as e:
        h = e.seconds // 3600
        m = (e.seconds % 3600) // 60
        await bot.send_message(
            sender_id,
            f"⏳ **Telegram FloodWait xatosi**\n\n"
            f"Userbot hisobi vaqtincha cheklangan.\n"
            f"⏱ Kutish vaqti: **{h} soat {m} daqiqa**\n\n"
            f"Bu vaqt o'tgach qayta urinib ko'ring."
        )
        try:
            await status_msg.delete()
        except Exception:
            pass
    except Exception as e:
        await bot.send_message(sender_id, _scan_err(e))
        try:
            await status_msg.delete()
        except Exception:
            pass


async def run_background_scan(sender_id, target, fpath, status_msg, pre_entity=None):
    global _last_scan_end_time
    try:
        count = await engine.deep_scan_group(
            userbot, target, fpath, status_msg, pre_entity=pre_entity
        )
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            await db.execute(
                "INSERT INTO archive_bin (file_name, file_path, created_date) VALUES (?, ?, ?)",
                (os.path.basename(fpath), fpath, datetime.now().strftime("%Y-%m-%d"))
            )
            await db.commit()
        await bot.send_file(
            sender_id, fpath,
            caption=f"✅ Skanerlash yakunlandi! `{count}` ta profil yozildi."
        )
        # ── AI XULOSA: skanerlash natijalari tahlili ─────────────────
        if _AI_ENABLED and ai_mod and count > 0:
            try:
                ai_status = await bot.send_message(sender_id, "🤖 AI xulosa tayyorlanmoqda...")
                users_sample = []
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
                    async with _db.execute(
                        "SELECT name, bio, personal_channel, private_channels, is_bot, is_premium "
                        "FROM users_memory_bank WHERE group_link LIKE ? ORDER BY rowid DESC LIMIT 50",
                        (f"%{target}%",)
                    ) as _cur:
                        for row in await _cur.fetchall():
                            users_sample.append({
                                "name": row[0], "bio": row[1],
                                "personal_channel": row[2], "private_channels": row[3],
                                "is_bot": row[4], "is_premium": row[5],
                            })
                summary = await ai_mod.summarize_scan(str(target), count, users_sample)
                await ai_status.edit(f"🤖 **AI Xulosa — Skanerlash natijasi**\n\n{summary}")
            except Exception:
                pass
        # ─────────────────────────────────────────────────────────────
    except FloodWaitError as e:
        h = e.seconds // 3600
        m = (e.seconds % 3600) // 60
        await bot.send_message(
            sender_id,
            f"⏳ **Telegram FloodWait xatosi**\n\n"
            f"Skanerlash jarayonida userbot hisobi cheklandi.\n"
            f"⏱ Kutish vaqti: **{h} soat {m} daqiqa**\n\n"
            f"Bu vaqt o'tgach qayta urinib ko'ring."
        )
    except Exception as e:
        await bot.send_message(sender_id, _scan_err(e))
    finally:
        _last_scan_end_time = time.time()
        try:
            await status_msg.delete()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────
# ISHGA TUSHIRISH
# ─────────────────────────────────────────────────────────────────────



# ─────────────────────────────────────────────────────────────────────
# TUGMA: KO'P MUSIQA QIDIRISH
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="🎼 Ko'p Musiqa Qidirish"))
async def btn_multi_music_search(event):
    if not await is_admin(event.sender_id):
        return
    total_fp, total_ch = await music.get_stats()
    if total_fp == 0:
        await event.respond(
            "⚠️ Baza bo\'sh!\n"
            "Monitoring ishlayotganini tekshiring."
        )
        return
    USER_STATES[event.sender_id] = {
        'state': 'waiting_multi_music',
        'files': []
    }
    await event.respond(
        f"🎼 **Ko\'p Musiqa Qidirish**\n\n"
        f"📊 Bazada `{total_fp}` ta fingerprint (`{total_ch}` manba)\n\n"
        f"Musiqalarni birin-ketin yuboring (10+ ta bo\'lishi mumkin)\n"
        f"Yuborib bo\'lgach `/search_done` yozing"
    )


@bot.on(events.NewMessage(pattern='/search_done'))
async def multi_music_done(event):
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict) or state.get('state') != 'waiting_multi_music':
        await event.respond("⚠️ Avval **🎼 Ko\'p Musiqa Qidirish** tugmasini bosing.")
        return
    files = state.get('files', [])
    if not files:
        await event.respond("⚠️ Hech qanday musiqa yuborilmadi.")
        return
    USER_STATES[event.sender_id] = None
    status = await event.respond(
        f"🎼 `{len(files)}` ta musiqa tahlil qilinmoqda...\n"
        f"Biroz kuting."
    )
    asyncio.create_task(run_multi_music_search(event.sender_id, files, status))


@bot.on(events.NewMessage(func=lambda e: True))
async def multi_music_file_handler(event):
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict) or state.get('state') != 'waiting_multi_music':
        return
    if not (event.audio or event.voice or event.document):
        return

    # Faylni saqlash
    tmp_path = os.path.join(BASE_DIR, f"multi_{event.sender_id}_{len(state['files'])}.ogg")
    try:
        await event.download_media(file=tmp_path)
        state['files'].append(tmp_path)
        count = len(state['files'])
        await event.respond(
            f"✅ `{count}` ta musiqa qabul qilindi\n"
            f"Yana musiqa yuboring yoki qidiruvni boshlang:",
            buttons=[
                [Button.inline("🔍 Qidiruvni Boshlash", data="multi_search_start")],
            ]
        )
    except Exception as e:
        await event.respond(f"❌ Xatolik: {e}")


@bot.on(events.CallbackQuery(pattern=b"multi_search_start"))
async def multi_search_start_callback(event):
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict) or state.get('state') != 'waiting_multi_music':
        await event.answer("⚠️ Sessiya tugagan.")
        return
    files = state.get('files', [])
    if not files:
        await event.answer("⚠️ Hech qanday musiqa yuborilmadi.")
        return
    USER_STATES[event.sender_id] = None
    await event.answer("🔍 Qidiruv boshlandi!")
    status = await event.respond(
        f"🎼 `{len(files)}` ta musiqa tahlil qilinmoqda...\n"
        f"Biroz kuting."
    )
    asyncio.create_task(run_multi_music_search(event.sender_id, files, status))


async def run_multi_music_search(sender_id, files, status_msg):
    try:
        # Excel tayyorlash
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = "Ko'p Musiqa Qidiruv"
        sh.append([
            "№", "Sizning Musiqa", "Manba Turi", "Kanal/Profil Nomi",
            "Telegram ID", "Kanal ID", "Ism", "Username", "Telefon",
            "Bio", "Maxfiy Kanal", "Shaxsiy Kanal", "O\'xshashlik %", "Manba Guruh"
        ])

        row_num   = 0
        found_any = False

        for idx, fpath in enumerate(files):
            music_name = f"Musiqa {idx + 1}"
            try:
                await status_msg.edit(
                    f"🎼 `{idx+1}/{len(files)}` tahlil qilinmoqda..."
                )
            except Exception:
                pass

            try:
                results = await music.search_music(fpath)
            except Exception:
                results = []

            if not results:
                row_num += 1
                sh.append([
                    row_num, music_name, "—", "Topilmadi",
                    "", "", "", "", "", "", "", "", "0%", ""
                ])
                continue

            found_any = True
            for r in results[:20]:  # Har musiqa uchun top 20
                row_num += 1
                is_profile = r['file_name'].startswith('profile_')
                manba_turi = "👤 Profil" if is_profile else "📢 Kanal"

                fname = uname = phone = bio = has_hidden = open_ch = grp_link = ""
                kanal_id = ""
                async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
                    async with db.execute(
                        "SELECT first_name, username, phone, bio, "
                        "has_hidden, open_channels, group_link "
                        "FROM users_memory_bank WHERE user_id=? LIMIT 1",
                        (r['channel_id'],)
                    ) as cur:
                        row = await cur.fetchone()
                    if row:
                        fname, uname, phone, bio, has_hidden, open_ch, grp_link = row
                        uname = ('@' + uname) if uname else ''
                        phone = ('+' + phone) if phone else ''

                    # Kanal ID (-100 formatida)
                    async with db.execute(
                        "SELECT channel_id FROM hidden_channel_knocker "
                        "WHERE creator_id=? LIMIT 1",
                        (r['channel_id'],)
                    ) as cur:
                        crow = await cur.fetchone()
                    if crow:
                        ch_raw = str(crow[0]).lstrip('-')
                        kanal_id = f"-100{ch_raw}" if not str(crow[0]).startswith('-100') else crow[0]

                shaxsiy = has_hidden if (has_hidden and has_hidden.startswith("http") and "t.me/+" not in (has_hidden or "")) else ""
                maxfiy  = has_hidden if (has_hidden and ("t.me/+" in (has_hidden or "") or "Maxfiy" in (has_hidden or ""))) else ""

                sh.append([
                    row_num, music_name, manba_turi, r['channel_name'],
                    r['channel_id'], kanal_id, fname, uname, phone,
                    bio or "", maxfiy, shaxsiy,
                    f"{r['score']}%", grp_link
                ])

        if not found_any:
            await bot.send_message(
                sender_id,
                f"📭 `{len(files)}` ta musiqadan hech biri bazada topilmadi."
            )
            return

        _loop = asyncio.get_event_loop()
        await _loop.run_in_executor(None, engine.apply_excel_styles, sh, row_num)
        out_name = f"Multi_Musiqa_{datetime.now().strftime('%d_%H%M')}.xlsx"
        out_path = os.path.join(BASE_DIR, out_name)
        await _loop.run_in_executor(None, wb.save, out_path)

        await bot.send_file(
            sender_id, out_path,
            caption=(
                f"🎼 **Ko\'p Musiqa Qidiruv yakunlandi!**\n\n"
                f"🎵 Tekshirildi: `{len(files)}` ta musiqa\n"
                f"📊 Natijalar: `{row_num}` ta\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
        )

    except Exception as e:
        await bot.send_message(sender_id, f"❌ Xatolik: {e}")
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass
        for fpath in files:
            try:
                if os.path.exists(fpath):
                    os.remove(fpath)
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────
# TUGMA: KUZATILADIGAN MUSIQALAR
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="🔔 Kuzatiladigan Musiqalar"))
async def btn_watch_music(event):
    if not await is_admin(event.sender_id):
        return
    watches = await music.get_watch_list()
    if watches:
        lines_txt = [f"🔔 **Kuzatiladigan musiqalar:** `{len(watches)}` ta\n"]
        for w_id, w_name, w_date in watches:
            lines_txt.append(f"  {w_id}. 🎵 `{w_name}` — {w_date}")
        buttons = []
        for w_id, w_name, w_date in watches:
            buttons.append([Button.inline(f"🗑 {w_id}. {w_name[:25]}", data=f"del_watch_{w_id}")])
        buttons.append([Button.inline("➕ Musiqa Qo\'shish", data="add_watch_music")])
        buttons.append([Button.inline("📋 Arxiv (Topilganlar)", data="watch_archive")])
        buttons.append([Button.inline("🔄 Bazani Qayta Skanerlash", data="watch_rescan")])
        await event.respond("\n".join(lines_txt), buttons=buttons)
    else:
        await event.respond(
            "🔔 **Kuzatiladigan musiqalar yo\'q**\n\n"
            "Musiqa qo\'shsangiz — darhol xabar beradi!",
            buttons=[
                [Button.inline("➕ Musiqa Qo\'shish", data="add_watch_music")],
                [Button.inline("📋 Arxiv (Topilganlar)", data="watch_archive")]
            ]
        )


@bot.on(events.CallbackQuery(pattern=b"del_watch_\\d+"))
async def del_watch_callback(event):
    if not await is_admin(event.sender_id):
        return
    w_id = int(event.data.decode().split("_")[-1])
    await music.delete_watch_music(w_id)
    await event.answer("✅ O\'chirildi!")
    watches = await music.get_watch_list()
    if watches:
        lines_txt = [f"🔔 **Kuzatiladigan musiqalar:** `{len(watches)}` ta\n"]
        for wid, wname, wdate in watches:
            lines_txt.append(f"  {wid}. 🎵 `{wname}` — {wdate}")
        buttons = []
        for wid, wname, wdate in watches:
            buttons.append([Button.inline(f"🗑 {wid}. {wname[:25]}", data=f"del_watch_{wid}")])
        buttons.append([Button.inline("➕ Musiqa Qo\'shish", data="add_watch_music")])
        await event.edit("\n".join(lines_txt), buttons=buttons)
    else:
        await event.edit(
            "🔔 Kuzatiladigan musiqalar yo\'q.",
            buttons=[[Button.inline("➕ Musiqa Qo\'shish", data="add_watch_music")]]
        )


@bot.on(events.CallbackQuery(pattern=b"add_watch_music"))
async def add_watch_callback(event):
    if not await is_admin(event.sender_id):
        return
    await event.answer()
    USER_STATES[event.sender_id] = {
        'state': 'waiting_watch_music',
        'name':  None
    }
    await event.respond(
        "🎵 **Kuzatiladigan musiqa qo\'shish**\n\n"
        "Avval musiqa nomini yozing (masalan: `Jakone - Mutny tip`):"
    )




@bot.on(events.NewMessage(func=lambda e: True))
async def watch_music_handler(event):
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict) or state.get('state') != 'waiting_watch_music':
        return

    # Avval nom so'raladi
    if state.get('name') is None:
        if not event.text:
            return
        state['name'] = event.text.strip()
        await event.respond(
            f"✅ Nom: `{state['name']}`\n\n"
            f"Endi musiqani yuboring 👇"
        )
        return

    # Keyin musiqa fayli
    if not (event.audio or event.voice or event.document):
        await event.respond("⚠️ Audio fayl yuboring!")
        return

    USER_STATES[event.sender_id] = None
    status = await event.respond("🎵 Musiqa tahlil qilinmoqda...")

    tmp_path = os.path.join(BASE_DIR, f"watch_{event.sender_id}.ogg")
    try:
        await event.download_media(file=tmp_path)
        fp, duration = await engine.music_mod.get_fingerprint_async(tmp_path)
        if not fp:
            await status.edit("❌ Fingerprint olishda xatolik. fpcalc.exe borligini tekshiring.")
            return
        await music.add_watch_music(fp, duration or 0, state['name'], event.sender_id)
        await status.edit(
            f"✅ **`{state['name']}`** kuzatuvga qo\'shildi!\n\n"
            f"Endi kimdir bu musiqani profil yoki kanalga qo\'ysa — darhol xabar beraman! 🔔"
        )
    except Exception as e:
        await status.edit(f"❌ Xatolik: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# ─────────────────────────────────────────────────────────────────────
# REAL-TIME ALERT TIZIMI
# ─────────────────────────────────────────────────────────────────────

async def watch_alert_sender():
    """Kuzatiladigan musiqa yoki kalit so'z alert topilganda xabar yuboradi."""
    while True:
        item = await engine._WATCH_ALERTS.get()
        try:
            # Yangi format: ('alert', hit_dict)
            if isinstance(item, tuple) and len(item) == 2 and item[0] == 'alert':
                await process_alert_hits([item[1]])
                continue
            alert = item
            source_id = alert['source_id']
            source_type = alert['source_type']
            source_name = alert['source_name']

            # Kanal linki va -100 ID formatini tayyorlash
            extra = ""
            if source_type == 'kanal':
                ch_id_clean = str(source_id).lstrip('-')
                ch_id_fmt = f"-100{ch_id_clean}" if not str(source_id).startswith('-100') else source_id
                ch_link = ""
                try:
                    entity = await userbot.get_entity(int(source_id))
                    username = getattr(entity, 'username', None)
                    if username:
                        ch_link = f"https://t.me/{username}"
                except Exception:
                    pass
                if not ch_link:
                    try:
                        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
                            async with _db.execute(
                                "SELECT channel_id FROM hidden_channel_knocker "
                                "WHERE channel_id LIKE ? LIMIT 1",
                                (f"%{ch_id_clean}%",)
                            ) as _cur:
                                _row = await _cur.fetchone()
                        if _row and str(_row[0]).startswith('https://'):
                            ch_link = _row[0]
                    except Exception:
                        pass
                extra = (f"\n🔗 Link: {ch_link}\n🆔 Kanal ID: `{ch_id_fmt}`" if ch_link
                         else f"\n🆔 Kanal ID: `{ch_id_fmt}`")

            msg_text = (
                f"🚨 **KUZATILADIGAN MUSIQA TOPILDI!**\n\n"
                f"🎵 Musiqa: `{alert['watch_name']}`\n"
                f"{'👤' if source_type == 'profil' else '📢'} "
                f"{'Profil' if source_type == 'profil' else 'Kanal'}: "
                f"`{source_name}`\n"
                f"🆔 ID: `{source_id}`\n"
                f"🎯 O\'xshashlik: `{alert['score']}%`\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                + extra
            )

            # Barcha adminlarga xabar yuborish
            all_admins = set()
            all_admins.add(SUPER_ADMIN_ID)
            all_admins.update(EXTRA_ADMIN_IDS)
            async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
                async with _db.execute("SELECT admin_id FROM trusted_admins") as _cur:
                    for (aid,) in await _cur.fetchall():
                        all_admins.add(aid)

            try:
                await music.save_watch_alert_log(
                    alert['watch_name'], source_name, source_id,
                    source_type, alert['score']
                )
            except Exception:
                pass

            for admin_id in all_admins:
                try:
                    await bot.send_message(admin_id, msg_text)
                    await asyncio.sleep(0.1)  # Telegram 30 msg/s chegarasidan o'tmaslik
                except Exception:
                    pass
        except Exception:
            pass
        finally:
            engine._WATCH_ALERTS.task_done()

    # ── Alert xabarlari uchun handler ──────────────────────────────
    # (yuqoridagi loop dan keyin hech qachon yetib kelmaydi, lekin xavfsizlik uchun)


async def _handle_alert_queue_item(item):
    """Alert queue dan kelgan xabarni qayta ishlash."""
    try:
        kind, data = item
        if kind == 'alert':
            await process_alert_hits([data])
    except Exception:
        pass


@bot.on(events.NewMessage(pattern='/clean_tmp'))
async def clean_tmp(event):
    if not await is_admin(event.sender_id):
        return
    import glob
    files = glob.glob(os.path.join(BASE_DIR, "tmp_*.ogg"))
    count = 0
    for f in files:
        try:
            os.remove(f)
            count += 1
        except Exception:
            pass
    await event.respond(f"🗑 {count} ta vaqtinchalik fayl o\'chirildi.")


@bot.on(events.NewMessage(pattern='/check_media'))
async def check_media(event):
    if not await is_admin(event.sender_id):
        return
    # .env dagi birinchi manbadan oxirgi 20 ta xabarni tekshirish
    status = await event.respond("🔍 Tekshirilmoqda...")
    try:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            # Avval joined kanallardan qidirish
            async with db.execute(
                "SELECT channel_id FROM hidden_channel_knocker "
                "WHERE status='joined' LIMIT 1"
            ) as cur:
                row = await cur.fetchone()
            # Topilmasa open_channels dan
            if not row:
                async with db.execute(
                    "SELECT DISTINCT open_channels FROM users_memory_bank "
                    "WHERE open_channels IS NOT NULL AND open_channels != '' "
                    "AND open_channels != 'Yo\'q' LIMIT 1"
                ) as cur:
                    row = await cur.fetchone()
        
        if not row:
            await status.edit("❌ Manba topilmadi.")
            return
        
        source = row[0]
        entity = await userbot.get_entity(source)
        
        lines = ["📊 **Oxirgi 50 ta xabar tahlili:**\n"]
        audio_count = 0
        total = 0
        async for msg in userbot.iter_messages(entity, limit=50):
            total += 1
            if msg.audio or msg.voice or msg.document:
                media = msg.audio or msg.voice or msg.document
                mime = getattr(media, 'mime_type', 'unknown')
                size = getattr(media, 'size', 0)
                duration = getattr(media, 'duration', 0)
                lines.append(
                    f"`{msg.id}` | {mime} | {size//1024} KB | {duration}s"
                )
                audio_count += 1

        lines.append(f"\n📊 Jami: {total} xabar, {audio_count} ta media")
        await status.edit("\n".join(lines[:25]))
    except Exception as e:
        await status.edit(f"❌ Xatolik: {e}")


@bot.on(events.NewMessage(pattern=r'/rescan_channel (.+)'))
async def rescan_channel_cmd(event):
    if not await is_admin(event.sender_id):
        return
    target = _pm(event, 1).strip()
    status = await event.respond(f"🔄 `{target}` qayta skanerlanmoqda...")

    try:
        entity = await userbot.get_entity(target)
        channel_id = str(entity.id)
        channel_name = getattr(entity, 'title', target)

        # Progressni o'chirish — to'liq qayta skanerlash uchun
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            try:
                await db.execute(
                    "CREATE TABLE IF NOT EXISTS music_channel_progress "
                    "(channel_id TEXT PRIMARY KEY, last_msg_id INTEGER)"
                )
                await db.execute(
                    "DELETE FROM music_channel_progress WHERE channel_id=?",
                    (channel_id,)
                )
                await db.commit()
            except Exception:
                pass

        await status.edit(
            f"🎵 **`{channel_name}`** qayta skanerlanmoqda...\n"
            f"Jami xabarlar tekshiriladi. Biroz kuting."
        )

        asyncio.create_task(_do_rescan_channel(event.sender_id, entity, channel_id, channel_name, target))

    except Exception as e:
        await status.edit(f"❌ Xatolik: `{e}`")


async def _do_rescan_channel(sender_id, entity, channel_id, channel_name, target):
    """Kanalning barcha musiqalarini qayta skanerlash"""
    import music_scanner as music_mod_local
    await music_mod_local.init_music_db()
    BASE_DIR_LOCAL = os.path.dirname(os.path.abspath(__file__))
    count = 0
    total_audio = 0

    try:
        # Barcha audio xabarlarni yig'ish
        audio_msgs = []
        async for msg in userbot.iter_messages(entity, limit=None):
            if engine.is_music_file(msg):
                audio_msgs.append(msg)
        total_audio = len(audio_msgs)

        await bot.send_message(
            sender_id,
            f"🎵 `{channel_name}` da `{total_audio}` ta musiqa topildi. Skanerlanmoqda..."
        )

        async def _dl(m):
            tmp = os.path.join(BASE_DIR_LOCAL, f"tmp_rescan_{channel_id}_{m.id}.ogg")
            for attempt in range(3):
                try:
                    await m.download_media(file=tmp)
                    if os.path.exists(tmp) and os.path.getsize(tmp) > 0:
                        return (m.id, tmp)
                    if os.path.exists(tmp):
                        os.remove(tmp)
                except Exception:
                    if os.path.exists(tmp):
                        os.remove(tmp)
            return (m.id, None)

        BATCH = 5
        for i in range(0, len(audio_msgs), BATCH):
            batch = audio_msgs[i:i+BATCH]
            results = await asyncio.gather(*[_dl(m) for m in batch])
            for msg_id, tmp in results:
                if tmp and os.path.exists(tmp):
                    try:
                        fp, dur = await engine.music_mod.get_fingerprint_async(tmp)
                        if fp:
                            await music_mod_local.save_fingerprint(
                                channel_id, channel_name,
                                f"msg_{msg_id}", fp, dur or 0
                            )
                            count += 1
                    except Exception:
                        pass
                    finally:
                        if os.path.exists(tmp):
                            os.remove(tmp)

        await bot.send_message(
            sender_id,
            f"✅ **`{channel_name}`** qayta skanerlash yakunlandi!\n\n"
            f"🎵 Jami musiqa: `{total_audio}` ta\n"
            f"🎶 Fingerprint saqlandi: `{count}` ta"
        )

    except Exception as e:
        await bot.send_message(sender_id, f"❌ Xatolik: `{e}`")


@bot.on(events.NewMessage(pattern='/test_fpcalc'))
async def test_fpcalc(event):
    if not await is_admin(event.sender_id):
        return
    import subprocess, glob
    fpcalc_paths = [
        os.path.join(BASE_DIR, "fpcalc.exe"),
        os.path.join(BASE_DIR, "fpcalc"),
        "fpcalc.exe", "fpcalc"
    ]
    found_path = next((p for p in fpcalc_paths if os.path.exists(p)), None)
    if not found_path:
        await event.respond(
            "❌ **fpcalc.exe topilmadi!**\n" +
            "\n".join(f"`{p}`" for p in fpcalc_paths[:2])
        )
        return
    try:
        result = subprocess.run([found_path, "-version"], capture_output=True, text=True, timeout=10)
        await event.respond(
            f"✅ **fpcalc topildi!**\n"
            f"📁 Yo\'li: `{found_path}`\n"
            f"📋 Versiya: `{result.stdout.strip() or result.stderr.strip()}`"
        )
    except Exception as e:
        await event.respond(f"❌ fpcalc ishga tushmadi: `{e}`")


@bot.on(events.NewMessage(pattern=r'/check_profile (\d+)'))
async def check_profile_music(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    try:
        fi = await userbot(engine.GetFullUserRequest(uid))
        fu = fi.full_user
        fields = []
        for attr in ['saved_music', 'profile_song', 'profile_songs', 'music']:
            val = getattr(fu, attr, None)
            if val is not None:
                fields.append(f"`{attr}`: {type(val).__name__}")
        if fields:
            await event.respond(f"🎵 **Profil {uid} musiqa fieldlari:**\n\n" + "\n".join(fields))
        else:
            await event.respond(f"📭 Profil {uid} da musiqa field topilmadi")
    except Exception as e:
        await event.respond(f"❌ Xatolik: {e}")


@bot.on(events.NewMessage(pattern='/rescan_music'))
async def rescan_music_handler(event):
    if not await is_admin(event.sender_id):
        return
    status = await event.respond(
        "🎵 Barcha profillar musiqasi qayta skanerlanmoqda...\n"
        "Fon da ishlaydi."
    )
    asyncio.create_task(_rescan_all_profiles_music(event.sender_id, status))


async def _rescan_all_profiles_music(sender_id, status_msg):
    count = 0
    found = 0
    try:
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute("SELECT DISTINCT user_id FROM users_memory_bank") as cur:
                users = await cur.fetchall()
        total = len(users)
        await status_msg.edit(f"🎵 Jami `{total}` ta profil tekshiriladi...")
        for (uid,) in users:
            try:
                fi = await engine._safe_api_call(lambda: userbot(engine.GetFullUserRequest(uid)))
                if fi is None:
                    count += 1
                    continue
                fu = fi.full_user
                music_docs = []
                for field in ['saved_music', 'profile_song', 'profile_songs', 'music']:
                    val = getattr(fu, field, None)
                    if val is None: continue
                    if isinstance(val, list): music_docs.extend(val)
                    else: music_docs.append(val)
                for idx, doc in enumerate(music_docs):
                    if hasattr(doc, 'document'): doc = doc.document
                    if not hasattr(doc, 'id'): continue
                    tmp_path = os.path.join(BASE_DIR, f"tmp_rescan_{uid}_{idx}.ogg")
                    try:
                        await userbot.download_media(doc, file=tmp_path)
                        if os.path.exists(tmp_path):
                            fp, duration = await engine.music_mod.get_fingerprint_async(tmp_path)
                            if fp:
                                already = await music.is_profile_music_saved(uid, fp)
                                if not already:
                                    await music.init_music_db()
                                    await music.save_fingerprint(str(uid), f"Profil: {uid}", f"profile_{uid}_{idx}", fp, duration or 0)
                                    found += 1
                    except Exception:
                        pass
                    finally:
                        if os.path.exists(tmp_path): os.remove(tmp_path)
                count += 1
                if count % 100 == 0:
                    try:
                        await status_msg.edit(f"🎵 `{count}/{total}` tekshirildi | 🎶 Topildi: `{found}` ta")
                    except Exception:
                        pass
            except Exception:
                pass
            await asyncio.sleep(0.3)
        await bot.send_message(sender_id,
            f"✅ **Musiqa rescan yakunlandi!**\n\n"
            f"👥 Tekshirildi: `{count}` ta\n"
            f"🎶 Topildi: `{found}` ta")
        try: await status_msg.delete()
        except: pass
    except Exception as e:
        await bot.send_message(sender_id, f"❌ Xatolik: {e}")


@bot.on(events.NewMessage(pattern=r'/test_channel (.+)'))
async def test_channel_cmd(event):
    if not await is_admin(event.sender_id):
        return
    target = _pm(event, 1).strip()
    status = await event.respond(f"🔍 `{target}` tekshirilmoqda...")

    try:
        entity = await userbot.get_entity(target)
        channel_name = getattr(entity, 'title', target)

        total = 0
        audio_count = 0
        async for msg in userbot.iter_messages(entity, limit=None):
            total += 1
            if engine.is_music_file(msg):
                audio_count += 1

        await status.edit(
            f"📊 **`{channel_name}` tahlili:**\n\n"
            f"📨 Jami xabarlar: `{total}` ta\n"
            f"🎵 Musiqa fayllar: `{audio_count}` ta\n\n"
            f"Bot shu `{audio_count}` ta musiqani fingerprint qiladi."
        )
    except Exception as e:
        await status.edit(f"❌ Xatolik: {e}")


@bot.on(events.NewMessage(pattern='/check_knocker'))
async def check_knocker(event):
    if not await is_admin(event.sender_id):
        return
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        async with db.execute(
            "SELECT COUNT(*) FROM hidden_channel_knocker WHERE status='pending'"
        ) as cur:
            pending = (await cur.fetchone())[0]
        async with db.execute(
            "SELECT channel_id, last_request_time FROM hidden_channel_knocker "
            "WHERE status='pending' LIMIT 5"
        ) as cur:
            samples = await cur.fetchall()

    lines = [
        f"🔍 **Knocker holati:**\n",
        f"📥 Pending kanallar: `{pending}` ta",
        f"📨 Bugungi so\'rovnomalar: `{engine._daily_knock_count}/{engine.MAX_DAILY_KNOCKS}`",
        f"⏱ Interval: `{engine.KNOCK_INTERVAL // 60}` daqiqa",
        f"⚡️ MONITORING_PAUSED: `{engine.MONITORING_PAUSED}`",
        f"\n**Namuna (5 ta):**"
    ]
    for ch_id, last_req in samples:
        from datetime import datetime as dt
        elapsed = "—"
        if last_req:
            try:
                diff = (dt.now() - dt.strptime(last_req, "%Y-%m-%d %H:%M")).total_seconds()
                elapsed = f"{int(diff//3600)}s {int((diff%3600)//60)}d o\'tgan"
            except:
                elapsed = last_req
        lines.append(f"  `{ch_id[:30]}` — {elapsed}")

    await event.respond("\n".join(lines))


@bot.on(events.NewMessage(pattern="🤖 AI Xulosa"))
async def btn_ai_xulosa(event):
    if not await is_admin(event.sender_id):
        return
    if not _AI_ENABLED or not ai_mod:
        await bot.send_message(
            event.sender_id,
            "⚠️ **AI moduli faol emas.**\n\n"
            "`.env` fayliga `ANTHROPIC_API_KEY=sk-ant-...` qo'shing.",
            buttons=MAIN_KEYBOARD,
        )
        return
    status = await bot.send_message(event.sender_id, "🤖 AI hisobot tayyorlanmoqda...")
    try:
        stats = {}
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as _db:
            async with _db.execute("SELECT COUNT(*) FROM users_memory_bank") as c:
                stats["total_users"] = (await c.fetchone() or [0])[0]
            async with _db.execute("SELECT COUNT(*) FROM messages_cache") as c:
                stats["total_messages"] = (await c.fetchone() or [0])[0]
            async with _db.execute(
                "SELECT COUNT(DISTINCT group_link) FROM users_memory_bank WHERE group_link != ''"
            ) as c:
                stats["total_sources"] = (await c.fetchone() or [0])[0]
            async with _db.execute(
                "SELECT COUNT(*) FROM hidden_channel_knocker"
            ) as c:
                stats["knocker_count"] = (await c.fetchone() or [0])[0]
            async with _db.execute(
                "SELECT created_date FROM archive_bin ORDER BY rowid DESC LIMIT 1"
            ) as c:
                row = await c.fetchone()
                stats["last_scan"] = row[0] if row else "yo'q"
        try:
            import music_scanner as _ms
            fp_total, _ = await _ms.get_stats()
            watch_total, _ = await _ms.get_watch_stats()
            stats["music_fingerprints"] = fp_total
            stats["watch_count"] = watch_total
        except Exception:
            pass
        summary = await ai_mod.generate_status_report(stats)
        await status.edit(
            f"🤖 **AI Umumiy Hisobot**\n\n"
            f"📊 Bazadagi ma'lumotlar:\n"
            f"  👤 Foydalanuvchilar: `{stats['total_users']:,}`\n"
            f"  💬 Kesh xabarlari: `{stats['total_messages']:,}`\n"
            f"  📡 Manbalar: `{stats['total_sources']}`\n"
            f"  🔒 Knocker kanallar: `{stats['knocker_count']}`\n\n"
            f"🤖 **AI Tahlil:**\n{summary}"
        )
    except Exception as e:
        await status.edit(f"❌ AI xulosa xatosi: {e}")


@bot.on(events.NewMessage(pattern="🔄 Botni Qayta Yuklash"))
async def btn_restart(event):
    if not await is_admin(event.sender_id):
        return
    await show_control_panel(event)


async def show_control_panel(event_or_msg, edit=False):
    """Nazorat panelini ko'rsatish"""
    r = engine._RESOURCE
    paused = engine.MONITORING_PAUSED
    music_paused = r['music_paused']
    profile_slow = r['profile_slow']
    heavy = r['heavy_scan']

    # Holat
    if paused:
        holat = "⏸ HAMMASI TO'XTATILGAN"
    elif heavy:
        holat = f"🔴 {r['current_task']} ishlayapti"
    else:
        holat = "🟢 NORMAL ISHLAYAPTI"

    navbat = _SCAN_QUEUE.qsize() if _SCAN_QUEUE else 0

    text = (
        f"🎛 **BOT NAZORAT PANELI**\n\n"
        f"⚙️ Holat: **{holat}**\n"
        f"📋 Navbat: `{navbat}` ta topshiriq\n\n"
        f"**Fon jarayonlar:**\n"
        f"{'⏸' if paused else '▶️'} Skanerlash\n"
        f"{'⏸' if music_paused else '▶️'} Musiqa tracker\n"
        f"{'🐢' if profile_slow else '▶️'} Profil tracker\n"
        f"{'⏸' if paused else '▶️'} So\'rovnoma tizimi\n"
    )

    buttons = [
        [Button.inline("🔄 Qayta Yuklash", data="bot_restart")],
        [
            Button.inline("⏸ Hammasini To'xtatish", data="ctrl_pause_all") if not paused
            else Button.inline("▶️ Hammasini Yoqish", data="ctrl_resume_all")
        ],
        [
            Button.inline("⏸ Musiqa Tracker", data="ctrl_pause_music") if not music_paused
            else Button.inline("▶️ Musiqa Tracker", data="ctrl_resume_music"),
            Button.inline("⏸ Profil Tracker", data="ctrl_pause_profile") if not profile_slow
            else Button.inline("▶️ Profil Tracker", data="ctrl_resume_profile"),
        ],
        [Button.inline("🗑 Navbatni Tozalash", data="ctrl_clear_queue")],
        [Button.inline("🔄 Yangilash", data="ctrl_refresh")],
    ]

    if edit:
        try:
            await event_or_msg.edit(text, buttons=buttons)
        except Exception:
            pass
    else:
        await event_or_msg.respond(text, buttons=buttons)


@bot.on(events.CallbackQuery(pattern=b"bot_restart"))
async def bot_restart_cb(event):
    if not await is_admin(event.sender_id):
        return
    await event.answer("🔄 Qayta yuklanmoqda...")
    await event.edit("🔄 **Bot qayta yuklanmoqda...**")
    await asyncio.sleep(1)
    try:
        await userbot.disconnect()
        await bot.disconnect()
    except Exception:
        pass
    os.execv(sys.executable, [sys.executable] + sys.argv)


@bot.on(events.CallbackQuery(pattern=b"ctrl_pause_all"))
async def ctrl_pause_all(event):
    if not await is_admin(event.sender_id):
        return
    engine.MONITORING_PAUSED = True
    engine._RESOURCE['music_paused'] = True
    engine._RESOURCE['profile_slow'] = True
    await event.answer("⏸ Hammasi to'xtatildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_resume_all"))
async def ctrl_resume_all(event):
    if not await is_admin(event.sender_id):
        return
    engine.MONITORING_PAUSED = False
    engine._SCAN_COUNT = 0
    engine._RESOURCE['music_paused'] = False
    engine._RESOURCE['profile_slow'] = False
    engine._RESOURCE['heavy_scan'] = False
    engine._RESOURCE['current_task'] = None
    await event.answer("▶️ Hammasi yoqildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_pause_music"))
async def ctrl_pause_music(event):
    if not await is_admin(event.sender_id):
        return
    engine._RESOURCE['music_paused'] = True
    await event.answer("⏸ Musiqa tracker to'xtatildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_resume_music"))
async def ctrl_resume_music(event):
    if not await is_admin(event.sender_id):
        return
    engine._RESOURCE['music_paused'] = False
    await event.answer("▶️ Musiqa tracker yoqildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_pause_profile"))
async def ctrl_pause_profile(event):
    if not await is_admin(event.sender_id):
        return
    engine._RESOURCE['profile_slow'] = True
    await event.answer("🐢 Profil tracker sekinlashtirildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_resume_profile"))
async def ctrl_resume_profile(event):
    if not await is_admin(event.sender_id):
        return
    engine._RESOURCE['profile_slow'] = False
    await event.answer("▶️ Profil tracker tezlashtirildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_clear_queue"))
async def ctrl_clear_queue(event):
    if not await is_admin(event.sender_id):
        return
    count = _SCAN_QUEUE.qsize() if _SCAN_QUEUE else 0
    while _SCAN_QUEUE and not _SCAN_QUEUE.empty():
        try:
            _SCAN_QUEUE.get_nowait()
            _SCAN_QUEUE.task_done()
        except asyncio.QueueEmpty:
            break
    await event.answer(f"🗑 {count} ta topshiriq o'chirildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"ctrl_refresh"))
async def ctrl_refresh(event):
    if not await is_admin(event.sender_id):
        return
    await event.answer("🔄 Yangilandi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"bot_pause"))
async def bot_pause_callback(event):
    if not await is_admin(event.sender_id):
        return
    engine.MONITORING_PAUSED = True
    await event.answer("⏸ To\'xtatildi!")
    await show_control_panel(event, edit=True)


@bot.on(events.CallbackQuery(pattern=b"bot_resume"))
async def bot_resume_callback(event):
    if not await is_admin(event.sender_id):
        return
    engine.MONITORING_PAUSED = False
    engine._SCAN_COUNT = 0
    await event.answer("▶️ Davom ettirildi!")
    await show_control_panel(event, edit=True)


# Skanerlash navbati — asyncio.Queue, main() da ishga tushiriladi
_SCAN_QUEUE: asyncio.Queue = None
_last_scan_end_time: float = 0.0   # oxirgi skan tugagan vaqt (Unix timestamp)
SCAN_COOLDOWN: int = 15 * 60       # skanerlashlar orasidagi minimal tanaffus (15 daqiqa)


async def scan_queue_runner():
    """Navbatdagi skanerlashlarni ketma-ket bajaradi (polling yo'q)."""
    while True:
        item = await _SCAN_QUEUE.get()
        try:
            # Oldingi skanerlash tugashini kutish
            while engine._SCAN_COUNT > 0 or engine.MONITORING_PAUSED:
                await asyncio.sleep(2)

            # 15 daqiqalik tanaffus tekshiruvi
            scan_type, sender_id, target = item
            elapsed = time.time() - _last_scan_end_time
            if elapsed < SCAN_COOLDOWN:
                wait_secs = int(SCAN_COOLDOWN - elapsed)
                mins = wait_secs // 60
                secs = wait_secs % 60
                await bot.send_message(
                    sender_id,
                    f"⏳ **Tanaffus vaqti**\n\n"
                    f"Keyingi skanerlashgacha: **{mins} daqiqa {secs} soniya**\n"
                    f"Telegram API cheklovidan himoya uchun kutilmoqda..."
                )
                await asyncio.sleep(wait_secs)
                await bot.send_message(sender_id, "✅ Tanaffus tugadi, skanerlash boshlanmoqda...")
            try:
                await bot.send_message(
                    sender_id,
                    f"▶️ Navbatdagi skanerlash boshlandi: `{target}`"
                )
                if scan_type == 'auto':
                    status = await bot.send_message(sender_id, "⏳ Skanerlash boshlandi, tur aniqlanmoqda...")
                    asyncio.create_task(run_auto_scan(sender_id, target, status))
                elif scan_type == 'msg':
                    fname  = f"Msg_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
                    fpath  = os.path.join(BASE_DIR, fname)
                    status = await bot.send_message(sender_id, "⏳ Yopiq guruh xabarlari o'qilmoqda...")
                    asyncio.create_task(run_msg_scan(sender_id, target, fpath, status))
                elif scan_type == 'group':
                    fname  = f"Group_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
                    fpath  = os.path.join(BASE_DIR, fname)
                    status = await bot.send_message(sender_id, "⏳ Skanerlash boshlandi...")
                    asyncio.create_task(run_background_scan(sender_id, target, fpath, status))
                elif scan_type == 'comment':
                    fname  = f"Comment_OSINT_{datetime.now().strftime('%d_%H%M')}.xlsx"
                    fpath  = os.path.join(BASE_DIR, fname)
                    status = await bot.send_message(sender_id, "⏳ Kanal commentariyalari o'qilmoqda...")
                    asyncio.create_task(run_comment_scan(sender_id, target, fpath, status))
                elif scan_type == 'keyword':
                    keyword, days, st = target
                    asyncio.create_task(run_keyword_search(sender_id, keyword, st, days))
            except Exception as e:
                print(f"scan_queue_runner xatosi: {e}")
        finally:
            _SCAN_QUEUE.task_done()


@bot.on(events.CallbackQuery(pattern=b"watch_archive"))
async def watch_archive_callback(event):
    if not await is_admin(event.sender_id):
        return
    await event.answer()
    logs = await music.get_watch_alerts_log(limit=1000)
    if not logs:
        await event.respond("📋 Arxiv bo\'sh — hali hech narsa topilmagan.")
        return

    # Excel tayyorlash
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Topilganlar Arxivi"
    sh.append([
        "№", "Musiqa Nomi", "Manba Turi", "Kanal/Profil",
        "ID", "Ism", "Username", "Telefon", "Manba Guruh",
        "O'xshashlik %", "Topilgan Sana"
    ])

    # Barcha foydalanuvchi ma'lumotlarini bitta ulanishda oldindan yuklab olish
    all_src_ids = list({src_id for (_, _, src_id, _, _, _) in logs})
    user_data_map = {}
    async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
        for _sid in all_src_ids:
            async with db.execute(
                "SELECT first_name, username, phone, group_link "
                "FROM users_memory_bank WHERE user_id=? LIMIT 1", (_sid,)
            ) as cur:
                _urow = await cur.fetchone()
            if _urow:
                user_data_map[_sid] = _urow

    for i, (w_name, src_name, src_id, src_type, score, found_date) in enumerate(logs, 1):
        manba = "👤 Profil" if src_type == "profil" else "📢 Kanal"

        fname = uname = phone = grp_link = ""
        _urow = user_data_map.get(src_id)
        if _urow:
            fname = _urow[0] or ""
            uname = ('@' + _urow[1]) if _urow[1] else ""
            phone = ('+' + _urow[2]) if _urow[2] else ""
            grp_link = _urow[3] or ""

        sh.append([
            i, w_name, manba, src_name, src_id,
            fname, uname, phone, grp_link,
            f"{score}%", found_date
        ])

    _loop = asyncio.get_event_loop()
    await _loop.run_in_executor(None, engine.apply_excel_styles, sh, len(logs))

    out_name = f"Watch_Arxiv_{datetime.now().strftime('%d_%H%M')}.xlsx"
    out_path = os.path.join(BASE_DIR, out_name)
    await _loop.run_in_executor(None, wb.save, out_path)

    await bot.send_file(
        event.sender_id, out_path,
        caption=(
            f"📋 **Topilganlar arxivi**\n\n"
            f"🎵 Jami: `{len(logs)}` ta topilgan\n"
            f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
    )


@bot.on(events.CallbackQuery(pattern=rb"watch_archive_\d+"))
async def watch_archive_detail(event):
    if not await is_admin(event.sender_id):
        return
    w_id = int(event.data.decode().split("_")[-1])
    watches = await music.get_watch_list()
    watch = next((w for w in watches if w[0] == w_id), None)
    if not watch:
        await event.answer("Topilmadi")
        return
    logs = await music.get_watch_alerts_log(watch_name=watch[1], limit=20)
    lines = [f"📋 **`{watch[1]}` arxivi:** `{len(logs)}` ta\n"]
    for w_name, src_name, src_id, src_type, score, found_date in logs:
        icon = "👤" if src_type == "profil" else "📢"
        lines.append(f"{icon} `{src_name}` — {score}% — {found_date}")
    await event.respond("\n".join(lines) if lines else "📭 Bu musiqa hali topilmagan.")


@bot.on(events.NewMessage(pattern='/flood_stats'))
async def flood_stats_cmd(event):
    if not await is_admin(event.sender_id):
        return
    
    stats = engine._FLOOD_STATS
    if not stats:
        await event.respond("✅ Hozircha flood yo'q!")
        return

    # Saralash - eng ko'p flood bergani birinchi
    sorted_stats = sorted(stats.items(), key=lambda x: x[1]['count'], reverse=True)
    
    lines = ["📊 **FLOOD STATISTIKASI:**\n"]
    for func, data in sorted_stats:
        lines.append(
            f"🔴 **{func}**\n"
            f"   Soni: `{data['count']}` ta\n"
            f"   Jami: `{data['total_secs']}` sek\n"
            f"   Eng uzun: `{data['max_secs']}` sek\n"
        )
    
    lines.append("\n📁 To'liq log: `flood_log.txt`")
    await event.respond("\n".join(lines))


@bot.on(events.NewMessage(pattern='/flood_reset'))
async def flood_reset_cmd(event):
    if not await is_admin(event.sender_id):
        return
    engine._FLOOD_STATS.clear()
    # flood_log.txt ni ham tozalash
    log_path = os.path.join(BASE_DIR, 'flood_log.txt')
    if os.path.exists(log_path):
        open(log_path, 'w').close()
    await event.respond("✅ Flood statistikasi tozalandi!")


@bot.on(events.CallbackQuery(pattern=b"watch_rescan"))
async def watch_rescan_callback(event):
    if not await is_admin(event.sender_id):
        return
    await event.answer("🔄 Skanerlash boshlandi...")
    status = await bot.send_message(event.sender_id, "🔄 Bazadagi barcha fingerprintlar kuzatiladigan musiqalar bilan taqqoslanmoqda...")

    watches = await music.get_watch_list()
    if not watches:
        await status.edit("❌ Kuzatiladigan musiqalar yo\'q!")
        return

    total_found = 0
    async with aiosqlite.connect(music.MUSIC_DB, timeout=30) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS music_fingerprints (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                channel_id TEXT, channel_name TEXT, file_name TEXT,
                fingerprint TEXT, duration REAL, added_date TEXT
            )
        """)
        async with db.execute(
            "SELECT channel_id, channel_name, file_name, fingerprint FROM music_fingerprints"
        ) as cur:
            all_fps = await cur.fetchall()

    # Watch fingerprintlarini pre-parse qilish — har biri 1 marta parse (15 ta)
    watch_fps_parsed = []
    async with aiosqlite.connect(music.MUSIC_DB, timeout=30) as db:
        async with db.execute("SELECT id, fingerprint FROM watch_fingerprints") as cur:
            for (w_id, w_fp) in await cur.fetchall():
                watch_obj = next((w for w in watches if w[0] == w_id), None)
                w_name = watch_obj[1] if watch_obj else str(w_id)
                try:
                    # numpy mavjud bo'lsa numpy array, aks holda list
                    arr = music.parse_fingerprint(w_fp)
                    watch_fps_parsed.append((w_id, w_name, arr))
                except Exception:
                    pass

    await status.edit(
        f"🔍 `{len(all_fps)}` ta fingerprint `{len(watch_fps_parsed)}` ta musiqa bilan taqqoslanmoqda...\n"
        f"⚡️ Tezlashtirilgan batch rejim (bot muzlamaydi)"
    )

    # Barcha taqqoslashni thread'da bajarish — event loop bloklanmaydi
    _loop = asyncio.get_event_loop()
    matches = await _loop.run_in_executor(
        None,
        music.batch_compare_against_watches,
        watch_fps_parsed,
        all_fps
    )

    for ch_id, ch_name, fname, w_id, w_name, score_pct in matches:
        _src_type = 'profil' if (fname or '').startswith('profile_') else 'kanal'
        total_found += 1
        engine._WATCH_ALERTS.put_nowait({
            'admin_id':    event.sender_id,
            'watch_name':  w_name,
            'score':       score_pct,
            'source_name': ch_name or ch_id,
            'source_id':   ch_id,
            'source_type': _src_type
        })

    try:
        await status.edit(
            f"✅ **Skanerlash yakunlandi!**\n\n"
            f"🎵 Tekshirildi: `{len(all_fps)}` ta fingerprint\n"
            f"🚨 Topildi: `{total_found}` ta mos"
            + (f"\n\nXabarlar yuborilmoqda..." if total_found > 0 else "\n\nMos topilmadi.")
        )
    except Exception:
        await bot.send_message(
            event.sender_id,
            f"✅ **Skanerlash yakunlandi!**\n\n"
            f"🎵 Tekshirildi: `{len(all_fps)}` ta fingerprint\n"
            f"🚨 Topildi: `{total_found}` ta mos"
        )



@bot.on(events.NewMessage(pattern=r"/sync_cache"))
async def cmd_sync_cache(event):
    """Keshni qo'lda yangilash buyrug'i."""
    if not await is_admin(event.sender_id):
        return
    total, sources, last_sync = await engine.get_cache_stats()
    msg = await event.respond(
        f"🔄 **Kesh sinxronlash boshlandi...**\n"
        f"💾 Hozir: `{total:,}` xabar | `{sources}` manba\n"
        f"🕐 Oxirgi: {last_sync or 'hech qachon'}"
    )
    # Fon da ishga tushirish
    async def _do_sync():
        sources_set = set()
        async with aiosqlite.connect(db_mod.DB_NAME, timeout=30) as db:
            async with db.execute(
                "SELECT DISTINCT group_link FROM users_memory_bank "
                "WHERE group_link IS NOT NULL AND group_link != ''"
            ) as cur:
                for (link,) in await cur.fetchall():
                    sources_set.add(link)
            async with db.execute(
                "SELECT channel_id FROM hidden_channel_knocker WHERE status='joined'"
            ) as cur:
                for (ch_id,) in await cur.fetchall():
                    sources_set.add(ch_id)
        saved = 0
        for src in sources_set:
            try:
                saved += await engine.sync_source_messages(userbot, src, limit_days=90)
            except Exception:
                pass
            await asyncio.sleep(2)
        new_total, new_src, new_sync = await engine.get_cache_stats()
        try:
            await msg.edit(
                f"✅ **Kesh yangilandi!**\n"
                f"💾 Jami: `{new_total:,}` xabar | `{new_src}` manba\n"
                f"➕ Yangi saqlandi: `{saved}` ta\n"
                f"🕐 {new_sync}"
            )
        except Exception:
            pass
    asyncio.create_task(_do_sync())


async def run_id_search(sender_id: int, id_str: str, status_msg):
    """
    ID bo'yicha qidiruv:
    - Manfiy raqam → kanal ID → kanal nomi + linki
    - Musbat raqam → profil ID → u yozgan xabarlar + havolalar
    """
    try:
        num_id = int(id_str.strip())

        # ── KANAL ID (manfiy) ────────────────────────────────────────
        if num_id < 0:
            await status_msg.edit(f"🔍 Kanal `{num_id}` qidirilmoqda...")
            info, link = await engine.lookup_channel_by_id(userbot, num_id)
            if not info:
                await bot.send_message(
                    sender_id,
                    f"❌ `{num_id}` ID li kanal topilmadi.\n"
                    f"💡 Kanal bazada bo'lmasa yoki kirish imkoni yo'q bo'lsa topilmaydi."
                )
                return
            text = (
                f"📢 **Kanal ma'lumoti**\n\n"
                f"🆔 ID: `{num_id}`\n"
                f"📌 Nomi: **{info.get('title') or 'Nomalum'}**\n"
            )
            if info.get('username'):
                text += f"👤 Username: @{info['username']}\n"
            if info.get('members'):
                text += f"👥 A'zolar: `{info['members']:,}` ta\n"
            if link:
                text += f"\n🔗 **Havola:** {link}"
            await bot.send_message(sender_id, text)

        # ── PROFIL ID (musbat) ───────────────────────────────────────
        else:
            await status_msg.edit(f"🔍 Profil `{num_id}` qidirilmoqda...")
            profile, messages = await engine.lookup_user_by_id(num_id)

            if not profile and not messages:
                await bot.send_message(
                    sender_id,
                    f"❌ `{num_id}` ID li foydalanuvchi bazada topilmadi.\n"
                    f"💡 U hali skanerlangan guruhda bo'lmagan bo'lishi mumkin."
                )
                return

            # Profil bloki
            if profile:
                full_name = (profile['first_name'] + " " + profile['last_name']).strip() or "Nomsiz"
                un_part  = f"  @{profile['username']}" if profile['username'] else ""
                ph_part  = f"\n📱 Telefon: `{profile['phone']}`" if profile['phone'] else ""
                bio_part = f"\n📝 Bio: {profile['bio'][:150]}" if profile['bio'] else ""
                oc_part  = f"\n📢 Ochiq kanal: {profile['open_channels']}" if profile['open_channels'] and profile['open_channels'] not in ('', "Yo'q") else ""
                hh_part  = f"\n🔒 Maxfiy kanal: {profile['has_hidden']}" if profile['has_hidden'] and profile['has_hidden'] not in ('', '❌') else ""
                grp_list = "\n".join(f"  • {g}" for g in profile['groups'][:10]) if profile['groups'] else "  —"
                score, slabel, _ = engine.calculate_trust_score(profile, len(messages))
                profile_text = (
                    f"👤 **Profil ma'lumoti**\n"
                    f"🆔 ID: `{num_id}`\n"
                    f"📛 Ism: **{full_name}**{un_part}"
                    f"{ph_part}{bio_part}{oc_part}{hh_part}\n\n"
                    f"📋 **A'zo guruhlari ({len(profile['groups'])} ta):**\n{grp_list}\n\n"
                    f"💬 Xabarlar bazada: `{len(messages)}` ta\n"
                    f"🕐 Qo'shilgan: {profile['added_date']}\n"
                    f"🔐 **Trust Score: {score}/100 — {slabel}**\n"
                    f"{'─' * 30}"
                )
                # Profil rasmini yuborish
                photo_path = await engine.get_profile_photo(userbot, num_id)
                if photo_path and os.path.exists(photo_path):
                    try:
                        await bot.send_file(sender_id, photo_path, caption=profile_text)
                        os.remove(photo_path)
                    except Exception:
                        await bot.send_message(sender_id, profile_text)
                else:
                    await bot.send_message(sender_id, profile_text)

            if not messages:
                await bot.send_message(
                    sender_id,
                    "💬 Bu foydalanuvchi hech qaysi guruhda xabar yozmagan\n"
                    "(yoki xabarlar hali keshlanmagan)."
                )
                return

            # Xabarlar bloki
            chunk = f"💬 **Xabarlar ({len(messages)} ta):**\n\n"
            chunk_num = 1
            for r in messages:
                link_part = f"\n🔗 [O'tish]({r['link']})" if r.get('link') else ""
                line = (
                    f"📍 **{r['source_title']}**  📅 {r['date']}\n"
                    f"💬 {r['text']}"
                    f"{link_part}\n"
                    f"{'─' * 20}\n\n"
                )
                if len(chunk) + len(line) > 3800:
                    await bot.send_message(sender_id, chunk, link_preview=False)
                    await asyncio.sleep(0.4)
                    chunk = f"_(davomi {chunk_num + 1})_\n\n" + line
                    chunk_num += 1
                else:
                    chunk += line
            if chunk.strip():
                await bot.send_message(sender_id, chunk, link_preview=False)

    except Exception as e:
        await bot.send_message(sender_id, _scan_err(e))
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


# ═════════════════════════════════════════════════════════════════════
# 🔬 TERGOV VOSITALARI
# ═════════════════════════════════════════════════════════════════════

@bot.on(events.NewMessage(pattern=re.compile(r'^(/tergov|🔬 Tergov Vositalari)$')))
async def btn_tergov(event):
    if not await is_admin(event.sender_id):
        return
    await event.respond(
        "🔬 **Tergov Vositalari**\nQuyidagi bo'limdan kerakli amalni tanlang:",
        buttons=[
            # Profil tahlili
            [Button.inline("🎯 Trust Score",       data=b"tv_trust"),
             Button.inline("📸 Profil Rasmi",       data=b"tv_photo")],
            [Button.inline("📅 Faollik Tarixi",     data=b"tv_timeline"),
             Button.inline("📊 Akkaunt Hayoti",     data=b"tv_lifecycle")],
            [Button.inline("✍️ Yozuv Uslubi",       data=b"tv_style"),
             Button.inline("🔄 Uslub Solishtirish", data=b"tv_compare")],
            [Button.inline("📝 O'zgarishlar",       data=b"tv_changes"),
             Button.inline("📄 To'liq Hisobot",     data=b"tv_evidence")],
            # Qidiruv
            [Button.inline("🔍 Username Qidiruv",   data=b"tv_lookup"),
             Button.inline("📞 Telefon Qidiruv",    data=b"tv_phone")],
            # Tahlil
            [Button.inline("👥 Umumiy A'zolar",     data=b"tv_common"),
             Button.inline("🤖 Koordinatsiya",       data=b"tv_coordinated")],
            [Button.inline("⏰ Bir Vaqtda Faol",    data=b"tv_temporal"),
             Button.inline("🗑 O'chirilgan Xabar",  data=b"tv_deleted")],
            [Button.inline("🕸 Tarmoq Xaritasi",    data=b"tv_network")],
            # Tergov ishi
            [Button.inline("➕ Yangi Tergov",       data=b"tv_case_new"),
             Button.inline("📋 Tergovlar Ro'yxati", data=b"tv_case_list")],
            [Button.inline("👤 Shaxs Qo'shish",     data=b"tv_case_add"),
             Button.inline("📄 Tergov Hisoboti",    data=b"tv_case_report")],
            [Button.inline("🗑 Tergovni O'chirish",  data=b"tv_case_del")],
        ]
    )


# ── Tergov vositalari: har bir tugma callback ─────────────────────────

_TV_PROMPTS = {
    "tv_trust":        ("🎯 Trust Score", "Tekshiriladigan shaxs ID sini yuboring:\n`123456789`"),
    "tv_photo":        ("📸 Profil Rasmi", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_timeline":     ("📅 Faollik Tarixi", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_lifecycle":    ("📊 Akkaunt Hayoti", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_style":        ("✍️ Yozuv Uslubi", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_compare":      ("🔄 Uslub Solishtirish", "Ikkita ID ni bo'sh joy bilan yuboring:\n`ID1 ID2`"),
    "tv_changes":      ("📝 O'zgarishlar", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_evidence":     ("📄 To'liq Hisobot", "Shaxs ID sini yuboring:\n`123456789`"),
    "tv_lookup":       ("🔍 Username Qidiruv", "Username yuboring:\n`@username`"),
    "tv_phone":        ("📞 Telefon Qidiruv", "Telefon raqam yuboring:\n`+998901234567`"),
    "tv_common":       ("👥 Umumiy A'zolar", "Ikkita guruh havolasini yuboring:\n`@guruh1 @guruh2`"),
    "tv_coordinated":  ("🤖 Koordinatsiya", "Guruh havolasini yuboring:\n`@guruh`"),
    "tv_deleted":      ("🗑 O'chirilgan Xabar", "Guruh havolasini yuboring:\n`@guruh`"),
    "tv_network":      ("🕸 Tarmoq Xaritasi", "IDlarni vergul bilan yuboring:\n`ID1,ID2,ID3`"),
    "tv_case_new":     ("➕ Yangi Tergov", "Tergov nomini yuboring:\n`Tergov nomi`"),
    "tv_case_add":     ("👤 Shaxs Qo'shish", "Tergov ID va shaxsni yuboring:\n`tergov_ID @username_yoki_ID`"),
    "tv_case_report":  ("📄 Tergov Hisoboti", "Tergov ID sini yuboring:\n`1`"),
    "tv_case_del":     ("🗑 Tergovni O'chirish", "O'chiriladigan tergov ID sini yuboring:\n`1`"),
}

@bot.on(events.CallbackQuery(pattern=b"tv_.*"))
async def tv_callback(event):
    if not await is_admin(event.sender_id):
        return
    action = event.data.decode()

    # tv_temporal — input kerak emas, to'g'ridan-to'g'ri ishga tushirish
    if action == "tv_temporal":
        await event.answer()
        status = await event.respond("⏳ Bir vaqtda faol akkauntlar hisoblanmoqda...")
        pairs = await engine.get_temporal_correlations()
        if not pairs:
            await status.edit("📭 Korrelyatsiya topilmadi.")
            return
        lines = ["⏰ **Bir vaqtda faol akkauntlar:**\n"]
        for uid1, uid2, overlap in pairs[:20]:
            lines.append(f"👤 `{uid1}` ↔ `{uid2}` — {overlap} ta mos soat")
        await status.edit("\n".join(lines))
        return

    # tv_case_list — input kerak emas
    if action == "tv_case_list":
        await event.answer()
        cases = await db_mod.get_investigations()
        if not cases:
            await event.respond("📁 Hozircha hech qanday tergov yo'q.\n➕ Yangi tergov ochish uchun tugmani bosing.")
            return
        lines = ["📁 **Tergovlar ro'yxati:**\n"]
        for c in cases:
            lines.append(f"`{c[0]}`. **{c[1]}** — {c[4][:10]}")
        await event.respond("\n".join(lines))
        return

    # Qolgan amallar — input kutish
    if action not in _TV_PROMPTS:
        await event.answer("Noma'lum amal", alert=True)
        return

    title, prompt = _TV_PROMPTS[action]
    USER_STATES[event.sender_id] = {'state': action}
    await event.answer()
    await event.respond(f"**{title}**\n\n{prompt}")


@bot.on(events.NewMessage())
async def tv_input_handler(event):
    """Tergov vositalari uchun input qabul qilish."""
    if not await is_admin(event.sender_id):
        return
    state = USER_STATES.get(event.sender_id)
    if not isinstance(state, dict):
        return
    action = state.get('state', '')
    if not action.startswith('tv_'):
        return

    text = (event.message.text or '').strip()
    if not text:
        return

    USER_STATES[event.sender_id] = None

    # Har bir amal uchun mavjud handler funksiyalarini chaqirish
    fake = event  # event o'zini ishlatamiz
    parts = text.split()

    if action == 'tv_trust':
        fake.message.text = f"/trust {text}"
        await cmd_trust(fake)

    elif action == 'tv_photo':
        fake.message.text = f"/photo {text}"
        await cmd_photo(fake)

    elif action == 'tv_timeline':
        fake.message.text = f"/timeline {text}"
        await cmd_timeline(fake)

    elif action == 'tv_lifecycle':
        fake.message.text = f"/lifecycle {text}"
        await cmd_lifecycle(fake)

    elif action == 'tv_style':
        fake.message.text = f"/style {text}"
        await cmd_style(fake)

    elif action == 'tv_compare':
        fake.message.text = f"/compare {text}"
        await cmd_compare_style(fake)

    elif action == 'tv_changes':
        fake.message.text = f"/changes {text}"
        await cmd_changes(fake)

    elif action == 'tv_evidence':
        fake.message.text = f"/evidence {text}"
        await cmd_evidence(fake)

    elif action == 'tv_lookup':
        fake.message.text = f"/lookup {text}"
        await cmd_lookup(fake)

    elif action == 'tv_phone':
        fake.message.text = f"/phone {text}"
        await cmd_phone(fake)

    elif action == 'tv_common':
        fake.message.text = f"/common {text}"
        await cmd_common(fake)

    elif action == 'tv_coordinated':
        fake.message.text = f"/coordinated {text}"
        await cmd_coordinated(fake)

    elif action == 'tv_deleted':
        fake.message.text = f"/deleted {text}"
        await cmd_deleted(fake)

    elif action == 'tv_network':
        fake.message.text = f"/network {text}"
        await cmd_network(fake)

    elif action == 'tv_case_new':
        fake.message.text = f"/case new {text}"
        await cmd_case_new(fake)

    elif action == 'tv_case_add':
        fake.message.text = f"/case add {text}"
        await cmd_case_add(fake)

    elif action == 'tv_case_report':
        fake.message.text = f"/case report {text}"
        await cmd_case_report(fake)

    elif action == 'tv_case_del':
        fake.message.text = f"/case del {text}"
        await cmd_case_del(fake)


# ─────────────────────────────────────────────────────────────────────
# 🚨 ALERT BOSHQARUVI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=re.compile(r'^(/alert|🚨 Alert Boshqaruvi)$')))
async def btn_alert_mgmt(event):
    if not await is_admin(event.sender_id):
        return
    alerts = await db_mod.list_alerts(event.sender_id)
    if not alerts:
        text = (
            "🚨 **Alert Boshqaruvi**\n\n"
            "Hozircha alert yo'q.\n\n"
            "Qo'shish: `/alert add <kalit_soz>`\n"
            "O'chirish: `/alert del <ID>`\n"
            "Ro'yxat: `/alert list`"
        )
    else:
        lines = ["🚨 **Alertlar ro'yxati:**\n"]
        for a in alerts:
            status = "✅" if a[3] else "⏸"
            grp = f" | {a[2]}" if a[2] else ""
            lines.append(f"{status} `{a[0]}` — **{a[1]}**{grp}")
        lines.append("\n`/alert add <kalit_soz>` — yangi qo'shish")
        lines.append("`/alert del <ID>` — o'chirish")
        text = "\n".join(lines)
    await event.respond(text)


@bot.on(events.NewMessage(pattern=r'/alert add (.+)'))
async def cmd_alert_add(event):
    if not await is_admin(event.sender_id):
        return
    keyword = _pm(event, 1).strip()
    aid = await db_mod.add_alert(event.sender_id, keyword)
    await event.respond(f"✅ Alert qo'shildi!\n🆔 ID: `{aid}`\n🔑 Kalit so'z: `{keyword}`")

@bot.on(events.NewMessage(pattern=r'/alert del (\d+)'))
async def cmd_alert_del(event):
    if not await is_admin(event.sender_id):
        return
    aid = int(_pm(event, 1))
    await db_mod.delete_alert(aid)
    await event.respond(f"❌ Alert `{aid}` o'chirildi.")

@bot.on(events.NewMessage(pattern='/alert list'))
async def cmd_alert_list(event):
    if not await is_admin(event.sender_id):
        return
    alerts = await db_mod.list_alerts(event.sender_id)
    if not alerts:
        await event.respond("Alertlar yo'q. `/alert add <so'z>` bilan qo'shing.")
        return
    lines = ["🚨 **Alertlaringiz:**\n"]
    for a in alerts:
        st = "✅ faol" if a[3] else "⏸ to'xtatilgan"
        lines.append(f"`{a[0]}` | **{a[1]}** | {st}")
    await event.respond("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #7 TRUST SCORE
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/trust (-?\d+)'))
async def cmd_trust(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    profile, messages = await engine.lookup_user_by_id(uid)
    if not profile and not messages:
        await event.respond(f"❌ `{uid}` ID bazada topilmadi.")
        return
    score, label, reasons = engine.calculate_trust_score(profile or {}, len(messages))
    bar = "█" * (score // 10) + "░" * (10 - score // 10)
    text = (
        f"🔐 **Trust Score: {uid}**\n\n"
        f"[{bar}] **{score}/100**\n"
        f"{label}\n\n"
        f"📊 **Ballar:**\n" +
        "\n".join(f"  {r}" for r in reasons)
    )
    await event.respond(text)


# ─────────────────────────────────────────────────────────────────────
# #8 O'ZGARISHLAR TARIXI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/changes (-?\d+)'))
async def cmd_changes(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    changes = await db_mod.get_user_change_log(uid)
    if not changes:
        await event.respond(f"📝 `{uid}` ID uchun o'zgarishlar yo'q (yoki hali kuzatilmagan).")
        return
    lines = [f"📝 **O'zgarishlar tarixi — ID {uid}** ({len(changes)} ta):\n"]
    for ch in changes[:30]:
        field_names = {'first_name': 'Ism', 'last_name': 'Familiya',
                       'username': 'Username', 'phone': 'Telefon', 'bio': 'Bio'}
        fname = field_names.get(ch[0], ch[0])
        lines.append(f"🕐 `{ch[3]}`")
        lines.append(f"   {fname}: `{ch[1] or '—'}` → `{ch[2] or '—'}`\n")
    await event.respond("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #10 USERNAME / TELEFON QIDIRUV
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/lookup (.+)'))
async def cmd_lookup(event):
    if not await is_admin(event.sender_id):
        return
    query = _pm(event, 1).strip()
    msg = await event.respond("🔍 Qidirilmoqda...")
    if query.startswith('@'):
        results = await engine.search_by_username(query)
        if not results:
            await msg.edit(f"❌ `{query}` username topilmadi.")
            return
        lines = [f"👤 **Username qidiruvi: {query}**\n"]
        for r in results[:5]:
            full = (r['first_name'] + ' ' + r['last_name']).strip() or 'Nomsiz'
            ph = f" | 📱{r['phone']}" if r['phone'] else ""
            lines.append(f"🆔 `{r['user_id']}` — **{full}** @{r['username']}{ph}")
            lines.append(f"   Guruhlar: {', '.join(r['groups'][:3])}")
        await msg.edit("\n".join(lines))
    elif query.startswith('+') or query[0].isdigit():
        result = await engine.search_by_phone(userbot, query)
        if not result:
            await msg.edit(f"❌ `{query}` telefon raqami topilmadi.")
            return
        full = (result['first_name'] + ' ' + result['last_name']).strip() or 'Nomsiz'
        src = result.get('source', '')
        lines = [
            f"📱 **Telefon qidiruvi: {query}**\n",
            f"🆔 ID: `{result['user_id']}`",
            f"📛 Ism: **{full}**",
            f"👤 Username: @{result['username']}" if result['username'] else "",
            f"📲 Manba: {src}",
        ]
        if result.get('groups'):
            lines.append(f"📋 Guruhlar: {', '.join(result['groups'][:3])}")
        await msg.edit("\n".join(l for l in lines if l))
    else:
        await msg.edit("⚠️ Format: `/lookup @username` yoki `/lookup +998901234567`")


# ─────────────────────────────────────────────────────────────────────
# #24 PROFIL RASMI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/photo (-?\d+)'))
async def cmd_photo(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    msg = await event.respond("📸 Rasm yuklanmoqda...")
    path = await engine.get_profile_photo(userbot, uid)
    if path and os.path.exists(path):
        await bot.send_file(event.sender_id, path, caption=f"📸 Profil rasmi — ID `{uid}`")
        os.remove(path)
        await msg.delete()
    else:
        await msg.edit(f"🔒 `{uid}` ID profil rasmi ko'rinmaydi (yopiq yoki yo'q).")


# ─────────────────────────────────────────────────────────────────────
# #25 TELEFON → PROFIL (alohida buyruq)
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/phone (.+)'))
async def cmd_phone(event):
    if not await is_admin(event.sender_id):
        return
    phone = _pm(event, 1).strip()
    msg = await event.respond("📱 Telegram da qidirilmoqda...")
    result = await engine.search_by_phone(userbot, phone)
    if not result:
        await msg.edit(f"❌ `{phone}` — Telegram da topilmadi.")
        return
    full = (result['first_name'] + ' ' + result['last_name']).strip() or 'Nomsiz'
    src_txt = "✅ Telegramdan topildi" if result.get('source') == 'telegram' else "✅ Bazadan topildi"
    text = (
        f"📱 **Telefon: {phone}**\n"
        f"{src_txt}\n\n"
        f"🆔 ID: `{result['user_id']}`\n"
        f"📛 Ism: **{full}**\n"
    )
    if result['username']:
        text += f"👤 Username: @{result['username']}\n"
    if result.get('groups'):
        text += f"📋 Guruhlar: {', '.join(result['groups'][:3])}\n"
    text += f"\n💡 `/trust {result['user_id']}` — trust score ko'rish"
    await msg.edit(text)


# ─────────────────────────────────────────────────────────────────────
# #12 FAOLLIK TIMELINE
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/timeline (-?\d+)'))
async def cmd_timeline(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    data = await engine.get_user_timeline(uid, days=90)
    if not data['timeline']:
        await event.respond(f"📅 `{uid}` ID uchun kesh da xabar yo'q.")
        return
    lines = [f"📅 **Faollik — ID {uid}** (so'nggi 90 kun)\n"]
    lines.append(f"Jami: `{data['total']}` xabar")
    lines.append(f"Birinchi: {data['first_msg']}")
    lines.append(f"Oxirgi:   {data['last_msg']}\n")
    lines.append("**Kunlik faollik:**")
    for d in data['timeline'][-20:]:
        bar = "▓" * min(d['count'], 20)
        lines.append(f"`{d['date']}` {bar} ({d['count']})")
    await event.respond("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #20 AKKAUNT HAYOTI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/lifecycle (-?\d+)'))
async def cmd_lifecycle(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    data = await engine.get_account_lifecycle(uid)
    if not data['monthly']:
        await event.respond(f"📊 `{uid}` ID uchun faollik ma'lumoti yo'q.")
        return
    lines = [f"📊 **Akkaunt hayoti — ID {uid}**\n"]
    lines.append(f"🕐 Birinchi ko'rilgan: {data['first_seen']}")
    lines.append(f"📝 O'zgarishlar soni: {data['change_count']}\n")
    if data['peak_month']:
        lines.append(f"📈 Eng faol oy: `{data['peak_month']}`")
    if data['peak_hour'] is not None:
        lines.append(f"🕐 Eng faol soat: `{data['peak_hour']}:00`\n")
    lines.append("**Oylik faollik:**")
    for m in data['monthly']:
        bar = "█" * min(m['count'] // 2, 25)
        lines.append(f"`{m['month']}` {bar} ({m['count']})")
    await event.respond("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #17 YOZUV USLUBI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/style (-?\d+)'))
async def cmd_style(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    style = await engine.analyze_writing_style(uid)
    if not style:
        await event.respond(f"✍️ `{uid}` ID uchun tahlil qilish uchun xabarlar yetarli emas.")
        return
    top = ", ".join(f"`{w}`({c})" for w, c in style['top_words'][:8])
    text = (
        f"✍️ **Yozuv uslubi — ID {uid}**\n\n"
        f"📨 Tahlil qilingan: `{style['msg_count']}` xabar\n"
        f"📏 O'rtacha uzunlik: `{style['avg_len']}` belgi / `{style['avg_words']}` so'z\n"
        f"😀 Emoji/xabar: `{style['emoji_ratio']}`\n"
        f"❓ Savol (%): `{style['question_pct']}` | ❗ Undov (%): `{style['exclaim_pct']}`\n"
        f"🔠 Katta harf (%): `{style['caps_ratio']}`\n\n"
        f"📊 **Top so'zlar:** {top}"
    )
    await event.respond(text)

@bot.on(events.NewMessage(pattern=r'/compare (-?\d+) (-?\d+)'))
async def cmd_compare_style(event):
    if not await is_admin(event.sender_id):
        return
    uid1 = int(_pm(event, 1))
    uid2 = int(_pm(event, 2))
    msg = await event.respond("✍️ Uslublar tahlil qilinmoqda...")
    similarity = await engine.compare_writing_styles(uid1, uid2)
    bar = "█" * (similarity // 10) + "░" * (10 - similarity // 10)
    if similarity >= 70:
        verdict = "⚠️ **Bir odam bo'lishi mumkin!**"
    elif similarity >= 50:
        verdict = "🟡 O'rtacha o'xshashlik"
    else:
        verdict = "🟢 Turli uslublar"
    text = (
        f"✍️ **Uslub solishtirish**\n"
        f"ID {uid1} vs ID {uid2}\n\n"
        f"[{bar}] **{similarity}%** o'xshashlik\n\n"
        f"{verdict}"
    )
    await msg.edit(text)


# ─────────────────────────────────────────────────────────────────────
# #15 CROSS-GROUP TAHLIL
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/common (.+?) (.+)'))
async def cmd_common(event):
    if not await is_admin(event.sender_id):
        return
    g1 = _pm(event, 1).strip()
    g2 = _pm(event, 2).strip()
    msg = await event.respond("🔍 Umumiy a'zolar qidirilmoqda...")
    members = await engine.get_common_members(g1, g2)
    if not members:
        await msg.edit(f"❌ `{g1}` va `{g2}` da umumiy a'zo topilmadi.")
        return
    lines = [f"👥 **Umumiy a'zolar: {len(members)} ta**\n{g1} ∩ {g2}\n"]
    for m in members[:30]:
        full = (m['first_name'] + ' ' + m['last_name']).strip() or 'Nomsiz'
        un = f" @{m['username']}" if m['username'] else ""
        ph = f" | {m['phone']}" if m['phone'] else ""
        lines.append(f"🆔 `{m['user_id']}` {full}{un}{ph}")
    if len(members) > 30:
        lines.append(f"\n... va yana {len(members)-30} ta")
    await msg.edit("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #16 KOORDINATSIYALANGAN XATTI-HARAKAT
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/coordinated (.+)'))
async def cmd_coordinated(event):
    if not await is_admin(event.sender_id):
        return
    group = _pm(event, 1).strip()
    msg = await event.respond("🔍 Tahlil qilinmoqda...")
    data = await engine.detect_coordinated_behavior(group)
    risk = data['risk_score']
    if risk >= 70:
        risk_label = "🔴 YUQORI XAVF"
    elif risk >= 40:
        risk_label = "🟡 O'RTA XAVF"
    else:
        risk_label = "🟢 PAST XAVF"
    text = (
        f"🕵️ **Koordinatsiya tahlili**\n{group}\n\n"
        f"👥 Jami a'zolar: `{data['total']}`\n"
        f"📵 Bio yo'q: `{data['no_bio_pct']}%`\n"
        f"📵 Telefon yo'q: `{data['no_phone_pct']}%`\n"
        f"🤖 Bir vaqtda qo'shilgan guruhlari: `{data['cluster_count']}`\n\n"
        f"⚠️ **Xavf darajasi: {risk}/100 — {risk_label}**"
    )
    if data['clusters']:
        text += f"\n\n**Shubhali klasterlar (birinchisi):**\n"
        for m in data['clusters'][0][:5]:
            full = (m[1] or '') + (' @' + m[2] if m[2] else '')
            text += f"  🆔 `{m[0]}` {full} — {m[5]}\n"
    await msg.edit(text)


# ─────────────────────────────────────────────────────────────────────
# #18 VAQT KORRELYATSIYASI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern='/temporal'))
async def cmd_temporal(event):
    if not await is_admin(event.sender_id):
        return
    msg = await event.respond("⏰ Vaqt korrelyatsiyasi hisoblanmoqda...")
    pairs = await engine.get_temporal_correlations(min_overlap=5)
    if not pairs:
        await msg.edit("⏰ Korrelyatsiya topilmadi (keshda yetarli xabar yo'q).")
        return
    lines = [f"⏰ **Bir vaqtda faol akkauntlar** (top {len(pairs)} juft):\n"]
    for uid1, uid2, overlap in pairs[:15]:
        lines.append(f"👤 `{uid1}` ↔ `{uid2}` — `{overlap}` umumiy soat")
        lines.append(f"   ⚠️ Bir odam bo'lishi mumkin — `/compare {uid1} {uid2}`")
    await msg.edit("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #19 O'CHIRILGAN XABARLAR
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/deleted (.+)'))
async def cmd_deleted(event):
    if not await is_admin(event.sender_id):
        return
    source = _pm(event, 1).strip()
    msgs = await engine.get_deleted_messages(source)
    if not msgs:
        await event.respond(f"🗑️ `{source}` da o'chirilgan xabar yo'q (yoki kuzatilmagan).")
        return
    lines = [f"🗑️ **O'chirilgan xabarlar — {source}** ({len(msgs)} ta):\n"]
    for m in msgs[:20]:
        un = f"@{m['username']}" if m['username'] else f"ID:{m['sender_id']}"
        link_part = f"\n   🔗 {m['link']}" if m['link'] else ""
        lines.append(f"📅 `{m['date']}` | {un}")
        lines.append(f"   💬 {m['text'][:150]}{link_part}\n")
    await event.respond("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────
# #14 EVIDENCE PAKETI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/evidence (-?\d+)'))
async def cmd_evidence(event):
    if not await is_admin(event.sender_id):
        return
    uid = int(_pm(event, 1))
    msg = await event.respond(f"📄 `{uid}` uchun evidence hisoboti tayyorlanmoqda...")
    path = await engine.generate_evidence_report(uid)
    if path and os.path.exists(path):
        await bot.send_file(
            event.sender_id, path,
            caption=f"📄 **Evidence hisoboti — ID {uid}**\n{datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        os.remove(path)
        await msg.delete()
    else:
        await msg.edit(f"❌ `{uid}` ID bazada topilmadi.")


# ─────────────────────────────────────────────────────────────────────
# #23 TARMOQ XARITASI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/network (.+)'))
async def cmd_network(event):
    if not await is_admin(event.sender_id):
        return
    ids_raw = _pm(event, 1).strip()
    try:
        user_ids = [int(x.strip()) for x in ids_raw.split(',') if x.strip().lstrip('-').isdigit()]
    except Exception:
        await event.respond("⚠️ Format: `/network 123,456,789`")
        return
    if not user_ids:
        await event.respond("⚠️ ID lar vergul bilan ajrating: `/network 123,456,789`")
        return
    msg = await event.respond("🕸️ Tarmoq xaritasi yaratilmoqda...")
    path = await engine.generate_network_map(user_ids)
    if path and os.path.exists(path):
        await bot.send_file(
            event.sender_id, path,
            caption=f"🕸️ **Tarmoq xaritasi** — {len(user_ids)} ta shaxs\nBrauzerda oching"
        )
        os.remove(path)
        await msg.delete()
    else:
        await msg.edit("❌ Xarita yaratishda xatolik.")


# ─────────────────────────────────────────────────────────────────────
# #22 TERGOVCHI ISH MAYDONI
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern=r'/case new (.+)'))
async def cmd_case_new(event):
    if not await is_admin(event.sender_id):
        return
    name = _pm(event, 1).strip()
    inv_id = await db_mod.create_investigation(name, event.sender_id)
    await event.respond(
        f"📁 **Yangi tergov yaratildi!**\n"
        f"🆔 ID: `{inv_id}`\n"
        f"📌 Nom: {name}\n\n"
        f"Shaxs qo'shish: `/case add {inv_id} <user_id>`\n"
        f"Hisobot: `/case report {inv_id}`"
    )

@bot.on(events.NewMessage(pattern='/case list'))
async def cmd_case_list(event):
    if not await is_admin(event.sender_id):
        return
    invs = await db_mod.get_investigations(event.sender_id)
    if not invs:
        await event.respond("📁 Tergovlar yo'q. `/case new <nom>` bilan yarating.")
        return
    lines = ["📁 **Tergovlar ro'yxati:**\n"]
    for inv in invs:
        targets = await db_mod.get_investigation_targets(inv[0])
        lines.append(f"🆔 `{inv[0]}` — **{inv[1]}** ({len(targets)} shaxs) | {inv[4][:10]}")
    lines.append("\n`/case report <ID>` — hisobot olish")
    await event.respond("\n".join(lines))

@bot.on(events.NewMessage(pattern=r'/case add (\d+) (.+)'))
async def cmd_case_add(event):
    if not await is_admin(event.sender_id):
        return
    inv_id     = int(_pm(event, 1))
    target_val = _pm(event, 2).strip()
    ttype = 'channel' if target_val.lstrip('-').isdigit() and int(target_val) < 0 else 'user'
    await db_mod.add_investigation_target(inv_id, ttype, target_val)
    await event.respond(f"✅ `{target_val}` tergov `{inv_id}` ga qo'shildi.")

@bot.on(events.NewMessage(pattern=r'/case report (\d+)'))
async def cmd_case_report(event):
    if not await is_admin(event.sender_id):
        return
    inv_id = int(_pm(event, 1))
    msg = await event.respond("📄 Tergov hisoboti tayyorlanmoqda...")
    path = await engine.generate_investigation_report(inv_id)
    if path and os.path.exists(path):
        await bot.send_file(
            event.sender_id, path,
            caption=f"📄 **Tergov hisoboti #{inv_id}**"
        )
        os.remove(path)
        await msg.delete()
    else:
        await msg.edit(f"❌ Tergov `{inv_id}` topilmadi.")

@bot.on(events.NewMessage(pattern=r'/case del (\d+)'))
async def cmd_case_del(event):
    if not await is_admin(event.sender_id):
        return
    inv_id = int(_pm(event, 1))
    await db_mod.delete_investigation(inv_id)
    await event.respond(f"🗑️ Tergov `{inv_id}` o'chirildi.")


# ─────────────────────────────────────────────────────────────────────
# ALERT SENDER — yangi kesh xabarlarda alert tekshiruvi
# ─────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────
# 🕵️ TERGOV MA'LUMOTI — TO'LIQ PDF
# ─────────────────────────────────────────────────────────────────────

@bot.on(events.NewMessage(pattern="💬 Kamentariya Xisoboti"))
async def btn_tergov_malumot(event):
    if not await is_admin(event.sender_id):
        return
    USER_STATES[event.sender_id] = 'waiting_tergov_id'
    await event.respond(
        "💬 **Kamentariya Xisoboti — To'liq PDF Hisobot**\n\n"
        "Quyidagilardan birini kiriting:\n"
        "📱 **Telefon:**  `+998901234567`\n"
        "👤 **Username:** `@username`\n"
        "🆔 **Profil ID:** `123456789`\n\n"
        "📄 Bot barcha ma'lumotlarni to'plab PDF fayl yuboradi."
    )


async def run_tergov_pdf(sender_id: int, identifier: str, status_msg):
    try:
        result = await engine.generate_tergov_pdf(userbot, identifier)
        if result is None or (isinstance(result, tuple) and result[0] is None):
            err = result[1] if result else "Noma'lum xato"
            await bot.send_message(sender_id, f"❌ Topilmadi: {err}")
            return
        pdf_path, label = result
        if not pdf_path or not os.path.exists(pdf_path):
            await bot.send_message(sender_id, "❌ PDF yaratishda xatolik yuz berdi.")
            return
        size_kb = round(os.path.getsize(pdf_path) / 1024)
        await bot.send_file(
            sender_id, pdf_path,
            caption=(
                f"💬 **Kamentariya Xisoboti**\n"
                f"📋 {label}\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                f"📄 Hajmi: {size_kb} KB"
            )
        )
        os.remove(pdf_path)
    except Exception as e:
        await bot.send_message(sender_id, f"❌ PDF xatolik: {e}")
    finally:
        try:
            await status_msg.delete()
        except Exception:
            pass


async def process_alert_hits(hits: list):
    for hit in hits:
        try:
            un = f"@{hit['sender_name']}" if hit['sender_name'] else f"ID:{hit['sender_id']}"
            link_part = f"\n🔗 {hit['link']}" if hit['link'] else ""
            text = (
                f"🚨 **ALERT ISHLADI!**\n"
                f"🔑 Kalit so'z: `{hit['keyword']}`\n"
                f"📍 Manba: {hit['source']}\n"
                f"👤 Kim: {un}\n"
                f"📅 Sana: {hit['date']}\n\n"
                f"💬 {hit['text'][:300]}"
                f"{link_part}"
            )
            await bot.send_message(hit['admin_id'], text, link_preview=False)
        except Exception:
            pass


async def main():
    # asyncio.Queue lar event loop ichida yaratilishi kerak
    global _SCAN_QUEUE
    _SCAN_QUEUE = asyncio.Queue()
    engine._WATCH_ALERTS = asyncio.Queue()

    await db_mod.init_db()
    await music.init_music_db()
    await userbot.start()
    await bot.start(bot_token=BOT_TOKEN)

    # Elektr uzilishi qolgan vaqtinchalik fayllarni tozalash
    import glob as _glob
    for _tmp in _glob.glob(os.path.join(BASE_DIR, "tmp_*.ogg")):
        try:
            os.remove(_tmp)
        except Exception:
            pass

    asyncio.create_task(engine.smart_channel_knocker(userbot, bot, SUPER_ADMIN_ID))
    asyncio.create_task(engine.background_profile_tracker(userbot))
    asyncio.create_task(engine.music_channel_tracker(userbot))
    asyncio.create_task(watch_alert_sender())
    asyncio.create_task(scan_queue_runner())
    print("✅ Kiber-Stansiya OSINT Pro ishga tushdi!")
    await bot.run_until_disconnected()

if __name__ == '__main__':
    asyncio.run(main())
